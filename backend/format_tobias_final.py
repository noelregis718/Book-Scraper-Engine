import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def format_tobias_final(input_file, output_file):
    # Read the data, skipping the first two title/source rows
    df_raw = pd.read_excel(input_file, skiprows=2, header=0)
    df_raw.columns = ["#", "Book Title", "Author", "Genre", "Agent", "Publisher", "Notes"]

    # Drop empty/header rows
    df_raw = df_raw.dropna(subset=["Book Title"])
    df_raw = df_raw[df_raw["Book Title"].astype(str).str.strip() != ""]
    df_raw = df_raw[~df_raw["#"].astype(str).str.strip().str.lower().isin(["#", "nan", ""])]

    # Build new 11-column dataframe
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    new_df = pd.DataFrame(columns=new_columns, index=range(len(df_raw)))

    new_df["Name of Series"] = df_raw["Book Title"].values
    new_df["Author Name"] = df_raw["Author"].values

    # Publisher: use actual if not dash/blank
    def clean_publisher(x):
        if pd.isna(x) or str(x).strip() in ["—", "", "-"]:
            return "Tobias Literary Agency"
        return str(x).strip()

    # Agent: use actual if not TLA/dash/blank
    def clean_agent(x):
        if pd.isna(x) or str(x).strip() in ["TLA", "—", "", "-"]:
            return "Tobias Literary Agency"
        return str(x).strip()

    new_df["Publisher"] = df_raw["Publisher"].apply(clean_publisher).values
    new_df["Name of agent"] = df_raw["Agent"].apply(clean_agent).values

    # Romantasy detection from genre
    def check_romantasy(genre):
        if not isinstance(genre, str): return "No"
        g = genre.lower()
        keywords = ["fantasy", "paranormal", "romantasy", "dragon", "magic",
                    "supernatural", "fae", "witch", "vampire", "shifter",
                    "werewolf", "gothic", "mytholog"]
        return "Yes" if any(k in g for k in keywords) else "No"

    new_df["Romantasy = Yes or No?"] = df_raw["Genre"].apply(check_romantasy).values
    new_df["Romantasy Sub-Genre of series"] = df_raw["Genre"].values

    # Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Tobias Final"

    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, len(new_columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(new_columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_file)
    print(f"Done! Saved to {output_file}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file = os.path.join(base, "Tobias_All_Books_FINAL.xlsx")
    output_file = os.path.join(base, "Tobias_All_Books_FINAL.xlsx")
    format_tobias_final(input_file, output_file)
