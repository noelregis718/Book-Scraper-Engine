import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def format_azantian():
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file = os.path.join(base, "Azantian_LitAgency_FULL.xlsx")
    output_file = os.path.join(base, "Azantian_LitAgency_Combined_Formatted.xlsx")

    if not os.path.exists(input_file):
        print(f"Error: {input_file} not found!")
        return

    # Read both sheets
    df_adult = pd.read_excel(input_file, sheet_name="Adult Books", skiprows=2)
    df_ya = pd.read_excel(input_file, sheet_name="Young Adult Books", skiprows=2)

    # Combine
    df_combined = pd.concat([df_adult, df_ya], ignore_index=True)

    # Drop fully empty rows or where Book Title is empty (but don't delete actual books)
    df_combined = df_combined.dropna(subset=["Book Title"])
    df_combined = df_combined[df_combined["Book Title"].astype(str).str.strip() != ""]

    # 11-column format
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    new_df = pd.DataFrame(columns=new_columns, index=range(len(df_combined)))

    new_df["Name of Series"] = df_combined["Book Title"].values
    if "Author Name" in df_combined.columns:
        new_df["Author Name"] = df_combined["Author Name"].values
    
    new_df["Publisher"] = ""
    new_df["Name of agent"] = "Azantian Literary Agency"

    # Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Books"

    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, len(new_columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(new_columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_file)
    print(f"Done! Saved to {output_file}")
    
    import subprocess
    subprocess.Popen(["start", output_file], shell=True)

if __name__ == "__main__":
    format_azantian()
