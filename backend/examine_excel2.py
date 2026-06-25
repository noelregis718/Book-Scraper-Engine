import openpyxl

def main():
    wb = openpyxl.load_workbook('LDLA_Combined_Mapped_Fixed.xlsx')
    ws = wb.active
    
    print('Checking rows where col 12 or 13 has data:')
    for r in range(2, ws.max_row + 1):
        col9 = ws.cell(r, 9).value
        col10 = ws.cell(r, 10).value
        col11 = ws.cell(r, 11).value
        col12 = ws.cell(r, 12).value
        col13 = ws.cell(r, 13).value
        
        if col12 is not None or col13 is not None:
            print(f"Row {r}: 9={col9}, 10={col10}, 11={col11}, 12={col12}, 13={col13}")

if __name__ == "__main__":
    main()
