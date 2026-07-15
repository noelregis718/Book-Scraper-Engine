import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

file_path = r"E:\Internship\PocketFM\New_Romantasy_Books.xlsx"

# Apply premium JRA styling
wb = load_workbook(file_path)
ws = wb.active

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
header_font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")

data_font = Font(size=11, name="Calibri")
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Gray

thin_border = Border(
    left=Side(style='thin', color="D9D9D9"),
    right=Side(style='thin', color="D9D9D9"),
    top=Side(style='thin', color="D9D9D9"),
    bottom=Side(style='thin', color="D9D9D9")
)

# Apply header styles
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    
    # Auto-adjust column width
    header_length = len(str(cell.value))
    ws.column_dimensions[cell.column_letter].width = max(15, header_length + 5)

# Apply data styles (alternating rows)
for row in range(2, ws.max_row + 1):
    fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.fill = fill
        cell.border = thin_border
        
        # Center align specific columns
        if col in [5, 6, 7, 9]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Freeze top row
ws.freeze_panes = "A2"

wb.save(file_path)
print("Data updated and premium styling applied to New_Romantasy_Books.xlsx successfully.")
