import os
import sys
from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from apply_jra_style import apply_styling

FILE_PATH = r"e:\Internship\PocketFM\LDLA_Combined.xlsx"

def safely_fill_agent_publisher():
    if not os.path.exists(FILE_PATH):
        print(f"Error: {FILE_PATH} not found!")
        return

    print(f"Safely opening {FILE_PATH} without overwriting data...")
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    # Find the column indices we need (1-based)
    header = {cell.value: cell.column for cell in ws[1]}
    
    agent_col = header.get('Name of agent in the main folder')
    publisher_col = header.get('Publisher')

    if not agent_col or not publisher_col:
        print("Error: Missing required columns in the sheet!")
        return

    print("Updating Agent and Publisher columns cell by cell...")
    
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=agent_col).value = "Katie Gisondi"
        ws.cell(row=row_idx, column=publisher_col).value = "Laura Dail Literary Agency"

    print("Surgically saving exact cell modifications...")
    wb.save(FILE_PATH)
    
    print("Re-applying JRA styling safely...")
    apply_styling(FILE_PATH)
    
    print("Agent and Publisher columns successfully populated!")

if __name__ == '__main__':
    safely_fill_agent_publisher()
