from openpyxl import load_workbook

def main():
    file_path = 'e:/Internship/PocketFM/PocketFM_CT_Analysis_Master.xlsx'
    print(f"Loading {file_path}...")
    wb = load_workbook(file_path)
    
    sheet_to_remove = 'Sheet3 (Calculated Hours)'
    
    if sheet_to_remove in wb.sheetnames:
        print(f"Removing '{sheet_to_remove}'...")
        del wb[sheet_to_remove]
        wb.save(file_path)
        print("Successfully removed the tab and saved the workbook!")
    else:
        print(f"Tab '{sheet_to_remove}' not found. It may have already been removed.")

if __name__ == '__main__':
    main()
