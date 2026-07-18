import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

file_path = r"e:\Internship\PocketFM\podium_data.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Style Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Adjust column widths and style cells
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        
        for cell in col:
            # apply borders and alignment to all cells
            if cell.row != 1:
                cell.alignment = wrap_align
                cell.border = thin_border
                
            try:
                val = str(cell.value) if cell.value is not None else ""
                # For long texts like summaries, don't use them to calculate max width
                if len(val) > max_length and len(val) < 100:
                    max_length = len(val)
            except:
                pass
                
        # Set column width
        adjusted_width = min(max_length + 2, 50)
        
        # Adjust specific columns manually for better readability
        # Book Title, Series Name, Summary, URLs
        if ws.cell(row=1, column=col[0].column).value == "Podium Summary / Description":
            ws.column_dimensions[column_letter].width = 80
        elif ws.cell(row=1, column=col[0].column).value in ["Podium URL", "Goodreads Series URL"]:
            ws.column_dimensions[column_letter].width = 40
        else:
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    wb.save(file_path)
    print("Styling applied successfully.")
except Exception as e:
    print(f"Error styling excel: {e}")
