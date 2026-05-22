import os
from openpyxl import load_workbook
import sys

def update_columns(file_path):
    print(f"Updating {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    updated_count = 0
    for row in range(2, ws.max_row + 1):
        # Only update if the row has a title
        title = ws.cell(row=row, column=1).value
        if title and str(title).strip():
            ws.cell(row=row, column=3).value = "Darley Anderson Literary Agency"
            ws.cell(row=row, column=11).value = "Camilla Bolton"
            updated_count += 1
            
    wb.save(file_path)
    print(f"Successfully updated {updated_count} rows!")
    
    import subprocess
    subprocess.Popen(["start", file_path], shell=True)

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    target = os.path.join(base, "Darley_Anderson_Formatted.xlsx")
    update_columns(target)
