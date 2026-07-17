from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import get_column_letter

file_path = r'e:\Internship\PocketFM\Small_Publishing_Agencies_Imprint_Status_Rechecked.xlsx'

wb = load_workbook(file_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]
new_col_title = "Is it Romantasy or not"

if new_col_title not in headers:
    new_col_idx = ws.max_column + 1
    new_header_cell = ws.cell(row=1, column=new_col_idx)
    new_header_cell.value = new_col_title

    # Copy formatting from the previous header
    prev_header_cell = ws.cell(row=1, column=new_col_idx - 1)
    if prev_header_cell.has_style:
        new_header_cell.font = copy(prev_header_cell.font)
        new_header_cell.fill = copy(prev_header_cell.fill)
        new_header_cell.border = copy(prev_header_cell.border)
        new_header_cell.alignment = copy(prev_header_cell.alignment)
        new_header_cell.number_format = copy(prev_header_cell.number_format)

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 25
else:
    new_col_idx = headers.index(new_col_title) + 1

try:
    genre_col_idx = headers.index("Genres they specialise in") + 1
except ValueError:
    genre_col_idx = None

# Populate the cells if we can find the genres column
if genre_col_idx:
    for row_idx in range(2, ws.max_row + 1):
        genre_val = str(ws.cell(row=row_idx, column=genre_col_idx).value or "")
        
        is_romantasy = "romantasy" in genre_val.lower()
        
        target_cell = ws.cell(row=row_idx, column=new_col_idx)
        target_cell.value = "Yes" if is_romantasy else "No"
        
        prev_cell = ws.cell(row=row_idx, column=new_col_idx - 1)
        if prev_cell.has_style:
            target_cell.font = copy(prev_cell.font)
            target_cell.fill = copy(prev_cell.fill)
            target_cell.border = copy(prev_cell.border)
            target_cell.alignment = copy(prev_cell.alignment)
            target_cell.number_format = copy(prev_cell.number_format)

wb.save(file_path)
print("Added and populated new column 'Is it Romantasy or not' to the sheet.")
