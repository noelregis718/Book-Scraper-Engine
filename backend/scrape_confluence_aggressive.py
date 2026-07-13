import asyncio
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from playwright.async_api import async_playwright
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

EXCEL_FILE = '../New_Agency_Template.xlsx'
PROFILE_DIR = '../playwright_goodreads_profile'

async def process_author_aggressive(context, author_name, sem, scraper, results_dict):
    async with sem:
        try:
            print(f"\n[AGGR] Searching: {author_name}...")
            books = await scraper.scrape_top_books_by_author(context, author_name, count=2)
            results_dict[author_name] = books
            print(f"  -> [{author_name}] Success: Extracted {len(books)} books.")
        except Exception as e:
            print(f"  -> [{author_name}] Error: {e}")
            results_dict[author_name] = []

async def main():
    print(f"Loading Excel file: {EXCEL_FILE}")
    df = pd.read_excel(EXCEL_FILE)
    
    missing_mask = df['Name of Series'].isna() & df['Author Name'].notna()
    missing_authors = df.loc[missing_mask, 'Author Name'].dropna().unique().tolist()
    
    missing_authors = [a for a in missing_authors if str(a).strip() not in ["", "N/A", "Unknown", "nan"]]
    print(f"Found {len(missing_authors)} authors requiring aggressive extraction.")
    
    if not missing_authors:
        print("No missing authors found. Exit.")
        return
        
    results_dict = {}
    sem = asyncio.Semaphore(3) # Safe concurrency for aggressive mode
    scraper = GoodreadsScraper(headless=False)
    
    async with async_playwright() as p:
        print("Launching browser with persistent profile...")
        context = await p.chromium.launch_persistent_context(
            user_data_dir=os.path.abspath(PROFILE_DIR),
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            args=['--disable-blink-features=AutomationControlled']
        )
        
        page = await context.new_page()
        await scraper.login_to_goodreads(page)
        await page.close()
        
        tasks = []
        for author in missing_authors:
            tasks.append(process_author_aggressive(context, author, sem, scraper, results_dict))
            
        await asyncio.gather(*tasks)
        await context.close()
        
    print("\nAggressive scraping complete! Rebuilding Excel sheet...")
    
    final_rows = []
    
    for idx, row in df.iterrows():
        title = row.get('Name of Series')
        author = row.get('Author Name')
        
        if pd.notna(title) and str(title).strip() != "":
            final_rows.append(row.to_dict())
        else:
            if author in results_dict and results_dict[author]:
                scraped_books = results_dict[author]
                for book in scraped_books:
                    new_row = {
                        'Name of Series': book.get('Book_Title', 'N/A'),
                        'Author Name': author,
                        'Publisher': row.get('Publisher'),
                        'GoodReads series link': book.get('GoodReads_Series_URL', book.get('GoodReads_Book_URL', 'N/A')),
                        'Number of PRIMARY books in the series': book.get('Num_Primary_Books', 'N/A'),
                        'Rating (out of 5) of Primary Book 1': book.get('GoodReads_Rating', 'N/A'),
                        'Ratings (#) of Primary Book 1': book.get('GoodReads_Rating_Count', 'N/A'),
                        'Synopsis (if available)': book.get('Description', 'N/A'),
                        'Romantasy = Yes or No?': book.get('Romantasy_Subgenre', 'N/A'),
                        'Romantasy Sub-Genre of series': book.get('Sub_Genre', 'N/A'),
                        'Name of agent in the main folder': row.get('Name of agent in the main folder', 'Confluence Literary Agency')
                    }
                    final_rows.append(new_row)
            else:
                final_rows.append(row.to_dict())
                
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Agency Data"
    
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin', color='D3D3D3'), 
                         right=Side(style='thin', color='D3D3D3'), 
                         top=Side(style='thin', color='D3D3D3'), 
                         bottom=Side(style='thin', color='D3D3D3'))
                         
    headers = [
        "Name of Series", "Author Name", "Publisher", "GoodReads series link",
        "Number of PRIMARY books in the series", "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1", "Synopsis (if available)",
        "Romantasy = Yes or No?", "Romantasy Sub-Genre of series",
        "Name of agent in the main folder"
    ]
    
    ws_new.append(headers)
    for cell in ws_new[1]:
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for r_dict in final_rows:
        row_data = [
            r_dict.get(col) if pd.notna(r_dict.get(col)) else None for col in headers
        ]
        ws_new.append(row_data)
        for cell in ws_new[ws_new.max_row]:
            cell.alignment = cell_alignment
            cell.border = thin_border
            
    column_widths = {
        'A': 25, 'B': 20, 'C': 15, 'D': 35, 'E': 15,
        'F': 12, 'G': 12, 'H': 50, 'I': 15, 'J': 15, 'K': 25
    }
    for col, width in column_widths.items():
        ws_new.column_dimensions[col].width = width
        
    ws_new.freeze_panes = 'A2'
    
    wb_new.save(EXCEL_FILE)
    print(f"Successfully saved {ws_new.max_row - 1} records to {EXCEL_FILE}")

if __name__ == "__main__":
    asyncio.run(main())
