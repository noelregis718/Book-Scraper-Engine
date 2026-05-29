import pandas as pd
import os
import sys

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def apply_azantian_style(file_path):
    df = pd.read_excel(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Books"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    columns_count = len(df.columns)
    for col in range(1, columns_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, columns_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Azantian_LitAgency_Combined_Formatted.xlsx")

KEYWORD_MAP = [
    ("Werewolf / Shifter Romance", ["werewolf", "shifter", "alpha", "pack", "omega", "luna", "wolf", "lycan"]),
    ("Monster Romance (Non-Shifter)", ["monster", "orc", "kraken", "alien", "beast", "creature", "demon lover", "tentacle"]),
    ("Mythology, Legend & Fairy Tale Retelling", ["retelling", "mythology", "greek god", "hades", "persephone", "fairy tale", "legend", "arthurian", "norse", "anansi", "medusa", "circe", "orpheus", "beauty and the beast"]),
    ("War College / Military Academy", ["war college", "military academy", "dragon rider", "fourth wing", "basgiath", "aerial", "flight school", "training camp", "rider", "bonded dragon", "academy"]),
    ("High-Stakes Games & Deadly Trials", ["trial", "deadly game", "tournament", "competition", "survival", "arena", "hunger game", "death match", "blood game", "lethal"]),
    ("Dark Academia Romantasy", ["dark academia", "secret society", "forbidden library", "ancient university", "campus", "scholarly", "cursed school", "arcane academy", "magic school"]),
    ("Gothic Dark Romantasy", ["gothic", "haunted", "manor", "dark romance", "gloomy castle", "shadow court", "cursed castle", "decaying estate", "vampire lord", "immortal lord", "macabre"]),
    ("Korean Romance Fantasy / Isekai", ["isekai", "reincarnated", "villainess", "otome", "korean", "manhwa", "transmigrated", "possessed", "regression", "saintess", "empress"]),
    ("Cozy / Cottagecore", ["cozy", "cottagecore", "small town magic", "bakery", "low stakes", "wholesome", "botanical", "village witch", "flower shop", "magical inn"]),
    ("Paranormal Romance", ["vampire", "ghost", "witch", "paranormal", "psychic", "medium", "warlock", "necromancer", "haunting", "supernatural", "fae romance", "fairy", "magic", "coven"]),
    ("High Fantasy Court Adventure", ["court", "throne", "kingdom", "royalty", "fae court", "high fantasy", "crown", "empire", "queen", "king", "prince", "realm", "dragon", "epic fantasy", "war", "political intrigue", "elven", "magical kingdom"]),
    ("Urban / Contemporary Fantasy Romance", ["urban fantasy", "modern day", "contemporary fantasy", "hidden world", "secret magic", "real world", "city magic", "supernatural city"])
]

def classify_subgenre(text):
    text = str(text).lower()
    for subgenre, keywords in KEYWORD_MAP:
        if any(k in text for k in keywords):
            return subgenre
    return None

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found!")
        return

    print(f"Loading {EXCEL_FILE}...")
    df = pd.read_excel(EXCEL_FILE)
            
    romantasy_col = "Romantasy = Yes or No?"
    subgenre_col = "Romantasy Sub-Genre of series"
    
    yes_count = 0
    
    for idx, row in df.iterrows():
        title = str(row.get('Name of Series', '')).strip()
        synopsis = str(row.get('Synopsis (if available)', '')).strip()
        
        if title.lower() == 'nan' or not title:
            continue
            
        current_yes_no = str(row.get(romantasy_col, '')).strip()
        
        # Combine all text to classify subgenre
        combined_text = synopsis + " " + title
        
        # Classify
        subgenre_result = classify_subgenre(combined_text)
        
        # Decide Yes/No
        if current_yes_no == 'Yes' or subgenre_result is not None:
            df.at[idx, romantasy_col] = "Yes"
            df.at[idx, subgenre_col] = subgenre_result if subgenre_result else "High Fantasy Court Adventure"
            yes_count += 1
        else:
            df.at[idx, romantasy_col] = "No"
            df.at[idx, subgenre_col] = ""

    print(f"Saving {EXCEL_FILE} with {yes_count} Romantasy matches...")
    df.to_excel(EXCEL_FILE, index=False)
    
    try:
        apply_azantian_style(EXCEL_FILE)
        print("--- Applied styling ---")
    except Exception as e:
        print(f"Could not apply styling: {e}")
        
    print("ALL DONE!")

if __name__ == '__main__':
    main()
