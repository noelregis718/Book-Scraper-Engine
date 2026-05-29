import asyncio
import os
import sys
import pandas as pd
import re
import json
from playwright.async_api import async_playwright

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper, clean_text, normalize_title_for_search

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Azantian_LitAgency_Combined_Formatted.xlsx")
MAX_CONCURRENT = 5

def apply_azantian_style(file_path):
    df = pd.read_excel(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Books"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    columns_count = len(df.columns)
    for col in range(1, columns_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, columns_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)

def normalize_title(title):
    if pd.isna(title) or not title:
        return ""
    t = str(title).lower()
    t = re.sub(r'[^\w\s]', '', t)
    return t.strip()

async def scrape_new_book_aggressively(context, title, author, semaphore):
    async with semaphore:
        page = await context.new_page()
        await page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        
        safe_title = title.encode('ascii', 'ignore').decode('ascii') if title else 'Unknown'
        safe_author = author.encode('ascii', 'ignore').decode('ascii') if author else 'Unknown'
        print(f"  [Aggressive Scrape Start] '{safe_title}' by {safe_author}")

        row_data = {
            "Name of Series": title,
            "Author Name": author,
            "Publisher": "",
            "GoodReads series link": "N/A",
            "Number of PRIMARY books in the series": 1,
            "Rating (out of 5) of Primary Book 1": "N/A",
            "Ratings (#) of Primary Book 1": "N/A",
            "Synopsis (if available)": "N/A",
            "Romantasy = Yes or No?": "No",
            "Romantasy Sub-Genre of series": "",
            "Name of agent": "Azantian Literary Agency"
        }

        try:
            norm_title = normalize_title_for_search(title)
            super_clean_title = re.sub(r'[^a-zA-Z0-9\s]', '', norm_title)
            
            author_for_search = author if author and str(author).strip() != "" else ""
            
            queries = [
                f"{super_clean_title} {author_for_search}".strip(),
                f"{norm_title} {author_for_search}".strip(),
                norm_title.strip(),
                super_clean_title.strip()
            ]
            
            book_url = None
            
            for query in queries:
                if not query: continue
                search_url = f"https://www.goodreads.com/search?q={query.replace(' ', '+')}"
                
                await page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)

                current_url = page.url
                if "/book/show/" in current_url:
                    book_url = current_url
                    break
                else:
                    try:
                        first_link = await page.wait_for_selector(
                            'a.bookTitle, [data-testid="bookTitle"] a, h3 a[href*="/book/show/"]',
                            timeout=3000
                        )
                        book_url = await first_link.evaluate("el => el.href")
                        break
                    except:
                        pass
            
            if not book_url:
                print(f"  [Not Found] '{safe_title}'")
                await page.close()
                return row_data

            await page.goto(book_url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(1.5)

            avg_rating, rating_count = "N/A", "N/A"
            try:
                ld_el = await page.query_selector('script[type="application/ld+json"]')
                if ld_el:
                    ld_data = json.loads(await ld_el.inner_text())
                    if isinstance(ld_data, list): ld_data = ld_data[0]
                    avg_rating = str(ld_data.get('aggregateRating', {}).get('ratingValue', 'N/A'))
                    rating_count = str(ld_data.get('aggregateRating', {}).get('ratingCount', 'N/A'))
            except: pass

            description = "N/A"
            desc_el = await page.query_selector('[data-testid="description"] .Formatted, .readable')
            if desc_el:
                description = clean_text(await desc_el.inner_text())

            series_url = book_url
            series_link_el = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a')
            if series_link_el:
                series_url = await series_link_el.evaluate("el => el.href")

            num_primary = "1"
            book1_rating = avg_rating
            book1_ratings = rating_count
            if "/series/" in series_url:
                try:
                    await page.goto(series_url, wait_until="domcontentloaded", timeout=45000)
                    content = await page.content()
                    m = re.search(r'(\d+)\s+primary\s+works', content, re.IGNORECASE)
                    if m: num_primary = m.group(1)
                    row1 = await page.query_selector('.listWithDividers__item, .seriesWork')
                    if row1:
                        rtxt = (await row1.inner_text()).lower()
                        r_match = re.search(r'([\d.]+)\s+avg\s+rating\s+[—\-]\s+([\d,]+)\s+ratings', rtxt)
                        if r_match:
                            book1_rating = r_match.group(1)
                            book1_ratings = r_match.group(2).replace(',', '')
                except: pass

            row_data["GoodReads series link"] = series_url
            row_data["Number of PRIMARY books in the series"] = num_primary
            row_data["Rating (out of 5) of Primary Book 1"] = book1_rating
            row_data["Ratings (#) of Primary Book 1"] = book1_ratings
            row_data["Synopsis (if available)"] = description
            
            # Keyword check for romantasy
            g = description.lower()
            keywords = ["fantasy", "paranormal", "romantasy", "dragon", "magic", "supernatural", "fae", "witch", "vampire", "shifter"]
            if any(k in g for k in keywords):
                row_data["Romantasy = Yes or No?"] = "Yes"

            print(f"  [Done] '{safe_title}'")
            
        except Exception as e:
            print(f"  [Error on '{safe_title}']: {e}")
        finally:
            try:
                await page.close()
            except: pass
            
        return row_data

async def run_scrape():
    df = pd.read_excel(EXCEL_FILE)
    
    authors = df['Author Name'].dropna().unique()
    authors = [a for a in authors if str(a).strip() != '']
    print(f"Found {len(authors)} unique authors.")
    
    scraper = GoodreadsScraper()
    new_books_to_scrape = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        login_page = await context.new_page()
        await scraper.login_to_goodreads(login_page)
        
        print("--- Finding top 3 books for each author ---")
        for author in authors:
            safe_author = author.encode('ascii', 'ignore').decode('ascii')
            print(f"Checking author: {safe_author}")
            # Search author and get up to 3 books
            top_books = await scraper.search_author_books_with_links(login_page, author, max_books=3)
            
            if not top_books:
                print(f"  No books found for {safe_author}")
                continue
                
            existing_titles = df[df['Author Name'] == author]['Name of Series'].apply(normalize_title).tolist()
            
            for book in top_books:
                found_title = book['title']
                norm_found = normalize_title(found_title)
                
                exists = False
                for ex in existing_titles:
                    if not ex or not norm_found: continue
                    if ex in norm_found or norm_found in ex:
                        exists = True
                        break
                        
                if exists:
                    print(f"  [Skipping] '{found_title.encode('ascii', 'ignore').decode('ascii')}' - Already in sheet")
                else:
                    print(f"  [Adding to Queue] '{found_title.encode('ascii', 'ignore').decode('ascii')}'")
                    new_books_to_scrape.append({'title': found_title, 'author': author})
        
        await login_page.close()
        
        if not new_books_to_scrape:
            print("No new books to scrape. All top books are already in the sheet!")
        else:
            print(f"\n--- Scraping {len(new_books_to_scrape)} new books aggressively ---")
            semaphore = asyncio.Semaphore(MAX_CONCURRENT)
            tasks = []
            
            for b in new_books_to_scrape:
                tasks.append(scrape_new_book_aggressively(context, b['title'], b['author'], semaphore))
                
            new_rows = await asyncio.gather(*tasks)
            
            new_df = pd.DataFrame(new_rows)
            df = pd.concat([df, new_df], ignore_index=True)
            
            print("--- Rebuilding Excel File with new data ---")
            df.to_excel(EXCEL_FILE, index=False)
            
            try:
                apply_azantian_style(EXCEL_FILE)
                print("--- Applied styling ---")
            except Exception as e:
                print(f"Could not apply styling: {e}")
                
        await browser.close()
    
    print("Opening Excel file...")
    import subprocess
    subprocess.Popen(["start", EXCEL_FILE], shell=True)
    print("ALL DONE!")

if __name__ == '__main__':
    asyncio.run(run_scrape())
