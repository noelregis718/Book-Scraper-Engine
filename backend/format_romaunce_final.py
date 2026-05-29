import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def format_romaunce_final(input_file, output_file):
    df_raw = pd.read_excel(input_file)

    # Standard 11 columns
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    new_df = pd.DataFrame(columns=new_columns, index=range(len(df_raw)))

    # Map existing columns to new columns
    if "Book Title" in df_raw.columns:
        new_df["Name of Series"] = df_raw["Book Title"].values
    if "Author" in df_raw.columns:
        new_df["Author Name"] = df_raw["Author"].values
        
    new_df["Publisher"] = "Romaunce Books"
    
    # Check if romantasy based on genre
    def check_romantasy(genre):
        if not isinstance(genre, str): return "No"
        g = genre.lower()
        keywords = ["fantasy", "paranormal", "romantasy", "dragon", "magic",
                    "supernatural", "fae", "witch", "vampire", "shifter",
                    "werewolf", "gothic", "mytholog"]
        return "Yes" if any(k in g for k in keywords) else "No"
        
    if "Genre" in df_raw.columns:
        new_df["Romantasy = Yes or No?"] = df_raw["Genre"].apply(check_romantasy).values
        new_df["Romantasy Sub-Genre of series"] = df_raw["Genre"].values

    new_df["Name of agent"] = "N/A"
    
    # We shouldn't delete any books, so no dropna based on books (unless they are entirely empty, but let's keep all)
    
    # Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Romaunce Final"

    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, len(new_columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(new_columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_file)
    print(f"Done! Saved to {output_file}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file = os.path.join(base, "Romaunce_Books_Complete.xlsx")
    output_file = os.path.join(base, "Romaunce_Books_Complete_Final.xlsx")
    format_romaunce_final(input_file, output_file)
