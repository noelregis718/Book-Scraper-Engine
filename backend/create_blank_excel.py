import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

file_path = r"E:\Internship\PocketFM\New_Romantasy_Books.xlsx"

# Standard 11 columns
columns = [
    'Name of Series',
    'Author Name',
    'Publisher',
    'GoodReads series link',
    'Number of PRIMARY books in the series',
    'Rating (out of 5) of Primary Book 1',
    'Ratings (#) of Primary Book 1',
    'Synopsis (if available)',
    'Romantasy = Yes or No?',
    'Romantasy Sub-Genre of series',
    'Name of agent in the main folder'
]

# Create empty dataframe with columns
df = pd.DataFrame(columns=columns)
df.to_excel(file_path, index=False)

# Apply premium styling
wb = load_workbook(file_path)
ws = wb.active

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
header_font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
thin_border = Border(
    left=Side(style='thin', color="D9D9D9"),
    right=Side(style='thin', color="D9D9D9"),
    top=Side(style='thin', color="D9D9D9"),
    bottom=Side(style='thin', color="D9D9D9")
)

# Apply header styles
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    
    # Auto-adjust column width
    header_length = len(str(cell.value))
    ws.column_dimensions[cell.column_letter].width = max(15, header_length + 5)

# Freeze top row
ws.freeze_panes = "A2"

wb.save(file_path)
print(f"Successfully created fully formatted blank sheet at: {file_path}")
