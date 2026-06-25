import openpyxl

def main():
    wb = openpyxl.load_workbook('LDLA_Combined_Mapped_Fixed.xlsx')
    ws = wb.active
    print('Max row:', ws.max_row, 'Max col:', ws.max_column)
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print('Headers:', headers)
    
    print('Checking for romantasy outside table...')
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and 'romantasy' in val.lower():
                print(f"Row {r}, Col {c}: {val}")

if __name__ == "__main__":
    main()
