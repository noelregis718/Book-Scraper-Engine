import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = r'e:\Internship\PocketFM\Books_Scraping_Template.xlsx'
print("Loading workbook...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define styles
header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                     right=Side(style='thin', color='BFBFBF'), 
                     top=Side(style='thin', color='BFBFBF'), 
                     bottom=Side(style='thin', color='BFBFBF'))

hyperlink_font = Font(color="0563C1", underline="single", name='Calibri', size=11)
regular_font = Font(name='Calibri', size=11)

print("Formatting header...")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = thin_border

print("Formatting rows and columns...")
# Reasonable column widths
column_widths = {
    'A': 30, # Name of Series
    'B': 25, # Author Name
    'C': 20, # Publisher
    'D': 35, # GoodReads series link
    'E': 15, # Number of PRIMARY books in the series
    'F': 15, # Rating (out of 5) of Primary Book 1
    'G': 15, # Ratings (#) of Primary Book 1
    'H': 70, # Synopsis (if available) - make this wide for reading
    'I': 20, # Romantasy = Yes or No?
    'J': 35, # Romantasy Sub-Genre of series
    'K': 20  # Name of agent
}

for i, col in enumerate(range(1, ws.max_column + 1)):
    letter = get_column_letter(col)
    if letter in column_widths:
        ws.column_dimensions[letter].width = column_widths[letter]

print("Applying data styles...")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment_left_top
        cell.border = thin_border
        
        # Format the hyperlink column (Column D is 4)
        if cell.column == 4 and cell.value and str(cell.value).startswith('http'):
            cell.font = hyperlink_font
            cell.hyperlink = cell.value
        else:
            cell.font = regular_font

# Try to clear specific row heights so Excel auto-fits the wrapped text
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = None

ws.freeze_panes = 'A2'

print("Saving workbook...")
wb.save(file_path)
print("Done!")
