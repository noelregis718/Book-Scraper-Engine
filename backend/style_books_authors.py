import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def apply_styling(file_path):
    print(f"Applying styles to {file_path}...")
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply Header Styles
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # Apply Body Styles
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border
            
    # Set Dynamic Column Widths based on header name
    for col in range(1, ws.max_column + 1):
        header_val = str(ws.cell(row=1, column=col).value).lower()
        col_letter = ws.cell(row=1, column=col).column_letter
        
        width = 20 # default
        if 'synopsis' in header_val:
            width = 60
        elif 'link' in header_val:
            width = 35
        elif 'series' in header_val and 'name' in header_val:
            width = 30
        elif 'title' in header_val:
            width = 35
        elif 'publisher' in header_val:
            width = 25
        elif 'agent' in header_val:
            width = 25
            
        ws.column_dimensions[col_letter].width = width
        
    # Freeze the top row
    ws.freeze_panes = 'A2'
        
    # Save the updated workbook
    wb.save(file_path)
    print("Styling applied successfully. None of the data was deleted.")
    
    import subprocess
    subprocess.Popen(["start", file_path], shell=True)

if __name__ == "__main__":
    apply_styling(r'E:\Internship\PocketFM\books_authors.xlsx')
