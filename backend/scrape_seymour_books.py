import asyncio
import os
import re
import json
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

AUTHORS_EXCEL = r"E:\Internship\PocketFM\Seymour_Agency_Authors.xlsx"
CATALOG_EXCEL = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"
SCRAPED_BOOKS_JSON = r"E:\Internship\PocketFM\backend\scratch\seymour_scraped_books.json"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

CONCURRENCY_LIMIT = 10

def load_scraped_books():
    if os.path.exists(SCRAPED_BOOKS_JSON):
        try:
            with open(SCRAPED_BOOKS_JSON, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_scraped_books(data):
    os.makedirs(os.path.dirname(SCRAPED_BOOKS_JSON), exist_ok=True)
    with open(SCRAPED_BOOKS_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def pre_populate_json_from_excel():
    if os.path.exists(SCRAPED_BOOKS_JSON):
        return
    if not os.path.exists(CATALOG_EXCEL):
        return
        
    print(">>> Recovering previous scraping progress from active Excel catalog...")
    try:
        wb = openpyxl.load_workbook(CATALOG_EXCEL, data_only=True)
        ws = wb.active
        
        recovered = {}
        for r in range(2, ws.max_row + 1):
            book = ws.cell(row=r, column=1).value
            author = ws.cell(row=r, column=2).value
            if author and book:  # Only recover if there is an actual book
                author_str = str(author).strip()
                book_str = str(book).strip()
                if author_str not in recovered:
                    recovered[author_str] = []
                if book_str not in recovered[author_str]:
                    recovered[author_str].append(book_str)
                        
        save_scraped_books(recovered)
        print(f"Success! Recovered previous progress for {len(recovered)} authors.")
    except Exception as e:
        print(f"Warning: Failed to recover previous progress from excel: {e}")

def clean_book_title(img_url, author_name):
    # Extract filename from url
    filename = img_url.split('/')[-1]
    filename_no_ext = os.path.splitext(filename)[0]
    
    # Remove thumb sizes, scaled, and common suffix suffixes
    filename_no_ext = re.sub(r'-\d+x\d+$', '', filename_no_ext)
    filename_no_ext = re.sub(r'-scaled$', '', filename_no_ext)
    
    # Create author slug safely
    author_slug = re.sub(r'[^a-zA-Z0-9]', '-', author_name.lower())
    author_slug = re.sub(r'-+', '-', author_slug).strip('-')
    
    title_slug = filename_no_ext
    # Replace underscores with dashes to unify them
    title_slug = title_slug.replace('_', '-')
    
    if title_slug.lower().startswith(author_slug):
        title_slug = title_slug[len(author_slug):].strip('-')
    else:
        author_parts = author_slug.split('-')
        title_parts = title_slug.lower().split('-')
        match_len = 0
        for i in range(min(len(author_parts), len(title_parts))):
            if author_parts[i] == title_parts[i]:
                match_len += 1
            else:
                break
        if match_len > 0:
            title_slug = '-'.join(title_parts[match_len:])
            
    title_words = title_slug.replace('-', ' ').split()
    
    # Strip common platform and technical size suffixes (e.g. ebook, kobo, sl1500) from the end
    platform_suffixes = {'ebook', 'kobo', 'nook', 'bn'}
    while title_words:
        last_word = title_words[-1].lower()
        if last_word in platform_suffixes or re.match(r'^sl\d+$', last_word):
            title_words.pop()
        else:
            break
        
    minor_words = {'a', 'an', 'the', 'and', 'but', 'or', 'for', 'nor', 'on', 'at', 'to', 'by', 'of', 'in'}
    
    cleaned_words = []
    for idx, word in enumerate(title_words):
        if idx == 0 or word.lower() not in minor_words:
            cleaned_words.append(word.capitalize())
        else:
            cleaned_words.append(word.lower())
            
    return ' '.join(cleaned_words)

async def scrape_author_page(sem, browser, idx, total, name, url, results):
    if not url or url == "N/A":
        print(f"[{idx}/{total}] Skipping {name} (No profile URL)")
        return

    async with sem:
        page = None
        try:
            page = await browser.new_page()
            # Disable image and stylesheet loading for ultra-fast performance since we only need the HTML DOM
            await page.route("**/*.{png,jpg,jpeg,gif,webp,css,woff,woff2,svg}", lambda route: route.abort())
            
            await page.goto(url, wait_until="domcontentloaded", timeout=45000)
            await page.wait_for_timeout(1000)  # Wait briefly for rendering to settle
            
            content = await page.content()
            soup = BeautifulSoup(content, "html.parser")
            
            # Find related books section title
            heading = soup.find(class_="xixs-related-books-section-title")
            
            books = []
            if heading:
                books_container = soup.find(class_="xixs-related-books")
                if books_container:
                    img_tags = books_container.find_all("img")
                    for img in img_tags:
                        img_url = img.get("src") or img.parent.get("href")
                        if img_url:
                            book_title = clean_book_title(img_url, name)
                            if book_title:
                                books.append(book_title)
            
            results[name] = books
            print(f"[{idx}/{total}] Scraped {name} -> Found {len(books)} books")
            
        except Exception as e:
            print(f"[{idx}/{total}] Error scraping {name}: {e}")
            results[name] = []
            
        finally:
            if page:
                await page.close()

async def crawl_all_profiles(authors):
    print(f">>> Initializing parallel Playwright crawl with concurrency={CONCURRENCY_LIMIT}...")
    sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
    
    results = {}
    async with async_playwright() as p:
        # Launch headed Chromium so the user can see the crawling live
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        tasks = []
        total = len(authors)
        for idx, (name, url) in enumerate(authors, 1):
            tasks.append(scrape_author_page(sem, context, idx, total, name, url, results))
            
        await asyncio.gather(*tasks, return_exceptions=True)
        await browser.close()
        
    return results

def duplicate_row(ws, source_row_idx, target_row_idx):
    ws.insert_rows(target_row_idx)
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)
        
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height

def format_excel_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def generate_master_catalog(scraped_books):
    if not os.path.exists(AUTHORS_EXCEL):
        print(f"Error: {AUTHORS_EXCEL} not found.")
        return
        
    print(">>> Generating completed 347-author styled catalog...")
    wb_src = openpyxl.load_workbook(AUTHORS_EXCEL, data_only=True)
    ws_src = wb_src.active
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Seymour Media Catalog"
    
    # Append the headers
    ws_out.append(HEADERS)
    
    # Load ALL 347 base authors
    base_rows = []
    for r in range(2, ws_src.max_row + 1):
        author_name = ws_src.cell(row=r, column=1).value
        if author_name:
            base_rows.append(str(author_name).strip())
            
    # Write default rows for all authors
    for author in base_rows:
        ws_out.append([
            "",
            author,
            "The Seymour Agency",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A"
        ])
        
    # Format initially
    format_excel_sheet(ws_out)
    
    print(">>> Merging cumulative scraped books into catalog...")
    total_books_added = 0
    # Loop backwards to safely handle duplicates
    for row_idx in range(ws_out.max_row, 1, -1):
        author = ws_out.cell(row=row_idx, column=2).value
        if not author:
            continue
            
        books = scraped_books.get(author, [])
        if books:
            # 1. Update the first book in-place
            ws_out.cell(row=row_idx, column=1, value=books[0])
            total_books_added += 1
            
            # 2. For subsequent books, duplicate row and set title
            for i in range(1, len(books)):
                target_row = row_idx + i
                duplicate_row(ws_out, row_idx, target_row)
                ws_out.cell(row=target_row, column=1, value=books[i])
                total_books_added += 1

    # Refresh formatting and auto-filters over final rows
    format_excel_sheet(ws_out)
    
    print(f">>>> Saving completed master catalog to {CATALOG_EXCEL}...")
    wb_out.save(CATALOG_EXCEL)
    print(f"Success! Saved master catalog with {ws_out.max_row - 1} rows! Total books mapped: {total_books_added}")

def main():
    if not os.path.exists(AUTHORS_EXCEL):
        print(f"Error: {AUTHORS_EXCEL} not found.")
        return
        
    # Recovery and load past progress
    pre_populate_json_from_excel()
    all_scraped = load_scraped_books()
    
    wb = openpyxl.load_workbook(AUTHORS_EXCEL, data_only=True)
    ws = wb.active
    
    authors = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        url = ws.cell(row=r, column=3).value
        if name and url:
            authors.append((str(name).strip(), str(url).strip()))
            
    print(f"Loaded {len(authors)} author profiles from excel.")
    
    # Slice the entire 347 authors list to complete the crawl
    authors_batch = authors
    
    # Filter out already scraped authors
    authors_to_crawl = []
    for name, url in authors_batch:
        if name not in all_scraped:
            authors_to_crawl.append((name, url))
            
    skipped_count = len(authors_batch) - len(authors_to_crawl)
    print(f"Skipping {skipped_count} authors already in database.")
    
    if authors_to_crawl:
        print(f"Crawling {len(authors_to_crawl)} new authors...")
        print(f"Authors to crawl: {[a[0] for a in authors_to_crawl]}")
        
        # Run the crawl for only the new authors
        batch_results = asyncio.run(crawl_all_profiles(authors_to_crawl))
        
        # Merge and persist results
        all_scraped.update(batch_results)
        save_scraped_books(all_scraped)
    else:
        print("All authors in this batch are already scraped! Skipping Playwright crawl.")
        
    # Generate the completed styled catalog
    generate_master_catalog(all_scraped)

if __name__ == "__main__":
    main()
