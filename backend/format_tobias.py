import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def format_tobias_excel(input_file, output_file):
    # Read the data, skipping the first two title/source rows
    df_raw = pd.read_excel(input_file, skiprows=2, header=0)

    # Rename columns from row 3 in the original file
    # Cols: #, Book Title, Author, Genre, Agent, Publisher, Notes / Deal Info
    df_raw.columns = ["#", "Book Title", "Author", "Genre", "Agent", "Publisher", "Notes"]

    # Drop empty rows
    df_raw = df_raw.dropna(subset=["Book Title"])
    df_raw = df_raw[df_raw["Book Title"].astype(str).str.strip() != ""]
    df_raw = df_raw[~df_raw["Book Title"].astype(str).str.startswith("#")]

    # Build new 11-column dataframe
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    new_df = pd.DataFrame(columns=new_columns, index=range(len(df_raw)))

    new_df["Name of Series"] = df_raw["Book Title"].values
    new_df["Author Name"] = df_raw["Author"].values
    new_df["Publisher"] = df_raw["Publisher"].apply(
        lambda x: "Tobias Literary Agency" if (str(x).strip() == "—" or str(x).strip() == "" or pd.isna(x)) else str(x)
    ).values
    new_df["Name of agent"] = df_raw["Agent"].apply(
        lambda x: "Tobias Literary Agency" if (str(x).strip() in ["TLA", "—", ""] or pd.isna(x)) else str(x)
    ).values

    # Detect Romantasy from genre
    def check_romantasy(genre):
        if not isinstance(genre, str): return "No"
        g = genre.lower()
        if any(k in g for k in ["fantasy", "paranormal", "romantasy", "dragon", "magic", "supernatural"]):
            return "Yes"
        return "No"

    new_df["Romantasy = Yes or No?"] = df_raw["Genre"].apply(check_romantasy).values
    new_df["Romantasy Sub-Genre of series"] = df_raw["Genre"].values

    # Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Tobias Bookshelf"

    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format Headers
    for col in range(1, len(new_columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Format Data
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(new_columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    # Set Column Widths
    widths = {
        "A": 30,  # Name of Series
        "B": 20,  # Author Name
        "C": 25,  # Publisher
        "D": 35,  # GoodReads series link
        "E": 15,  # Number of PRIMARY books
        "F": 15,  # Rating
        "G": 15,  # Ratings (#)
        "H": 50,  # Synopsis
        "I": 15,  # Romantasy
        "J": 20,  # Sub-Genre
        "K": 25   # Name of agent
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_file)
    print(f"Done! Saved to {output_file}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file = os.path.join(base, "Tobias_All_144_Books.xlsx")
    output_file = os.path.join(base, "Tobias_All_144_Books.xlsx")
    format_tobias_excel(input_file, output_file)
