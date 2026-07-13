import asyncio
import os
import openpyxl
from playwright.async_api import async_playwright
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

EXCEL_FILE = '../New_Agency_Template.xlsx'
PROFILE_DIR = '../playwright_goodreads_profile'

async def fetch_link(context, scraper, title, author, sem, row_idx, ws):
    async with sem:
        try:
            print(f"[Row {row_idx}] Goodreads Search: '{title}' by {author}...")
            # Use the robust internal scraper method that handles captchas/retries
            book_data = await scraper.scrape_goodreads_data(context, title, author, "")
            
            if book_data and book_data.get("GoodReads_Book_URL") and book_data["GoodReads_Book_URL"] != "N/A":
                url = book_data["GoodReads_Book_URL"]
                print(f"  -> [Row {row_idx}] Success: {url}")
                ws.cell(row=row_idx, column=4).value = url
            else:
                print(f"  -> [Row {row_idx}] Failed to extract.")
                ws.cell(row=row_idx, column=4).value = "Not Found"
                
        except Exception as e:
            print(f"  -> [Row {row_idx}] Error: {e}")
            ws.cell(row=row_idx, column=4).value = "Error"

async def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    tasks_data = []
    
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        author = ws.cell(row=row, column=2).value
        link = ws.cell(row=row, column=4).value
        
        if title and str(title).strip() and str(title).strip() != "N/A":
            if not link or str(link).strip() in ["N/A", "nan", "Not Found", "Error"]:
                tasks_data.append({
                    'row': row,
                    'title': str(title).strip(),
                    'author': str(author).strip() if author else ""
                })
                
    print(f"Found {len(tasks_data)} books still missing links.")
    if not tasks_data:
        return
        
    sem = asyncio.Semaphore(4) # 4 tabs concurrently to avoid crashing execution context
    scraper = GoodreadsScraper(headless=True)
    
    async with async_playwright() as p:
        print("Launching persistent browser...")
        context = await p.chromium.launch_persistent_context(
            user_data_dir=os.path.abspath(PROFILE_DIR),
            headless=True,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            args=['--disable-blink-features=AutomationControlled']
        )
        
        tasks = []
        for t in tasks_data:
            tasks.append(fetch_link(context, scraper, t['title'], t['author'], sem, t['row'], ws))
            
        print("--- Starting Goodreads Internal Link extraction ---")
        await asyncio.gather(*tasks)
        await context.close()
        
    wb.save(EXCEL_FILE)
    print("Excel file updated successfully with recovered links!")

if __name__ == "__main__":
    asyncio.run(main())
