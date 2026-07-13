import openpyxl
import re

TAXONOMY = {
    "High Fantasy Court Adventure": [
        "fae", "elven", "elf", "kingdom", "court", "empire", "epic", "crown", "prince", "princess", "queen", "king", 
        "throne", "heir", "castle", "realm", "rebellion", "sword", "magic", "sorcery", "high fantasy", "royalty", "emperor", "empress"
    ],
    "Gothic Dark Romantasy": [
        "gothic", "dark magic", "haunted", "vampiric", "macabre", "blood", "curse", "death", "graveyard", "necromancer", 
        "mansion", "shadows", "darkness", "brooding", "sinister", "monster", "demon"
    ],
    "Dark Academia Romantasy": [
        "academy", "secret society", "school", "university", "scholar", "campus", "library", "professors", "students", 
        "dark academia", "institute", "college", "boarding school", "dormitory"
    ],
    "Monster Romance (Non-Shifter)": [
        "monster", "tentacle", "orc", "kraken", "minotaur", "beast", "alien", "demon", "gargoyle", "creature", "non-human"
    ],
    "Werewolf / Shifter Romance": [
        "shifter", "wolf", "bear", "dragon", "pack", "alpha", "omega", "mate", "lycan", "werewolf", "fated", "luna"
    ],
    "High-Stakes Games & Deadly Trials": [
        "trial", "tournament", "game", "contest", "survival", "arena", "hunger games", "competition", "deadly", "compete"
    ],
    "Mythology, Legend & Fairy Tale Retelling": [
        "myth", "greek", "norse", "retelling", "fairy tale", "god", "goddess", "hades", "persephone", "olympus", 
        "beauty and the beast", "cinderella", "legend", "folklore"
    ],
    "War College / Military Academy": [
        "war college", "military", "dragon rider", "combat", "training", "soldier", "army", "warrior", "squad", "legion", "commander"
    ],
    "Korean Romance Fantasy / Isekai": [
        "isekai", "reincarnation", "villainess", "empress", "saintess", "transmigration", "manhwa", "korean", "past life", "reborn"
    ],
    "Paranormal Romance": [
        "vampire", "demon", "succubus", "witch", "coven", "warlock", "ghost", "spirit", "paranormal", "supernatural", "angel"
    ],
    "Cozy / Cottagecore": [
        "cozy", "cottage", "bakery", "small town", "tea", "coffee", "inn", "tavern", "low-stakes", "healing", "wholesome", "familiar"
    ],
    "Urban / Contemporary Fantasy Romance": [
        "urban", "city", "hidden world", "modern", "detective", "agency", "underworld", "new york", "london", "contemporary"
    ]
}

ROMANCE_KEYWORDS = ["love", "romance", "desire", "passion", "kiss", "heart", "lover", "mate", "bond", "seduction", "attraction", "feelings", "marriage", "betrothal", "wedding", "husband", "wife", "boyfriend", "girlfriend", "smut", "spicy"]

def classify_romantasy():
    excel_file = '../New_Agency_Template.xlsx'
    print(f"Loading {excel_file} for classification...")
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    count_yes = 0
    count_no = 0
    
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value or ""
        synopsis = ws.cell(row=row, column=8).value or ""
        
        text_to_analyze = (str(title) + " " + str(synopsis)).lower()
        
        if not text_to_analyze.strip() or text_to_analyze.strip() == "nan":
            ws.cell(row=row, column=9).value = 'No'
            ws.cell(row=row, column=10).value = 'N/A'
            continue
            
        scores = {genre: 0 for genre in TAXONOMY.keys()}
        
        for genre, keywords in TAXONOMY.items():
            for kw in keywords:
                pattern = rf"\b{re.escape(kw)}\b"
                matches = len(re.findall(pattern, text_to_analyze))
                scores[genre] += matches
                
        best_genre = max(scores, key=scores.get)
        max_score = scores[best_genre]
        
        romance_score = 0
        for kw in ROMANCE_KEYWORDS:
            pattern = rf"\b{re.escape(kw)}\b"
            romance_score += len(re.findall(pattern, text_to_analyze))
            
        inherently_romantic = ["Monster Romance (Non-Shifter)", "Werewolf / Shifter Romance", "Paranormal Romance", "Gothic Dark Romantasy", "Korean Romance Fantasy / Isekai"]
        
        is_romantasy = False
        if max_score > 0:
            if romance_score > 0 or best_genre in inherently_romantic:
                is_romantasy = True
            elif max_score > 1:
                is_romantasy = True
                
        if is_romantasy:
            ws.cell(row=row, column=9).value = 'Yes'
            ws.cell(row=row, column=10).value = best_genre
            count_yes += 1
        else:
            ws.cell(row=row, column=9).value = 'No'
            ws.cell(row=row, column=10).value = 'N/A'
            count_no += 1
            
    print(f"Classification complete! Found {count_yes} Romantasy books and {count_no} Non-Romantasy/Unknown.")
    
    wb.save(excel_file)
    print("Excel file successfully updated with Romantasy categorizations!")

if __name__ == "__main__":
    classify_romantasy()
