import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def format_root_literary(input_file, output_file):
    # -----------------------------------------------------------
    # 1. Read source data (skip rows 1-2: title + subtitle rows)
    #    Row 3 = header: #, Book Title, Author, Agent at Root Literary
    #    Rows 4+ = data
    # -----------------------------------------------------------
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in['Root Literary Books']

    rows = list(ws_in.iter_rows(min_row=4, max_row=ws_in.max_row, values_only=True))
    # Filter out completely blank rows
    rows = [r for r in rows if any(v is not None for v in r)]

    # -----------------------------------------------------------
    # 2. Build data in the 11-column format
    # -----------------------------------------------------------
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    data_rows = []
    for row in rows:
        # row = (#, Book Title, Author, Agent at Root Literary)
        _, book_title, author, agent = row[0], row[1], row[2], row[3]

        if book_title is None or str(book_title).strip() == "":
            continue

        agent_val = str(agent).strip() if agent else "Root Literary"

        new_row = [
            str(book_title).strip(),   # Name of Series  (= Book Title as-is)
            str(author).strip() if author else "",  # Author Name (as-is)
            "",                        # Publisher          — to be scraped
            "",                        # GoodReads series link — to be scraped
            "",                        # Number of PRIMARY books
            "",                        # Rating (out of 5) of Primary Book 1
            "",                        # Ratings (#) of Primary Book 1
            "",                        # Synopsis (if available)
            "",                        # Romantasy = Yes or No?
            "",                        # Romantasy Sub-Genre of series
            agent_val,                 # Name of agent
        ]
        data_rows.append(new_row)

    # -----------------------------------------------------------
    # 3. Write to new workbook with styling
    # -----------------------------------------------------------
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Root Literary"

    # Style constants
    header_fill   = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style='thin', color="CCCCCC")
    thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alternating row fills
    fill_even = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",  fill_type="solid")

    # Write header row
    ws_out.append(new_columns)

    # Write data rows
    for data_row in data_rows:
        ws_out.append(data_row)

    # Apply header styling
    for col_idx in range(1, len(new_columns) + 1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = align_center
        cell.border  = thin_border

    # Freeze header row
    ws_out.freeze_panes = "A2"

    # Apply data row styling
    for row_idx in range(2, ws_out.max_row + 1):
        row_fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx in range(1, len(new_columns) + 1):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            cell.fill      = row_fill
            cell.alignment = align_top
            cell.border    = thin_border

    # Column widths
    col_widths = {
        "A": 35,   # Name of Series
        "B": 28,   # Author Name
        "C": 22,   # Publisher
        "D": 40,   # GoodReads series link
        "E": 18,   # Number of PRIMARY books
        "F": 18,   # Rating (out of 5)
        "G": 16,   # Ratings (#)
        "H": 55,   # Synopsis
        "I": 16,   # Romantasy Y/N
        "J": 28,   # Romantasy Sub-Genre
        "K": 25,   # Name of agent
    }
    for col_letter, width in col_widths.items():
        ws_out.column_dimensions[col_letter].width = width

    # Row height for header
    ws_out.row_dimensions[1].height = 40

    wb_out.save(output_file)
    print(f"Done! {len(data_rows)} rows written -> {output_file}")


if __name__ == "__main__":
    base        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file  = os.path.join(base, "Root_Literary_Combined.xlsx")
    output_file = os.path.join(base, "Root_Literary_Formatted.xlsx")
    format_root_literary(input_file, output_file)
