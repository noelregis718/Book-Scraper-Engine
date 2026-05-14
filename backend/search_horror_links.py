import asyncio
import pandas as pd
import os
import sys
import random
import json
from playwright.async_api import async_playwright
from openpyxl import load_workbook

# Add current directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from goodreads_scraper import GoodreadsScraper, normalize_title_for_search

# Configuration
INPUT_FILE = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
OUTPUT_FILE = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
STATE_FILE = "horror_progress.json"
BATCH_SIZE = 10 
# Target End Row (Excel Numbering)
TARGET_END_ROW = 3700 

def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                return json.load(f).get("last_processed_row", 1)
        except:
            return 1
    return 1

def save_state(last_row):
    with open(STATE_FILE, 'w') as f:
        json.dump({"last_processed_row": last_row}, f)

async def process_book(index, row, context, gr_scraper, ws, total_in_batch, counter):
    counter[0] += 1
    curr = counter[0]
    
    excel_title = str(row.get("Title", "Unknown"))
    
    # Title Only Search
    print(f"  [{curr}/{total_in_batch}] Searching (Title Only): {excel_title}...")
    
    match_found = False
    gr_data = {}
    
    try:
        await asyncio.sleep(random.uniform(0.1, 0.5))
        gr_data = await gr_scraper.scrape_goodreads_data(context, excel_title, "")
        
        if gr_data:
            found_url = gr_data.get('GoodReads_Book_URL', 'N/A')
            found_title = gr_data.get('Book_Title', '')
            norm_excel = normalize_title_for_search(excel_title)
            norm_found = normalize_title_for_search(found_title)
            
            if found_url != "N/A" and (norm_excel in norm_found or norm_found in norm_excel or norm_excel[:15] in norm_found):
                match_found = True
    except Exception as e:
        print(f"    [Error] {excel_title}: {e}")

    # Update Worksheet
    row_idx = index + 2
    header = [cell.value for cell in ws[1]]
    
    def clean_val(val):
        return str(val) if val and str(val).upper() != "N/A" else ""

    mapping = {
        "Goodreads Rating": clean_val(gr_data.get("GoodReads_Rating")),
        "Goodreads No. of Ratings": clean_val(gr_data.get("GoodReads_Rating_Count")),
        "Goodread Link": clean_val(gr_data.get("GoodReads_Book_URL")),
        "Series Link": clean_val(gr_data.get("GoodReads_Series_URL")),
        "# of primary books": clean_val(gr_data.get("Num_Primary_Books")),
        "GR Book 1 Rating": clean_val(gr_data.get("Book1_Rating"))
    }
    
    if match_found:
        for col_name, value in mapping.items():
            if col_name in header:
                col_idx = header.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).value = value
        print(f"    [Success] {excel_title}")
    else:
        # Clear if not found
        cols_to_clear = ["Goodreads Rating", "Goodreads No. of Ratings", "Goodread Link", 
                         "Series Link", "# of primary books", "GR Book 1 Rating"]
        for col_name in cols_to_clear:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).value = ""
        print(f"    [Not Found] {excel_title}")

async def main():
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(root_dir, INPUT_FILE)
    
    print(f"Loading {INPUT_FILE}...")
    df = pd.read_excel(file_path)
    
    last_done_index = load_state()
    start_idx = last_done_index + 1
    end_idx = TARGET_END_ROW - 2 
    
    if start_idx > end_idx:
        print(f"Already finished up to {TARGET_END_ROW}. Check horror_progress.json.")
        return

    print(f"Starting batch process for rows {start_idx+2} to {end_idx+2}...")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    async with async_playwright() as p:
        print("Launching browser...")
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        gr_scraper = GoodreadsScraper(headless=True)
        
        print("Logging in to Goodreads...")
        login_page = await context.new_page()
        await gr_scraper.login_to_goodreads(login_page)
        await login_page.close()
        
        # Process in waves
        for i in range(start_idx, end_idx + 1, BATCH_SIZE):
            current_batch_end = min(i + BATCH_SIZE, end_idx + 1)
            batch_df = df.iloc[i:current_batch_end]
            
            print(f"\n>>> WAVE: Rows {i+2} to {current_batch_end+1}...")
            
            counter = [0]
            tasks = []
            for index, row in batch_df.iterrows():
                tasks.append(process_book(index, row, context, gr_scraper, ws, len(batch_df), counter))
            
            await asyncio.gather(*tasks)
            
            print(f"Wave complete. Saving progress...")
            try:
                wb.save(file_path)
                save_state(current_batch_end - 1)
            except Exception as e:
                print(f"  [Warning] Could not save Excel file: {e}")
            
            # Brief cooldown between waves
            await asyncio.sleep(2)
            
        print("\nAll target rows complete! Opening file...")
        await browser.close()
        
        if os.name == 'nt':
            os.startfile(file_path)

if __name__ == "__main__":
    asyncio.run(main())
