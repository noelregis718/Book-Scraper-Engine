import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import pandas as pd

def format_madwoman(input_file, output_file):
    # Read the existing excel file
    df = pd.read_excel(input_file)
    
    # -----------------------------------------------------------
    # 2. Build data in the 11-column format
    # -----------------------------------------------------------
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    data_rows = []
    
    # Iterate through the DataFrame rows
    for index, row in df.iterrows():
        book_title = row.get('Name of Series', '')
        author = row.get('Author Name', '')
        
        # If both title and author are empty, skip
        if pd.isna(book_title) and pd.isna(author):
            continue
            
        book_title_str = str(book_title).strip() if pd.notna(book_title) else ""
        author_str = str(author).strip() if pd.notna(author) else ""
        
        new_row = [
            book_title_str,            # Name of Series 
            author_str,                # Author Name
            row.get('Publisher', 'Mad Woman Literary Agent'),
            row.get('GoodReads series link', ''),
            row.get('Number of PRIMARY books in the series', ''),
            row.get('Rating (out of 5) of Primary Book 1', ''),
            row.get('Ratings (#) of Primary Book 1', ''),
            row.get('Synopsis (if available)', ''),
            row.get('Romantasy = Yes or No?', ''),
            row.get('Romantasy Sub-Genre of series', ''),
            row.get('Name of agent', 'Alex Land'),
        ]
        data_rows.append(new_row)

    # -----------------------------------------------------------
    # 3. Write to new workbook with styling
    # -----------------------------------------------------------
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Mad Woman Literary"

    # Style constants
    header_fill   = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style='thin', color="CCCCCC")
    thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alternating row fills
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",  fill_type="solid")

    # Write header row
    ws_out.append(new_columns)

    # Write data rows
    for data_row in data_rows:
        ws_out.append(data_row)

    # Apply header styling
    for col_idx in range(1, len(new_columns) + 1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = align_center
        cell.border  = thin_border

    # Freeze header row
    ws_out.freeze_panes = "A2"

    # Apply data row styling
    for row_idx in range(2, ws_out.max_row + 1):
        row_fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx in range(1, len(new_columns) + 1):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            cell.fill      = row_fill
            cell.alignment = align_top
            cell.border    = thin_border

    # Column widths
    col_widths = {
        "A": 35,   # Name of Series
        "B": 28,   # Author Name
        "C": 22,   # Publisher
        "D": 40,   # GoodReads series link
        "E": 18,   # Number of PRIMARY books
        "F": 18,   # Rating (out of 5)
        "G": 16,   # Ratings (#)
        "H": 55,   # Synopsis
        "I": 16,   # Romantasy Y/N
        "J": 28,   # Romantasy Sub-Genre
        "K": 25,   # Name of agent
    }
    for col_letter, width in col_widths.items():
        ws_out.column_dimensions[col_letter].width = width

    # Row height for header
    ws_out.row_dimensions[1].height = 40

    wb_out.save(output_file)
    print(f"Done! {len(data_rows)} rows written -> {output_file}")


if __name__ == "__main__":
    base        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file  = os.path.join(base, "madwoman_literary_scraped_books.xlsx")
    output_file = os.path.join(base, "madwoman_literary_scraped_books.xlsx")
    format_madwoman(input_file, output_file)
