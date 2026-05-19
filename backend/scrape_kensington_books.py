import asyncio
import os
import sys
import re
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from playwright.async_api import async_playwright

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper
from romantasy_analyzer import identify_romantasy_subgenre

CATALOG_FILE = r"E:\Internship\PocketFM\Kensington_Master_Merged_Catalog.xlsx"
HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def duplicate_row(ws, source_row_idx, target_row_idx):
    ws.insert_rows(target_row_idx)
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)
        
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height

async def process_author(idx, author_name, context, gr_scraper, author_books_data, sem):
    async with sem:
        print(f"[{idx}] Processing Author: '{author_name}'...")
        search_page = await context.new_page()
        books_list = []
        try:
            # Find author page and get strictly 1 book
            books_list = await gr_scraper.search_author_books_with_links(search_page, author_name, max_books=1)
            
            # Resilient Fallback
            if len(books_list) < 1:
                search_url = f"https://www.goodreads.com/search?q={author_name.replace(' ', '+')}"
                await search_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(0.5)
                
                search_book_els = await search_page.query_selector_all('a.bookTitle, [data-testid="bookTitle"] a')
                for el in search_book_els:
                    title = (await el.inner_text()).strip()
                    link = await el.evaluate("el => el.href")
                    if link:
                        link = link.split('?')[0] # Strip tracking parameters
                        if link not in [b['link'].split('?')[0] for b in books_list]:
                            books_list.append({'title': title, 'link': link})
                    if len(books_list) >= 1:
                        break
        except Exception as e:
            print(f"  [Error] Failed finding books for '{author_name}': {e}")
        finally:
            await search_page.close()
            
        books_list = books_list[:1]
        print(f"  Found {len(books_list)} proper books for '{author_name}'.", flush=True)

        async def scrape_detail(b_idx, b):
            book_page = await context.new_page()
            try:
                await book_page.goto(b['link'], wait_until="domcontentloaded", timeout=45000)
                details = await gr_scraper.extract_book_details(book_page)
                if details:
                    print(f"    [Success] '{author_name}' - '{b['title']}' Rating: {details.get('GoodReads_Rating')}")
                    return details
            except Exception as e:
                pass
            finally:
                if not book_page.is_closed():
                    await book_page.close()
            return None

        # Scrape book details in parallel
        detail_tasks = [scrape_detail(b_idx, b) for b_idx, b in enumerate(books_list, 1)]
        results = await asyncio.gather(*detail_tasks)
        scraped_details = [d for d in results if d]
        
        author_books_data[author_name] = scraped_details

async def scrape_kensington_books():
    print(">>> Starting AGGRESSIVE Kensington Goodreads Book Enrichment pipeline...")
    
    if not os.path.exists(CATALOG_FILE):
        return
        
    wb = openpyxl.load_workbook(CATALOG_FILE)
    ws = wb.active
    
    authors_to_process = []
    seen_authors = set()
    for r in range(2, ws.max_row + 1):
        author_name = ws.cell(row=r, column=2).value
        gr_link = ws.cell(row=r, column=4).value
        
        if author_name:
            author_str = str(author_name).strip()
            # If gr_link is empty or N/A, this author hasn't been scraped yet
            if author_str not in seen_authors and (not gr_link or str(gr_link).strip() == "N/A"):
                authors_to_process.append((r, author_str))
                seen_authors.add(author_str)
            else:
                seen_authors.add(author_str) # Mark as seen even if scraped to skip duplicates
                
        if len(authors_to_process) >= 100:
            break
            
    print(f"Loaded {len(authors_to_process)} unique unprocessed authors from catalog:")
    for r, name in authors_to_process:
        print(f"  Row {r}: '{name}'")
        
    async with async_playwright() as p:
        print("[System] Launching Playwright browser in headed mode...")
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        gr_scraper = GoodreadsScraper(headless=False)
        author_books_data = {}
        sem = asyncio.Semaphore(5) # Up to 5 authors concurrently
        
        # Process in waves of 5 to run 5 authors simultaneously
        chunk_size = 5
        for i in range(0, len(authors_to_process), chunk_size):
            chunk = authors_to_process[i:i + chunk_size]
            tasks = [process_author(idx, name, context, gr_scraper, author_books_data, sem) for idx, (_, name) in enumerate(chunk, i + 1)]
            await asyncio.gather(*tasks)
            await asyncio.sleep(1) # Breath between waves
            
        await browser.close()
        
    print("\n>>> Merging and writing scraped details back to Excel...")
    wb = openpyxl.load_workbook(CATALOG_FILE)
    ws = wb.active
    
    for r, author_name in reversed(authors_to_process):
        author_str = str(author_name).strip()
        books = author_books_data.get(author_str, [])
        
        if not books:
            ws.cell(row=r, column=1, value="N/A")
            ws.cell(row=r, column=3, value="Kensington")
            ws.cell(row=r, column=4, value="N/A")
            ws.cell(row=r, column=5, value=1)
            ws.cell(row=r, column=9, value="No")
            continue
            
        first_book = books[0]
        title_1 = first_book.get("Book_Title", "N/A")
        series_name_1 = re.sub(r'\(.*?\)|\[.*?\]', '', title_1).strip()
        
        synopsis_1 = first_book.get("Description", "N/A")
        rom_sub_1 = identify_romantasy_subgenre(synopsis_1, first_book.get("Genre", ""))
        
        series_link_1 = first_book.get("GoodReads_Series_URL")
        book_link_1 = first_book.get("GoodReads_Book_URL")
        final_link_1 = series_link_1 if series_link_1 and series_link_1 != "N/A" else (book_link_1 or "N/A")
        
        ws.cell(row=r, column=1, value=series_name_1)
        ws.cell(row=r, column=3, value="Kensington")
        ws.cell(row=r, column=4, value=final_link_1)
        ws.cell(row=r, column=5, value=1)
        ws.cell(row=r, column=6, value=first_book.get("GoodReads_Rating", "N/A"))
        ws.cell(row=r, column=7, value=first_book.get("GoodReads_Rating_Count", "N/A"))
        ws.cell(row=r, column=8, value=synopsis_1)
        ws.cell(row=r, column=9, value="Yes" if rom_sub_1 != "N/A" else "No")
        ws.cell(row=r, column=10, value=rom_sub_1)
        
        for b_idx in range(1, len(books)):
            target_row = r + b_idx
            duplicate_row(ws, r, target_row)
            nth_book = books[b_idx]
            title_n = nth_book.get("Book_Title", "N/A")
            series_name_n = re.sub(r'\(.*?\)|\[.*?\]', '', title_n).strip()
            synopsis_n = nth_book.get("Description", "N/A")
            rom_sub_n = identify_romantasy_subgenre(synopsis_n, nth_book.get("Genre", ""))
            
            series_link_n = nth_book.get("GoodReads_Series_URL")
            book_link_n = nth_book.get("GoodReads_Book_URL")
            final_link_n = series_link_n if series_link_n and series_link_n != "N/A" else (book_link_n or "N/A")
            
            ws.cell(row=target_row, column=1, value=series_name_n)
            ws.cell(row=target_row, column=4, value=final_link_n)
            ws.cell(row=target_row, column=6, value=nth_book.get("GoodReads_Rating", "N/A"))
            ws.cell(row=target_row, column=7, value=nth_book.get("GoodReads_Rating_Count", "N/A"))
            ws.cell(row=target_row, column=8, value=synopsis_n)
            ws.cell(row=target_row, column=9, value="Yes" if rom_sub_n != "N/A" else "No")
            ws.cell(row=target_row, column=10, value=rom_sub_n)

    format_excel_sheet(ws)
    wb.save(CATALOG_FILE)
    print(f"\n>>> Success! Kensington catalog aggressively enriched for the next {len(authors_to_process)} authors.")
    
    # Auto-open the Excel file
    if os.name == 'nt':
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    asyncio.run(scrape_kensington_books())
