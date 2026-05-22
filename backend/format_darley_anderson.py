import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def format_darley_anderson(input_file, output_file):
    df = pd.read_excel(input_file)
    
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]

    data_rows = []
    # Columns in original: ['Author', 'Book Title', 'Source URL', 'Scrape Basis']
    for index, row in df.iterrows():
        book_title = str(row.get("Book Title", "")).strip()
        author = str(row.get("Author", "")).strip()
        
        if book_title == "" or book_title == "nan":
            continue
            
        if author == "nan":
            author = ""

        agent_val = "Darley Anderson"

        new_row = [
            book_title,                # Name of Series
            author,                    # Author Name
            "",                        # Publisher
            "",                        # GoodReads series link
            "",                        # Number of PRIMARY books
            "",                        # Rating (out of 5)
            "",                        # Ratings (#)
            "",                        # Synopsis
            "",                        # Romantasy = Yes or No?
            "",                        # Romantasy Sub-Genre
            agent_val,                 # Name of agent
        ]
        data_rows.append(new_row)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Darley Anderson"

    # Style constants
    header_fill   = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style='thin', color="CCCCCC")
    thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_even = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",  fill_type="solid")

    ws_out.append(new_columns)

    for data_row in data_rows:
        ws_out.append(data_row)

    for col_idx in range(1, len(new_columns) + 1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.fill    = header_fill
        cell.font    = header_font
        cell.alignment = align_center
        cell.border  = thin_border

    ws_out.freeze_panes = "A2"

    for row_idx in range(2, ws_out.max_row + 1):
        row_fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx in range(1, len(new_columns) + 1):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            cell.fill      = row_fill
            cell.alignment = align_top
            cell.border    = thin_border

    col_widths = {
        "A": 35,   # Name of Series
        "B": 28,   # Author Name
        "C": 22,   # Publisher
        "D": 40,   # GoodReads series link
        "E": 18,   # Number of PRIMARY books
        "F": 18,   # Rating (out of 5)
        "G": 16,   # Ratings (#)
        "H": 55,   # Synopsis
        "I": 16,   # Romantasy Y/N
        "J": 28,   # Romantasy Sub-Genre
        "K": 25,   # Name of agent
    }
    for col_letter, width in col_widths.items():
        ws_out.column_dimensions[col_letter].width = width

    ws_out.row_dimensions[1].height = 40

    wb_out.save(output_file)
    print(f"Done! {len(data_rows)} rows written -> {output_file}")


if __name__ == "__main__":
    base        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file  = os.path.join(base, "darley_anderson_books_final_updated.xlsx")
    output_file = os.path.join(base, "Darley_Anderson_Formatted.xlsx")
    format_darley_anderson(input_file, output_file)
