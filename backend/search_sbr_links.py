import asyncio
import pandas as pd
import os
import sys
import random
from playwright.async_api import async_playwright
from openpyxl import load_workbook

# Add current directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from goodreads_scraper import GoodreadsScraper, normalize_title_for_search

# Configuration
INPUT_FILE = "SBR Media.xlsx"
OUTPUT_FILE = "SBR Media.xlsx"
BATCH_SIZE = 10 

async def process_book(index, row, context, gr_scraper, ws):
    series_name = str(row.get("Name of Series", "Unknown"))
    author_name = str(row.get("Author Name", ""))
    
    # Use Series Name + Author for SBR (since it's a series-focused list)
    search_query = f"{series_name} {author_name}"
    print(f"  [SBR] Row {index+2}: {series_name} by {author_name}...")
    
    match_found = False
    gr_data = {}
    
    try:
        await asyncio.sleep(random.uniform(0.1, 0.5))
        # Search for series mainly
        gr_data = await gr_scraper.scrape_goodreads_data(context, series_name, author_name)
        
        if gr_data:
            found_url = gr_data.get('GoodReads_Series_URL', 'N/A')
            if found_url != "N/A":
                match_found = True
    except Exception as e:
        print(f"    [Error] {e}")

    # Update Worksheet
    row_idx = index + 2
    header = [cell.value for cell in ws[1]]
    
    def clean_val(val):
        return str(val) if val and str(val).upper() != "N/A" else ""

    mapping = {
        "GoodReads series link": clean_val(gr_data.get("GoodReads_Series_URL")),
        "Number of PRIMARY books in the series": clean_val(gr_data.get("Num_Primary_Books")),
        "Rating (out of 5) of Primary Book 1": clean_val(gr_data.get("Book1_Rating")),
        "Ratings (#) of Primary Book 1": clean_val(gr_data.get("Book1_Num_Ratings")),
        "Synopsis (if available)": clean_val(gr_data.get("Book_Synopsis"))
    }
    
    if match_found:
        for col_name, value in mapping.items():
            if col_name in header:
                col_idx = header.index(col_name) + 1
                # Only fill if empty to avoid overwriting existing good data
                existing = ws.cell(row=row_idx, column=col_idx).value
                if not existing:
                    ws.cell(row=row_idx, column=col_idx).value = value
        print(f"    [Success] Updated.")
    else:
        print(f"    [Skip] Not found.")

async def main():
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(root_dir, INPUT_FILE)
    
    print(f"Loading {INPUT_FILE}...")
    df = pd.read_excel(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find empty rows in 'GoodReads series link'
    target_indices = []
    col_name = "GoodReads series link"
    for idx, row in df.iterrows():
        val = str(row.get(col_name, ""))
        if not val or val.lower() == "nan" or val.lower() == "n/a":
            target_indices.append(idx)
    
    if not target_indices:
        print("All rows already filled!")
        return

    print(f"Targeting {len(target_indices)} empty rows for SBR Media enrichment...")
    
    async with async_playwright() as p:
        print("Launching browser...")
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        gr_scraper = GoodreadsScraper(headless=False)
        
        print("Logging in to Goodreads...")
        login_page = await context.new_page()
        await gr_scraper.login_to_goodreads(login_page)
        await login_page.close()
        
        # Process in batches
        for i in range(0, len(target_indices), BATCH_SIZE):
            chunk = target_indices[i:i + BATCH_SIZE]
            print(f"\n>>> SBR WAVE: Processing {len(chunk)} rows...")
            
            tasks = []
            for idx in chunk:
                row = df.iloc[idx]
                tasks.append(process_book(idx, row, context, gr_scraper, ws))
            
            await asyncio.gather(*tasks)
            
            print(f"Wave complete. Saving progress...")
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"  [Warning] Could not save Excel file: {e}")
            
            await asyncio.sleep(1)
            
        print("\nSBR Media enrichment complete! Opening file...")
        await browser.close()
        
        if os.name == 'nt':
            os.startfile(file_path)

if __name__ == "__main__":
    asyncio.run(main())
