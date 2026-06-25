import openpyxl

def fix_excel():
    filename = 'LDLA_Combined_Mapped_Fixed.xlsx'
    print(f"Loading {filename}...")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Check max row
    max_row = ws.max_row
    print(f"Processing {max_row} rows...")
    
    # Move values for each row
    for r in range(2, max_row + 1):
        # Read from outside columns
        romantasy_val = ws.cell(row=r, column=12).value
        subgenre_val = ws.cell(row=r, column=13).value
        
        # If there's something to move, move it
        if romantasy_val is not None:
            ws.cell(row=r, column=9).value = romantasy_val
            ws.cell(row=r, column=12).value = None
            
        if subgenre_val is not None:
            ws.cell(row=r, column=10).value = subgenre_val
            ws.cell(row=r, column=13).value = None

    # Clear headers for 12 and 13
    ws.cell(row=1, column=12).value = None
    ws.cell(row=1, column=13).value = None
    
    # Save the modified workbook
    print("Saving the modified workbook...")
    wb.save(filename)
    print("Done!")

if __name__ == "__main__":
    fix_excel()
