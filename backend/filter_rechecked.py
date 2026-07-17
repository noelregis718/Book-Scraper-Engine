from openpyxl import load_workbook

file_path = r'e:\Internship\PocketFM\Small_Publishing_Agencies_Imprint_Status_Rechecked.xlsx'

wb = load_workbook(file_path)
ws = wb.active

header = [cell.value for cell in ws[1]]
try:
    imprint_col_idx = header.index('Imprint? (Rechecked)') + 1
except ValueError:
    print("Could not find column 'Imprint? (Rechecked)'")
    exit(1)

# Delete from bottom to top to preserve row indices during deletion
deleted_count = 0
for row_idx in range(ws.max_row, 1, -1):
    cell_val = ws.cell(row=row_idx, column=imprint_col_idx).value
    if str(cell_val).strip() != 'No imprint':
        ws.delete_rows(row_idx, 1)
        deleted_count += 1

wb.save(file_path)
print(f"Deleted {deleted_count} imprint rows while preserving formatting.")
