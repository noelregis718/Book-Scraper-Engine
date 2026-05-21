import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os


MAIN_SHEET = "Mushens Entertainment"

NEW_COLUMNS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent",
]


def _read_data_rows(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        book_title = row[0] if len(row) > 0 else None
        author     = row[1] if len(row) > 1 else None
        if book_title is None or str(book_title).strip() == "":
            continue
        rows.append([
            str(book_title).strip(),
            str(author).strip() if author else "",
            "Mushens Entertainment",
            row[3] if len(row) > 3 and row[3] is not None else "",
            row[4] if len(row) > 4 and row[4] is not None else "",
            row[5] if len(row) > 5 and row[5] is not None else "",
            row[6] if len(row) > 6 and row[6] is not None else "",
            row[7] if len(row) > 7 and row[7] is not None else "",
            row[8] if len(row) > 8 and row[8] is not None else "",
            row[9] if len(row) > 9 and row[9] is not None else "",
            "Juliet Mushens",
        ])
    return rows


def _apply_styling(ws):
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx in range(1, len(NEW_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(NEW_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill_white
            cell.alignment = align_top
            cell.border = thin_border

    col_widths = {
        "A": 35, "B": 28, "C": 22, "D": 40, "E": 18,
        "F": 18, "G": 16, "H": 55, "I": 16, "J": 28, "K": 25,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 40


def format_mushens(input_file, output_file):
    """
    Re-applies the 11-column white-row styling to the Mushens Entertainment
    sheet IN PLACE — every other sheet in the workbook (e.g. 'All Authors')
    is preserved untouched.
    """
    wb = openpyxl.load_workbook(input_file)

    # Pick the main sheet (by name if present, else first sheet)
    if MAIN_SHEET in wb.sheetnames:
        ws_old = wb[MAIN_SHEET]
        position = wb.sheetnames.index(MAIN_SHEET)
    else:
        ws_old = wb.worksheets[0]
        position = 0

    data_rows = _read_data_rows(ws_old)

    # Replace the main sheet at the same position
    old_title = ws_old.title
    del wb[old_title]
    ws_new = wb.create_sheet(MAIN_SHEET, position)

    ws_new.append(NEW_COLUMNS)
    for r in data_rows:
        ws_new.append(r)

    _apply_styling(ws_new)

    wb.save(output_file)
    print(f"Done! {len(data_rows)} rows written -> {output_file}")


if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file  = os.path.join(base, "Mushens_Entertainment_Bestsellers.xlsx")
    output_file = os.path.join(base, "Mushens_Entertainment_Bestsellers.xlsx")
    format_mushens(input_file, output_file)
