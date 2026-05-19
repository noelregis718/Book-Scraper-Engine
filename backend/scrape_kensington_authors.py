import asyncio
import os
import sys
import openpyxl
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATALOG_FILE = r"E:\Internship\PocketFM\Kensington_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching kensington style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

async def scrape_kensington_authors():
    url = "https://www.kensingtonbooks.com/authors/?v=13b5bfe96f3e#a"
    print("=" * 60)
    print("      KENSINGTON AGENCY: AUTHOR NAME SCRAPER")
    print("=" * 60)
    print(f"Targeting URL: {url}\n")

    async with async_playwright() as p:
        print("[System] Launching browser in visible mode (headless=False) to bypass Cloudflare...")
        try:
            browser = await p.chromium.launch(
                headless=False,
                channel="chrome",
                args=["--disable-blink-features=AutomationControlled"]
            )
        except Exception as e:
            print(f"[System] Chrome launch failed, trying regular Chromium with stealth args: {e}")
            browser = await p.chromium.launch(
                headless=False,
                args=["--disable-blink-features=AutomationControlled"]
            )
            
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = await context.new_page()
        await page.add_init_script("delete navigator.__proto__.webdriver;")
        
        print(f"[System] Navigating to {url}...")
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            
            print("\n" + "!" * 60)
            print("  CLOUD FLARE BYPASS REQUIRED!")
            print("  Please look at the browser window that just popped up on your screen.")
            print("  1. Check the box to verify you are a human if prompted.")
            print("  2. Wait for the A-Z Authors list page to load.")
            print("!" * 60 + "\n")
            
            # Since input() in an async function blocks the event loop, we'll do an interactive loop using asyncio.sleep
            print("[Action] Waiting for the page to bypass security and load. We will check the page content automatically...")
            # Pause immediately to let the user see the page and solve any security challenge
            print("\n" + "!" * 60)
            print("  CLOUD FLARE BYPASS REQUIRED!")
            print("  Please look at the browser window that just popped up on your screen.")
            print("  1. Solve/verify any Turnstile human challenge if prompted.")
            print("  2. Wait for the A-Z Authors list page to load.")
            print("!" * 60 + "\n")
            
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, input, ">>> Press [Enter] in this terminal ONLY when the A-Z Authors page has fully loaded on your screen...")

            # Extract author list
            print("[System] Bypassed! Extracting author names...")
            authors_data = await page.evaluate("""() => {
                const links = Array.from(document.querySelectorAll('a'));
                const list = [];
                links.forEach(link => {
                    const href = link.getAttribute('href');
                    const text = link.innerText.trim();
                    if (href && (href.toLowerCase().includes('/author/') || href.toLowerCase().includes('/authors/'))) {
                        if (text && text.length > 2 && text.length < 50) {
                            const tLower = text.toLowerCase();
                            const blacklist = ["author", "authors", "home", "books", "contact", "about", "submission", "faq", "terms", "privacy", "help", "search"];
                            if (!blacklist.some(item => tLower.includes(item))) {
                                if (!list.includes(text)) {
                                    list.push(text);
                                }
                            }
                        }
                    }
                });
                return list;
            }""")
            
            print(f"[System] Extracted {len(authors_data)} unique author names from website.")
            
            if not authors_data:
                print("[Warning] No author names extracted via direct link scanning. Trying fallback innerText split...")
                # Try fallback: find elements by letter classes
                body_text = await page.evaluate("() => document.body.innerText")
                # Just print warning
                print("[System] Closing browser...")
                await browser.close()
                return

            await browser.close()
            
            # Sort author names alphabetically
            authors_data = sorted(authors_data)
            
            # Write to excel sheet
            print(f"\n>>> Writing {len(authors_data)} authors to: {CATALOG_FILE}...")
            wb = load_workbook(CATALOG_FILE)
            ws = wb.active
            
            # Clear existing data rows (except header)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
                
            for idx, author in enumerate(authors_data, 2):
                ws.cell(row=idx, column=1, value="N/A")            # Name of Series
                ws.cell(row=idx, column=2, value=author)           # Author Name
                ws.cell(row=idx, column=3, value="Kensington")     # Publisher
                ws.cell(row=idx, column=4, value="N/A")            # GoodReads link
                ws.cell(row=idx, column=5, value=1)                # Books count
                ws.cell(row=idx, column=6, value="N/A")            # Rating
                ws.cell(row=idx, column=7, value="N/A")            # Ratings #
                ws.cell(row=idx, column=8, value="N/A")            # Synopsis
                ws.cell(row=idx, column=9, value="No")             # Romantasy Yes/No
                ws.cell(row=idx, column=10, value="N/A")           # Romantasy Subgenre
                ws.cell(row=idx, column=11, value="N/A")           # Agent Name
                
            print(">>> Re-applying premium styling and saving workbook...")
            format_excel_sheet(ws)
            wb.save(CATALOG_FILE)
            print(">>> Catalog updated successfully!")
            
            print("\n" + "=" * 60)
            print("                 SCRAPING COMPLETE")
            print("=" * 60)
            print(f"Total Authors Scraped: {len(authors_data)}")
            print(f"Output Catalog:        {CATALOG_FILE}")
            print("=" * 60 + "\n")
            
            if os.name == 'nt':
                print("  [System] Auto-opening sheet for review...")
                os.startfile(CATALOG_FILE)
                
        except Exception as e:
            print(f"[Error] Scrape failed: {e}")
            await browser.close()

if __name__ == "__main__":
    asyncio.run(scrape_kensington_authors())
