import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

FILE_1488 = r"E:\Internship\PocketFM\kensington_authors_MASTER_1_to_1488.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching kensington style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def format_1488_catalog():
    print(f">>> Reading 1488 File: {FILE_1488}")
    # Read the file and use the second row (index 1) as the header
    df = pd.read_excel(FILE_1488, header=1)
    print(f"    Loaded {len(df)} raw rows.")

    # Drop the grouping rows which have no book title
    df = df.dropna(subset=['Book Title'])
    print(f"    Found {len(df)} actual books.")

    # Create the new dataframe with mapped columns
    mapped_df = pd.DataFrame()
    mapped_df['Name of Series'] = df['Book Title']
    mapped_df['Author Name'] = df['Author Name']
    mapped_df['Publisher'] = "Kensington"
    
    # Fill the remaining columns with N/A
    for h in HEADERS:
        if h not in mapped_df.columns:
            mapped_df[h] = "N/A"
            
    # Enforce exact 11-column order
    mapped_df = mapped_df[HEADERS]
    
    print(f">>> Total Books mapped: {len(mapped_df)}")
    
    print(f">>> Saving raw mapped data back to {FILE_1488}...")
    mapped_df.to_excel(FILE_1488, index=False)
    
    print(">>> Applying premium 11-column formatting...")
    wb = openpyxl.load_workbook(FILE_1488)
    ws = wb.active
    
    format_excel_sheet(ws)
    
    wb.save(FILE_1488)
    print(">>> Formatting complete!")
    
    if os.name == 'nt':
        os.startfile(FILE_1488)

if __name__ == "__main__":
    format_1488_catalog()
