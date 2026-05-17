import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MASTER_FILE = "Deep_Catalog_Enrichment.xlsx"

def format_excel():
    if not os.path.exists(MASTER_FILE):
        print(f"Error: {MASTER_FILE} not found.")
        return

    print(f"Applying premium styling to {MASTER_FILE}...")
    
    # Load with openpyxl
    wb = load_workbook(MASTER_FILE)
    ws = wb.active
    if ws is None:
        raise ValueError("Active worksheet is None")

    # Define Styles
    header_font = Font(name='Outfit', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    body_font = Font(name='Outfit', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="D9D9D9"),
        right=Side(style='thin', color="D9D9D9"),
        top=Side(style='thin', color="D9D9D9"),
        bottom=Side(style='thin', color="D9D9D9")
    )

    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Column Widths
    widths = {
        'A': 40, # Name of Series
        'B': 25, # Author Name
        'C': 20, # Publisher
        'D': 35, # GoodReads series link
        'E': 15, # Number of PRIMARY books
        'F': 12, # Rating
        'G': 15, # Ratings count
        'H': 60, # Synopsis
        'I': 15, # Romantasy Yes/No
        'J': 30, # Romantasy Sub-genre
        'K': 20  # Agent
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Format Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Format Body
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = thin_border
            
            # Alignment based on column
            if get_column_letter(cell.column or 1) in ['E', 'F', 'G', 'I']:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            
            # All white background
            cell.fill = white_fill

    # Freeze Header
    ws.freeze_panes = "A2"

    print(f"Saving formatted file to {MASTER_FILE}...")
    wb.save(MASTER_FILE)
    print("Done! The catalog is now neat, clean, and professional.")

if __name__ == "__main__":
    format_excel()
