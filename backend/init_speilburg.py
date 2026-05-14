import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
CATALOG_FILE = "Speilburg_Media_Catalog_Final.xlsx"
COLS = [
    'Name of Series', 'Author Name', 'Publisher', 'GoodReads series link',
    'Number of Primary books in series', 'Average Rating of Series',
    'Ratings count of the first book of the series', 'Synopsis (if available)',
    'Romantasy = Yes or No?', 'Romantasy Sub-Genre of series', 'Agent'
]

def init_speilburg():
    print(f"Initializing {CATALOG_FILE}...")
    
    # Create DataFrame and Save
    df = pd.DataFrame(columns=COLS)
    df.to_excel(CATALOG_FILE, index=False)
    
    # Load and Format
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active
    
    # Styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    thin_side = Side(border_style='thin', color='000000')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Header Styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Column Widths
    widths = {1: 35, 2: 25, 3: 15, 4: 40, 5: 15, 6: 12, 7: 15, 8: 70, 9: 15, 10: 40, 11: 15}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(CATALOG_FILE)
    print(f"Success! {CATALOG_FILE} is ready and formatted.")
    if os.name == 'nt':
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    init_speilburg()
