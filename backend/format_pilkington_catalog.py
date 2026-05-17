import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

CATALOG_FILE = "Pilkington_Agency_Catalog_Final.xlsx"

def format_catalog():
    if not os.path.exists(CATALOG_FILE):
        print(f"Error: {CATALOG_FILE} not found!")
        return

    # Load the workbook
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active
    if ws is None:
        raise ValueError("Active worksheet is None")

    # Define Styles
    header_font = Font(name='Calibri', size=12, bold=True, color='000000')
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') # Light Grey
    
    body_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # 1. Format Headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 2. Format Body
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = thin_border
            
            # Specific Alignments
            col_letter = get_column_letter(int(cell.column))
            header_val = str(ws[f"{col_letter}1"].value)
            
            if any(k in header_val for k in ['Rating', 'Number', 'Ratings (#)', 'Yes or No?']):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # 3. Adjust Column Widths
    column_widths = {
        'A': 35, # Name of Series
        'B': 25, # Author Name
        'C': 20, # Publisher
        'D': 40, # GoodReads series link
        'E': 15, # Number of PRIMARY
        'F': 15, # Rating
        'G': 15, # Ratings (#)
        'H': 60, # Synopsis
        'I': 15, # Romantasy Yes/No
        'J': 35, # Sub-Genre
        'K': 20, # Name of Agent
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 4. Freeze Top Row
    ws.freeze_panes = "A2"

    wb.save(CATALOG_FILE)
    print(f"Successfully formatted {CATALOG_FILE} with premium agency styling.")
    os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    format_catalog()
