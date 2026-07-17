from openpyxl import load_workbook
import os

file_path = r'e:\Internship\PocketFM\Small_Mid_Sized_Publishers_Crime_Series_Expanded_30_Per_Publisher.xlsx'

wb = load_workbook(file_path)
ws = wb.active

header = [cell.value for cell in ws[1]]
try:
    books_col_idx = header.index('Number of Primary Books') + 1
except ValueError:
    print("Could not find column 'Number of Primary Books'")
    exit(1)

# Delete from bottom to top to preserve row indices during deletion
deleted_count = 0
for row_idx in range(ws.max_row, 1, -1):
    cell_val = ws.cell(row=row_idx, column=books_col_idx).value
    # Convert to string and handle floats like 1.0 or strings like "1"
    val_str = str(cell_val).strip()
    if val_str == '1' or val_str == '1.0':
        ws.delete_rows(row_idx, 1)
        deleted_count += 1

wb.save(file_path)
print(f"Deleted {deleted_count} rows where Number of Primary Books was 1, while preserving formatting.")
