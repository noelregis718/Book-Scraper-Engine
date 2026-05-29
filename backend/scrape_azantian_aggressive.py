import asyncio
import os
import sys
import pandas as pd
from playwright.async_api import async_playwright

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper, clean_text, normalize_title_for_search

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

file_lock = asyncio.Lock()

def apply_azantian_style(file_path):
    df = pd.read_excel(file_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Books"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    columns_count = len(df.columns)
    for col in range(1, columns_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, columns_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)

async def safe_save(df, excel_path):
    async with file_lock:
        try:
            df.to_excel(excel_path, index=False)
        except Exception as e:
            print(f"  [Save Error] {e}")

async def scrape_one_book(index, title, author_for_search, df, excel_path, scraper, shared_context, semaphore):
    async with semaphore:
        page = await shared_context.new_page()
        await page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })

        try:
            norm_title = normalize_title_for_search(title)
            import re
            super_clean_title = re.sub(r'[^a-zA-Z0-9\s]', '', norm_title)
            
            queries = [
                f"{super_clean_title} {author_for_search}".strip(),
                f"{norm_title} {author_for_search}".strip(),
                norm_title.strip(),
                super_clean_title.strip()
            ]
            
            book_url = None
            
            for query in queries:
                if not query: continue
                search_url = f"https://www.goodreads.com/search?q={query.replace(' ', '+')}"
                print(f"[{index+1}] Aggressive Searching: '{query}'")

                await page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(2)

                current_url = page.url
                if "/book/show/" in current_url:
                    book_url = current_url
                    print(f"[{index+1}] Auto-redirected to: {book_url}")
                    break
                else:
                    try:
                        first_link = await page.wait_for_selector(
                            'a.bookTitle, [data-testid="bookTitle"] a, h3 a[href*="/book/show/"]',
                            timeout=3000
                        )
                        book_url = await first_link.evaluate("el => el.href")
                        print(f"[{index+1}] Found via search: {book_url}")
                        break
                    except:
                        pass
            
            if not book_url:
                print(f"[{index+1}] Still no result for '{title}' after aggressive search.")
                return

            await page.goto(book_url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(1.5)

            import json, re

            avg_rating, rating_count = "N/A", "N/A"
            try:
                ld_el = await page.query_selector('script[type="application/ld+json"]')
                if ld_el:
                    ld_data = json.loads(await ld_el.inner_text())
                    if isinstance(ld_data, list): ld_data = ld_data[0]
                    avg_rating = str(ld_data.get('aggregateRating', {}).get('ratingValue', 'N/A'))
                    rating_count = str(ld_data.get('aggregateRating', {}).get('ratingCount', 'N/A'))
            except: pass

            description = "N/A"
            desc_el = await page.query_selector('[data-testid="description"] .Formatted, .readable')
            if desc_el:
                description = clean_text(await desc_el.inner_text())

            series_url = book_url
            series_link_el = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a')
            if series_link_el:
                series_url = await series_link_el.evaluate("el => el.href")

            num_primary = "1"
            book1_rating = avg_rating
            book1_ratings = rating_count
            if "/series/" in series_url:
                try:
                    await page.goto(series_url, wait_until="domcontentloaded", timeout=45000)
                    content = await page.content()
                    m = re.search(r'(\d+)\s+primary\s+works', content, re.IGNORECASE)
                    if m: num_primary = m.group(1)
                    row1 = await page.query_selector('.listWithDividers__item, .seriesWork')
                    if row1:
                        rtxt = (await row1.inner_text()).lower()
                        r_match = re.search(r'([\d.]+)\s+avg\s+rating\s+[—\-]\s+([\d,]+)\s+ratings', rtxt)
                        if r_match:
                            book1_rating = r_match.group(1)
                            book1_ratings = r_match.group(2).replace(',', '')
                except: pass

            df.at[index, "GoodReads series link"] = series_url
            df.at[index, "Number of PRIMARY books in the series"] = num_primary
            df.at[index, "Rating (out of 5) of Primary Book 1"] = book1_rating
            df.at[index, "Ratings (#) of Primary Book 1"] = book1_ratings
            df.at[index, "Synopsis (if available)"] = description

            print(f"[{index+1}] Done aggressive scrape: '{title}'")
            await safe_save(df, excel_path)

        except Exception as e:
            print(f"[{index+1}] Error on '{title}': {e}")
        finally:
            try:
                await page.close()
            except: pass

async def scrape_azantian_aggressive(excel_path):
    print(f"Loading: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"Error: File not found — {excel_path}")
        return

    df = pd.read_excel(excel_path)

    for col in ["GoodReads series link", "Number of PRIMARY books in the series",
                "Rating (out of 5) of Primary Book 1", "Ratings (#) of Primary Book 1",
                "Synopsis (if available)"]:
        if col in df.columns:
            df[col] = df[col].astype(object)

    scraper = GoodreadsScraper(headless=False)
    semaphore = asyncio.Semaphore(5)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)

        login_context = await browser.new_context()
        login_page = await login_context.new_page()
        print("Logging in to Goodreads...")
        await scraper.login_to_goodreads(login_page)
        storage_state = await login_context.storage_state()
        await login_context.close()
        print("Login done. Starting aggressive scrape for books...\n")

        shared_context = await browser.new_context(storage_state=storage_state)
        tasks = []
        for index, row in df.iterrows():
            title = str(row.get("Name of Series", "")).strip()
            author = str(row.get("Author Name", "")).strip()

            if not title or title in ["nan", ""]:
                continue

            # Scrape everything aggressively
            author_for_search = author if author not in ["nan", ""] else ""

            tasks.append(scrape_one_book(index, title, author_for_search, df, excel_path, scraper, shared_context, semaphore))

        if not tasks:
            print("No books found to scrape!")
            await shared_context.close()
            await browser.close()
            return

        await asyncio.gather(*tasks)
        print("\nAggressive scraping complete!")
        try:
            await shared_context.close()
            await browser.close()
        except: pass

    print("Reapplying perfect styling...")
    try:
        apply_azantian_style(excel_path)
    except Exception as e:
        print(f"Error applying style: {e}")
        
    print("Opening Excel file...")
    import subprocess
    subprocess.Popen(["start", excel_path], shell=True)
    print("All done!")

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    target = os.path.join(base, "Azantian_LitAgency_Combined_Formatted.xlsx")
    asyncio.run(scrape_azantian_aggressive(target))
