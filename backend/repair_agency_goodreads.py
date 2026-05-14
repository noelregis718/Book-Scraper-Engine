import asyncio
import os
import sys
import pandas as pd
import re
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURATION ---
EXCEL_FILE = r"E:\Internship\PocketFM\Sandra Djikstra.xlsx"
MAX_CONCURRENT_TABS = 1

# Try to import existing utilities
try:
    from goodreads_scraper import GoodreadsScraper
    from ai_classifier import identify_subgenre
except ImportError:
    # Fallback if imports fail in this context
    class GoodreadsScraper:
        def __init__(self, **kwargs): pass
        async def scrape_goodreads_data(self, *args, **kwargs): return {}
    def identify_subgenre(synopsis, tags): return "N/A"

def style_excel(file_path):
    try:
        if not os.path.exists(file_path): return
        wb = load_workbook(file_path)
        ws = wb.active
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(file_path)
    except: pass

# --- INDUSTRIAL SAVE UTILITY (Amazon Technique) ---
def save_agency_excel(df, filename):
    import time
    from datetime import datetime
    
    def _write(path):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Agency Catalog')
            worksheet = writer.sheets['Agency Catalog']
            worksheet.freeze_panes = "A2"
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            col_widths = {"A": 35, "B": 25, "C": 20, "D": 40, "E": 12, "F": 12, "G": 12, "H": 70, "I": 15, "J": 30, "K": 25}
            for col_idx, col_name in enumerate(df.columns):
                col_letter = chr(65 + col_idx) if col_idx < 26 else chr(64 + (col_idx // 26)) + chr(65 + (col_idx % 26))
                worksheet.column_dimensions[col_letter].width = col_widths.get(col_letter, 20)
                
                for cell in worksheet[col_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    try:
        _write(filename)
        print(f"  [System] Successfully saved to {filename}")
        return filename
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(filename)
        fallback = f"{base}_{timestamp}{ext}"
        print(f"  [!!!] '{filename}' is open. Saving to '{fallback}' instead.")
        _write(fallback)
        return fallback

async def repair_row(index, row, context, gr_scraper, semaphore, df, total_to_repair, counter):
    async with semaphore:
        series_name = str(row.get("Name of Series", "N/A"))
        author_name = str(row.get("Author Name", "N/A"))
        
        counter[0] += 1
        curr = counter[0]
        print(f"[{curr}/{total_to_repair}] Tech-Sync Repair: {series_name}...")
        
        try:
            gr_data = await gr_scraper.scrape_goodreads_data(context, series_name, author_name)
            
            if gr_data and gr_data.get("GoodReads_Series_URL"):
                synopsis = gr_data.get("Description", "N/A")
                tags = [gr_data.get("Genre", ""), gr_data.get("Sub_Genre", "")]
                matched_genre = identify_subgenre(synopsis, tags)
                
                mapped = {
                    "GoodReads series link": gr_data.get("GoodReads_Series_URL", "N/A"),
                    "Number of PRIMARY books in the series": gr_data.get("Num_Primary_Books", "N/A"),
                    "Rating (out of 5) of Primary Book 1": gr_data.get("Book1_Rating", "N/A"),
                    "Ratings (#) of Primary Book 1": gr_data.get("Book1_Num_Ratings", "N/A"),
                    "Synopsis (if available)": synopsis[:1000] if synopsis != "N/A" else "N/A",
                    "Is it Romantasy ?": "Yes" if matched_genre != "N/A" else "No",
                    "Romantasy Sub-Genre of series": matched_genre
                }
                
                # AMAZON TECHNIQUE: Direct DataFrame Update
                for col, val in mapped.items():
                    if col in df.columns:
                        df.at[index, col] = val
                
                # SPECIAL: Update Author if it was Unknown
                if str(df.at[index, "Author Name"]).lower() == "unknown" and gr_data.get("Author_Found"):
                    df.at[index, "Author Name"] = gr_data["Author_Found"]
                    print(f"  -> AUTHOR RECOVERY: Identified as {gr_data['Author_Found']}")
                        
                print(f"  -> SUCCESS [{curr}]: Mapped fields for {series_name}")
            else:
                print(f"  -> FAILED [{curr}]: No Goodreads match for {series_name}")
        except Exception as e:
            print(f"  -> ERROR [{curr}]: {e}")

async def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found.")
        return

    print(f"Loading {EXCEL_FILE} for Amazon-style metadata repair...")
    df = pd.read_excel(EXCEL_FILE)
    
    # INDUSTRIAL FIX: Ensure columns can hold strings
    for col in df.columns:
        df[col] = df[col].astype(object)
    
    # Identify missing rows (Missing metadata OR Unknown authors)
    mask = (df["GoodReads series link"].astype(str).str.lower().isin(["n/a", "nan", "none", ""])) | \
           (df["Author Name"].astype(str).str.lower() == "unknown")
    to_repair = df[mask]
    total_to_repair = len(to_repair)
    
    if total_to_repair == 0:
        print("Coverage is already 100%! No titles need repairing.")
        return

    print(f"Found {total_to_repair} titles missing Goodreads metadata. Starting Tech-Sync repair...")
    sys.stdout.flush()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        
        page = await context.new_page()
        print("\n" + "="*60)
        print("ACTION REQUIRED: AUTOMATING LOGIN TO GOODREADS")
        print("1. Solving initial redirects...")
        print("2. Entering credentials...")
        print("3. IF A CAPTCHA APPEARS, PLEASE SOLVE IT MANUALLY.")
        print("="*60 + "\n")
        
        await page.goto("https://www.goodreads.com/user/sign_in", timeout=60000)
        
        try:
            # If on the initial sign-in page, click "Sign in with email" if necessary
            email_btn = page.locator('a:has-text("Sign in with email")')
            if await email_btn.is_visible():
                await email_btn.click()
                await page.wait_for_load_state("networkidle")
            
            # Use Amazon-style selectors for Goodreads login (they often share the same auth system)
            if await page.locator("#ap_email").is_visible():
                await page.fill("#ap_email", "noel.regis04@gmail.com")
                await page.fill("#ap_password", "Noel@1024")
                await page.click("#signInSubmit")
                print("  [Auto-Login] Credentials submitted. Waiting for dashboard...")
            elif await page.locator('input[name="user[email]"]').is_visible():
                await page.fill('input[name="user[email]"]', "noel.regis04@gmail.com")
                await page.fill('input[name="user[password]"]', "Noel@1024")
                await page.click('input[type="submit"]')
                print("  [Auto-Login] Standard credentials submitted. Waiting for dashboard...")
        except Exception as e:
            print(f"  [Auto-Login] automation failed (expected if already logged in): {e}")

        # Wait for the user to be on the home page or dashboard (indicating successful login)
        login_selectors = ['.header-profile-link', '.topBarUser__profileLink', '[data-testid="profile-link"]', 'a[href*="sign_out"]']
        
        logged_in = False
        for _ in range(60): # Wait up to 5 minutes
            for sel in login_selectors:
                try:
                    if await page.locator(sel).is_visible():
                        logged_in = True
                        break
                except: pass
            if logged_in: break
            await asyncio.sleep(5)
            
        if logged_in:
            print("  [System] Login detected! Proceeding with repair...")
        else:
            print("  [Warning] Login not detected within 5 minutes. Attempting to proceed anyway...")
        
        await page.close()
        
        gr_scraper = GoodreadsScraper(headless=False)
        semaphore = asyncio.Semaphore(MAX_CONCURRENT_TABS)
        counter = [0]
        
        # Amazon Technique: Process all tasks, df is updated directly in memory
        tasks = [repair_row(index, row, context, gr_scraper, semaphore, df, total_to_repair, counter) for index, row in to_repair.iterrows()]
        await asyncio.gather(*tasks)
        
        print("\nRepair phase complete. Saving with Amazon-standard utility...")
        final_file = save_agency_excel(df, EXCEL_FILE)
        
        await browser.close()
        print(f"\n{'='*60}")
        print(f"MISSION ACCOMPLISHED: Final metadata sweep complete.")
        print(f"Final Data: {final_file}")
        print(f"{'='*60}\n")
        
        if os.name == 'nt': os.startfile(final_file)

if __name__ == "__main__":
    asyncio.run(main())
