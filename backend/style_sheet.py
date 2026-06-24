import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def apply_styling():
    filename = "Sheet 1 .xlsx"
    print(f"Loading {filename}...")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Define clean styles
    # A nice light blue background for the headers
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    # Thin borders for all cells
    border_style = Side(border_style="thin", color="D4D4D4") # Light grey border for a cleaner look
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Text alignment
    data_alignment = Alignment(vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    print("Applying header styling...")
    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = header_alignment

    print("Applying borders and alignment to data rows...")
    # Style all data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = data_alignment

    print("Adjusting column widths...")
    # Auto-adjust column widths (with a max limit so paragraphs don't make columns giant)
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        
        # Check the first 100 rows to estimate width to save processing time
        for cell in col[:100]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = max_length + 2
        if adjusted_width > 45:
            adjusted_width = 45  # Cap the width for text-heavy columns
        elif adjusted_width < 12:
            adjusted_width = 12  # Minimum width
            
        ws.column_dimensions[column_letter].width = adjusted_width

    print("Freezing the top row...")
    # Freeze the top header row so it stays visible when scrolling down
    ws.freeze_panes = 'A2'

    print(f"Saving styled workbook to {filename}...")
    wb.save(filename)
    print("Styling complete!")

if __name__ == "__main__":
    apply_styling()
