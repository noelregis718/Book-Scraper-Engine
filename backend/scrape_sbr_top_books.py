import asyncio
import os
import sys
import pandas as pd
from playwright.async_api import async_playwright
import traceback

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

EXCEL_FILE = r"E:\Internship\PocketFM\SBR Media.xlsx"
MAX_CONCURRENT = 8

def map_genres(tags):
    tags = [str(t).lower() for t in tags if str(t)]
    is_romantasy = any('romantasy' in t for t in tags)
    is_fantasy = any('fantasy' in t for t in tags)
    is_romance = any('romance' in t for t in tags)
    is_crime = any('crime' in t or 'thriller' in t or 'mystery' in t for t in tags)
    
    if is_romantasy:
        return 'Romantasy'
    elif is_crime:
        return 'Crime Thriller'
    elif is_romance and is_fantasy:
        return 'Romantasy'
    elif is_romance:
        return 'Romance Drama'
    elif is_fantasy:
        return 'Fantasy'
    return 'Unknown'

async def process_author(context, scraper, idx, author_name, df, semaphore, results_list):
    async with semaphore:
        print(f"[{idx}] Searching for Top 3 books for Author: {author_name}...")
        try:
            # Scrape top 3 books
            books_data = await scraper.scrape_top_books_by_author(context, author_name, count=3)
            
            if not books_data:
                print(f"[{idx}] No books found for {author_name}.")
                return
                
            for data in books_data:
                new_row = {col: '' for col in df.columns}
                new_row['Author Name'] = author_name
                new_row['Name of Series'] = data.get('Book_Title', '')
                new_row['GoodReads series link'] = data.get('GoodReads_Series_URL') or data.get('GoodReads_Book_URL', '')
                new_row['Number of PRIMARY books in the series'] = data.get('Num_Primary_Books', 1)
                new_row['Rating (out of 5) of Primary Book 1'] = data.get('Book1_Rating', '')
                new_row['Ratings (#) of Primary Book 1'] = data.get('Book1_Num_Ratings', '')
                new_row['Synopsis (if available)'] = data.get('Description', '')
                new_row['No. of pages in Book 1'] = data.get('Num_Pages', '')
                
                try:
                    num_pages = int(data.get('Num_Pages', 0))
                    num_books = int(data.get('Num_Primary_Books', 1))
                    new_row['Page Count (Sum of no. of pages in all primary books)'] = num_pages * num_books if num_pages else ''
                except:
                    pass
                
                all_genres = data.get('All_Genres', [])
                new_row['Genre tags- Up to 7 tags'] = ", ".join(all_genres)
                
                genre = map_genres(all_genres)
                new_row['Genre'] = genre
                new_row['Sub-Genre'] = 'Needs Mapping'
                
                results_list.append((idx, new_row))
                
            print(f"[{idx}] Successfully added {len(books_data)} books for '{author_name}'!")
        except Exception as e:
            print(f"[{idx}] Error scraping '{author_name}': {e}")

async def run_scrape():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found!")
        return

    df = pd.read_excel(EXCEL_FILE)
    
    scraper = GoodreadsScraper(headless=False)
    tasks = []
    results_list = []
    indices_to_drop = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        login_page = await context.new_page()
        await scraper.login_to_goodreads(login_page)
        
        semaphore = asyncio.Semaphore(MAX_CONCURRENT)
        
        # Find authors with missing books
        for idx, row in df.iterrows():
            title = str(row.get('Name of Series', '')).strip()
            author = str(row.get('Author Name', '')).strip()
            
            if author and (title.lower() == 'nan' or not title):
                indices_to_drop.append(idx)
                tasks.append(process_author(context, scraper, idx, author, df, semaphore, results_list))
                
        if tasks:
            await asyncio.gather(*tasks)
            
        await login_page.close()
        await browser.close()
        
    print("Processing results...")
    
    # Drop empty rows
    df = df.drop(indices_to_drop)
    
    # Append new rows
    new_rows = [r[1] for r in sorted(results_list, key=lambda x: x[0])]
    if new_rows:
        new_df = pd.DataFrame(new_rows)
        df = pd.concat([df, new_df], ignore_index=True)
        
    print("Saving Excel...")
    df.to_excel(EXCEL_FILE, index=False)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        align_left_nowrap = Alignment(horizontal='left', vertical='top', wrap_text=False)
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 18
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = align_left_nowrap
        wb.save(EXCEL_FILE)
        print("Styling applied successfully.")
    except Exception as e:
        print(f"Styling error: {e}")
    
    print("Done!")

if __name__ == '__main__':
    asyncio.run(run_scrape())
