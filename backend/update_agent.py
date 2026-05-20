from openpyxl import load_workbook
import os

file_name = 'kensington_authors_MASTER_1_to_1488.xlsx'
wb = load_workbook(file_name)
ws = wb.active

# Find the 'Name of agent' column index
agent_col_idx = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Name of agent":
        agent_col_idx = col
        break

if agent_col_idx:
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=agent_col_idx).value = "Tim Moore"
    
    wb.save(file_name)
    print("Agent name updated to 'Tim Moore' for all rows.")
else:
    print("'Name of agent' column not found.")

# Open the file on Windows
os.startfile(os.path.abspath(file_name))
print("File opened.")
