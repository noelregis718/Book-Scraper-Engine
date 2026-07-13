import asyncio
import os
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = '../New_Agency_Template.xlsx'

async def fetch_link(context, title, author, sem, row_idx, ws):
    async with sem:
        page = await context.new_page()
        try:
            print(f"[Row {row_idx}] Searching link for: '{title}' by {author}...")
            # We use brave search because it's fast and doesn't block
            query = f'"{title}" {author} goodreads'
            url = f"https://search.brave.com/search?q={query.replace(' ', '+')}"
            
            await page.goto(url, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(1)
            
            # Look for the first goodreads book link
            links = await page.query_selector_all('a[href*="goodreads.com/book/show/"]')
            found_link = None
            if links:
                for link in links:
                    href = await link.evaluate("el => el.href")
                    if href and "goodreads.com/book/show/" in href:
                        found_link = href
                        break
                        
            if not found_link:
                # Try a broader search without quotes
                query = f'{title} {author} goodreads book'
                url = f"https://search.brave.com/search?q={query.replace(' ', '+')}"
                await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(1)
                links = await page.query_selector_all('a[href*="goodreads.com/book/show/"]')
                if links:
                    for link in links:
                        href = await link.evaluate("el => el.href")
                        if href and "goodreads.com/book/show/" in href:
                            found_link = href
                            break

            if found_link:
                print(f"  -> [Row {row_idx}] Success: {found_link}")
                ws.cell(row=row_idx, column=4).value = found_link
            else:
                print(f"  -> [Row {row_idx}] Failed to find link.")
                ws.cell(row=row_idx, column=4).value = "Not Found"
                
        except Exception as e:
            print(f"  -> [Row {row_idx}] Error: {e}")
            ws.cell(row=row_idx, column=4).value = "Error"
        finally:
            await page.close()

async def main():
    print(f"Loading Excel file: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    tasks_data = []
    
    # Column 4 is GoodReads series link (D)
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        author = ws.cell(row=row, column=2).value
        link = ws.cell(row=row, column=4).value
        
        if title and str(title).strip() and str(title).strip() != "N/A":
            if not link or str(link).strip() == "N/A" or str(link).strip() == "nan":
                tasks_data.append({
                    'row': row,
                    'title': str(title).strip(),
                    'author': str(author).strip() if author else ""
                })
                
    print(f"Found {len(tasks_data)} books missing links.")
    if not tasks_data:
        return
        
    sem = asyncio.Semaphore(8) # 8 tabs concurrently for maximum speed
    
    async with async_playwright() as p:
        print("Launching headless browser...")
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        tasks = []
        for t in tasks_data:
            tasks.append(fetch_link(context, t['title'], t['author'], sem, t['row'], ws))
            
        print("\n--- Starting highly concurrent link extraction ---")
        await asyncio.gather(*tasks)
        await browser.close()
        
    wb.save(EXCEL_FILE)
    print("\nExcel file updated successfully with recovered links!")

if __name__ == "__main__":
    asyncio.run(main())
