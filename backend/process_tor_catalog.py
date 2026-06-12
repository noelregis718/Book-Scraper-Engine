import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def identify_romantasy_subgenre(synopsis):
    if not isinstance(synopsis, str) or len(synopsis) < 10:
        return "N/A", "No"
    
    text = synopsis.lower()
    
    # 1. Broad definition check (Is it Romantasy?)
    romance_kws = ['romance', 'love', 'passion', 'lover', 'heart', 'desire', 'marriage', 'betrothal', 'mate', 'soulmate', 'kiss', 'attraction', 'forbidden']
    fantasy_kws = ['magic', 'fantasy', 'sword', 'dragon', 'spell', 'curse', 'witch', 'wizard', 'demon', 'god', 'realm', 'kingdom', 'empire', 'fairy', 'fae', 'monster', 'vampire', 'werewolf', 'supernatural', 'power']
    
    is_romantasy = "No"
    # If it has at least one strong romance word AND one fantasy word, or if it explicitly says romantasy
    if (any(r in text for r in romance_kws) and any(f in text for f in fantasy_kws)) or 'romantasy' in text:
        is_romantasy = "Yes"
        
    if is_romantasy == "No":
        # Still check if it strongly hits any subgenre just in case
        pass

    subgenre = "N/A"
    
    # Priority matching for the 12 subgenres
    if any(k in text for k in ["werewolf", "shifter", "alpha", "pack", "omega", "luna"]):
        subgenre = "Werewolf / Shifter Romance"
    elif any(k in text for k in ["monster", "orc", "kraken", "alien", "beast"]):
        subgenre = "Monster Romance (Non-Shifter)"
    elif any(k in text for k in ["retelling", "mythology", "greek gods", "hades", "persephone", "fairy tale", "legend", "myth"]):
        subgenre = "Mythology, Legend & Fairy Tale Retelling"
    elif any(k in text for k in ["war college", "military academy", "dragon rider", "training", "fourth wing", "basgiath"]):
        subgenre = "War College / Military Academy"
    elif any(k in text for k in ["trials", "deadly games", "competition", "tournament", "survival", "hunger games"]):
        subgenre = "High-Stakes Games & Deadly Trials"
    elif any(k in text for k in ["university", "secret society", "library", "dark academia", "campus", "scholarly", "academy"]):
        subgenre = "Dark Academia Romantasy"
    elif any(k in text for k in ["gothic", "haunted", "manor", "dark romance", "gloomy", "castle", "shadows", "macabre"]):
        subgenre = "Gothic Dark Romantasy"
    elif any(k in text for k in ["isekai", "reincarnated", "villainess", "otome", "korean", "manhwa"]):
        subgenre = "Korean Romance Fantasy / Isekai"
    elif any(k in text for k in ["cozy", "cottagecore", "small town", "low stakes", "wholesome", "magical bakery", "inn", "tavern"]):
        subgenre = "Cozy / Cottagecore"
    elif any(k in text for k in ["vampire", "ghost", "witch", "paranormal", "psychic", "medium", "coven"]):
        subgenre = "Paranormal Romance"
    elif any(k in text for k in ["urban fantasy", "modern day", "city", "contemporary fantasy", "hidden world", "detective"]):
        subgenre = "Urban / Contemporary Fantasy Romance"
    elif any(k in text for k in ["court", "throne", "kingdom", "royalty", "epic fantasy", "fae", "high fantasy", "prince", "princess", "queen", "king", "empire"]):
        subgenre = "High Fantasy Court Adventure"
        
    # If we decided it IS romantasy based on broad keywords but didn't hit a specific subgenre rule:
    if is_romantasy == "Yes" and subgenre == "N/A":
        # Default fallback logic based on common tropes
        if any(w in text for w in ['city', 'modern', 'new york', 'london']):
            subgenre = "Urban / Contemporary Fantasy Romance"
        elif any(w in text for w in ['school', 'academy']):
            subgenre = "Dark Academia Romantasy"
        else:
            subgenre = "High Fantasy Court Adventure" # Safest default for Tor Fantasy

    # Conversely, if we found a strong subgenre but missed the broad romantasy check, we assume it is romantasy
    if subgenre != "N/A":
        is_romantasy = "Yes"

    return subgenre, is_romantasy

def process_and_style(file_path):
    print(f"Processing data in {file_path}...")
    df = pd.read_excel(file_path)
    
    df['Publisher'] = "Tor Publishing Group (US) / Pan Macmillan (UK)"
    df['Name of agent'] = "Gillian Green and Bella Pagan (UK)"
    
    count_yes = 0
    for idx, row in df.iterrows():
        synopsis = str(row.get('Synopsis (if available)', ''))
        subgenre, is_romantasy = identify_romantasy_subgenre(synopsis)
        
        df.at[idx, 'Romantasy = Yes or No?'] = is_romantasy
        df.at[idx, 'Romantasy Sub-Genre of series'] = subgenre
        if is_romantasy == "Yes":
            count_yes += 1

    df.to_excel(file_path, index=False)
    print(f"Data processing complete. {count_yes} books classified as Romantasy. Applying styles...")
    
    # Styling
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A1B38", end_color="2A1B38", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            col_letter = get_column_letter(cell.column)
            if col_letter in ['A', 'D', 'H']: 
                cell.alignment = left_wrap_alignment
            else:
                cell.alignment = center_alignment

    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 25 
    ws.column_dimensions['C'].width = 45 
    ws.column_dimensions['D'].width = 40 
    ws.column_dimensions['E'].width = 15 
    ws.column_dimensions['F'].width = 15 
    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 75 
    ws.column_dimensions['I'].width = 15 
    ws.column_dimensions['J'].width = 35 
    ws.column_dimensions['K'].width = 35 

    ws.freeze_panes = 'A2'
    wb.save(file_path)
    print("Styling applied successfully!")

if __name__ == "__main__":
    process_and_style(r"e:\Internship\PocketFM\agency_template.xlsx")
