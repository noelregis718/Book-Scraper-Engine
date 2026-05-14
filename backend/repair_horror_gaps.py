import asyncio
import pandas as pd
import os
import sys
import random
import json
from playwright.async_api import async_playwright
from openpyxl import load_workbook

# Add current directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from goodreads_scraper import GoodreadsScraper, normalize_title_for_search

# Configuration
INPUT_FILE = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
OUTPUT_FILE = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
BATCH_SIZE = 10

# Target Excel Ranges (Row numbers)
TARGET_RANGES = [
    (2840, 2882),
    (2341, 2391),
    (2239, 2282),
    (1673, 1710)
]

async def process_book(index, row, context, gr_scraper, ws):
    excel_title = str(row.get("Title", "Unknown"))
    print(f"  [Repair] Row {index+2}: {excel_title}...")
    
    match_found = False
    gr_data = {}
    
    try:
        await asyncio.sleep(random.uniform(0.1, 0.5))
        gr_data = await gr_scraper.scrape_goodreads_data(context, excel_title, "")
        
        if gr_data:
            found_url = gr_data.get('GoodReads_Book_URL', 'N/A')
            found_title = gr_data.get('Book_Title', '')
            norm_excel = normalize_title_for_search(excel_title)
            norm_found = normalize_title_for_search(found_title)
            
            if found_url != "N/A" and (norm_excel in norm_found or norm_found in norm_excel or norm_excel[:15] in norm_found):
                match_found = True
    except Exception as e:
        print(f"    [Error] {e}")

    # Update Worksheet
    row_idx = index + 2
    header = [cell.value for cell in ws[1]]
    
    def clean_val(val):
        return str(val) if val and str(val).upper() != "N/A" else ""

    mapping = {
        "Goodread Link": clean_val(gr_data.get("GoodReads_Book_URL")),
        "Series Link": clean_val(gr_data.get("GoodReads_Series_URL")),
        "Goodreads Rating": clean_val(gr_data.get("GoodReads_Rating")),
        "Goodreads No. of Ratings": clean_val(gr_data.get("GoodReads_Rating_Count")),
        "# of primary books": clean_val(gr_data.get("Num_Primary_Books")),
        "GR Book 1 Rating": clean_val(gr_data.get("Book1_Rating"))
    }
    
    if match_found:
        for col_name, value in mapping.items():
            if col_name in header:
                col_idx = header.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).value = value
        print(f"    [Success] Updated.")
    else:
        print(f"    [Skip] Not found or no match.")

async def main():
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(root_dir, INPUT_FILE)
    
    print(f"Loading {INPUT_FILE}...")
    df = pd.read_excel(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Flatten ranges into a single list of indices
    target_indices = []
    for start, end in TARGET_RANGES:
        # Excel Row X is index X-2
        target_indices.extend(range(start - 2, end - 1))
    
    print(f"Targeting {len(target_indices)} specific rows for repair...")
    
    async with async_playwright() as p:
        print("Launching browser...")
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        gr_scraper = GoodreadsScraper(headless=True)
        
        print("Logging in to Goodreads...")
        login_page = await context.new_page()
        await gr_scraper.login_to_goodreads(login_page)
        await login_page.close()
        
        # Process targeted rows in waves
        for i in range(0, len(target_indices), BATCH_SIZE):
            chunk = target_indices[i:i + BATCH_SIZE]
            print(f"\n>>> REPAIR WAVE: Processing {len(chunk)} rows...")
            
            tasks = []
            for idx in chunk:
                row = df.iloc[idx]
                tasks.append(process_book(idx, row, context, gr_scraper, ws))
            
            await asyncio.gather(*tasks)
            
            print(f"Wave complete. Saving progress...")
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"  [Warning] Could not save Excel file: {e}")
            
            await asyncio.sleep(2)
            
        print("\nRepair complete! Opening file...")
        await browser.close()
        
        if os.name == 'nt':
            os.startfile(file_path)

if __name__ == "__main__":
    asyncio.run(main())
