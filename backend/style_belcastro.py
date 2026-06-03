from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def apply_styling():
    file_name = r'E:\Internship\PocketFM\Belcastro_Agency_Formatted.xlsx'
    
    if not os.path.exists(file_name):
        print(f"Error: {file_name} not found!")
        return
        
    wb = load_workbook(file_name)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(left=Side(style='thin', color='D4D4D4'), right=Side(style='thin', color='D4D4D4'), 
                         top=Side(style='thin', color='D4D4D4'), bottom=Side(style='thin', color='D4D4D4'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Set preferred column widths based on expected content
    column_widths = {
        'A': 30, # Name of Series
        'B': 25, # Author Name
        'C': 20, # Publisher
        'D': 35, # GoodReads series link
        'E': 15, # Number of PRIMARY books in the series
        'F': 12, # Rating (out of 5)
        'G': 15, # Ratings (#)
        'H': 60, # Synopsis
        'I': 15, # Romantasy = Yes or No?
        'J': 30, # Romantasy Sub-Genre
        'K': 20  # Name of agent
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Content formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    # Save the styled file
    wb.save(file_name)
    print("Styling applied successfully.")
    
    # Open the file on Windows
    os.startfile(os.path.abspath(file_name))
    print("File opened.")

if __name__ == '__main__':
    apply_styling()
