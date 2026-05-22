import openpyxl

file_path = r"E:\Internship\PocketFM\Kensington_Media_Catalog.xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True)
ws = wb.active

print(f"Sheet Title: {ws.title}")
print(f"Max Rows: {ws.max_row}")
print(f"Max Columns: {ws.max_column}")

print("\nHeader Row:")
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(headers)

print("\nFirst 10 Rows:")
for r in range(2, 12):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {r}: {row_vals}".encode('ascii', 'ignore').decode('ascii'))
