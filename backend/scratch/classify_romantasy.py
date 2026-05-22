import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATALOG_FILE = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching seymour style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def classify_book(title, synopsis):
    title_lower = str(title or "").lower()
    syn_lower = str(synopsis or "").lower()
    combined = f"{title_lower} {syn_lower}"

    # General indicators of any magic/fantasy/paranormal context
    fantasy_indicators = [
        "magic", "fantasy", "fae", "spell", "elf", "dragon", "sorcerer", "curse", 
        "enchanted", "witch", "vampire", "werewolf", "shifter", "demon", "alien", 
        "mythology", "gods", "goddess", "kingdom", "throne", "crown", "paranormal",
        "supernatural", "coven", "warlock", "gargoyle", "monsters", "academy", "prophecy"
    ]
    
    # If no fantasy/paranormal indicators are present in the text, it's highly likely NOT a Romantasy book
    is_possibly_fantasy = any(word in combined for word in fantasy_indicators)
    if not is_possibly_fantasy:
        return "No", "N/A"

    # 1. Werewolf / Shifter Romance
    shifter_words = ["shifter", "werewolf", "pack", "alpha", "omega", "luna", "wolf", "mate", "howl", "werewolves", "shifters", "lycan"]
    if any(word in combined for word in shifter_words):
        return "Yes", "Werewolf / Shifter Romance"

    # 2. Mythology, Legend & Fairy Tale Retelling
    retelling_words = ["retelling", "fairy tale", "mythology", "myth", "hades", "persephone", "zeus", "greek", "norse", "gods", "goddess", "beauty and the beast", "cinderella", "rapunzel"]
    if any(word in combined for word in retelling_words):
        return "Yes", "Mythology, Legend & Fairy Tale Retelling"

    # 3. Monster Romance (Non-Shifter)
    monster_words = ["monster", "alien", "demon", "gargoyle", "minotaur", "kraken", "orc", "beast", "tentacle", "creature"]
    if any(word in combined for word in monster_words):
        return "Yes", "Monster Romance (Non-Shifter)"

    # 4. War College / Military Academy
    war_college_words = ["war college", "military academy", "cadet", "soldier", "military", "commander", "combat", "training ground"]
    if any(word in combined for word in war_college_words):
        return "Yes", "War College / Military Academy"

    # 5. Korean Romance Fantasy / Isekai
    isekai_words = ["isekai", "reincarnated", "reincarnation", "transmigrated", "transmigration", "villainess", "otome", "korean", "manhwa", "empress"]
    if any(word in combined for word in isekai_words):
        return "Yes", "Korean Romance Fantasy / Isekai"

    # 6. Dark Academia Romantasy
    academia_words = ["academy", "academia", "university", "college", "library", "secret society", "professor", "scholarly"]
    if any(word in combined for word in academia_words):
        return "Yes", "Dark Academia Romantasy"

    # 7. High-Stakes Games & Deadly Trials
    games_words = ["trials", "deadly games", "arena", "tournament", "lethal trials", "selection", "competition"]
    if any(word in combined for word in games_words):
        return "Yes", "High-Stakes Games & Deadly Trials"

    # 8. Gothic Dark Romantasy
    gothic_words = ["gothic", "dark romance", "haunted", "castle", "crypt", "shadow", "creepy", "macabre", "darkness", "grim", "morbid", "dracula"]
    if any(word in combined for word in gothic_words):
        return "Yes", "Gothic Dark Romantasy"

    # 9. Cozy / Cottagecore
    cozy_words = ["cozy", "cottagecore", "tea shop", "bakery", "inn", "village", "witchy cozy", "peaceful", "low-stakes", "comforting"]
    if any(word in combined for word in cozy_words):
        return "Yes", "Cozy / Cottagecore"

    # 10. High Fantasy Court Adventure
    court_words = ["court", "kingdom", "throne", "crown", "prince", "princess", "king", "queen", "royalty", "court intrigue", "empire", "high fantasy", "lord", "alliance"]
    if any(word in combined for word in court_words):
        return "Yes", "High Fantasy Court Adventure"

    # 11. Paranormal Romance
    paranormal_words = ["paranormal", "vampire", "witch", "ghost", "specter", "coven", "spell", "warlock", "spirit", "psychic", "medium"]
    if any(word in combined for word in paranormal_words):
        return "Yes", "Paranormal Romance"

    # 12. Urban / Contemporary Fantasy Romance
    urban_words = ["urban", "contemporary fantasy", "city", "modern", "streets", "detective", "supernatural city", "hidden magic", "metropolis"]
    if any(word in combined for word in urban_words):
        return "Yes", "Urban / Contemporary Fantasy Romance"

    # Default Fallbacks for General Fantasy
    if any(word in combined for word in ["court", "kingdom", "throne", "crown", "prince", "princess", "empress", "royalty"]):
        return "Yes", "High Fantasy Court Adventure"
    
    return "Yes", "Urban / Contemporary Fantasy Romance"

def run_classification():
    print("=" * 60)
    print("      SEYMOUR AGENCY: ROMANTASY SUBGENRE CLASSIFIER")
    print("=" * 60)
    print(f"Spreadsheet: {CATALOG_FILE}\n")

    if not os.path.exists(CATALOG_FILE):
        print(f"[Error] Catalog spreadsheet {CATALOG_FILE} not found!")
        return

    wb = load_workbook(CATALOG_FILE)
    ws = wb.active

    classified_yes = 0
    classified_no = 0
    counts = {}

    for r in range(2, ws.max_row + 1):
        book_title = ws.cell(row=r, column=1).value
        synopsis = ws.cell(row=r, column=8).value

        # Run high-precision classification
        is_romantasy, subgenre = classify_book(book_title, synopsis)

        # Write to Column 9 (Romantasy = Yes or No?)
        ws.cell(row=r, column=9, value=is_romantasy)
        
        # Write to Column 10 (Romantasy Sub-Genre of series)
        ws.cell(row=r, column=10, value=subgenre)

        if is_romantasy == "Yes":
            classified_yes += 1
            counts[subgenre] = counts.get(subgenre, 0) + 1
        else:
            classified_no += 1

    print(">>> Re-applying premium styling and saving workbook...")
    format_excel_sheet(ws)
    wb.save(CATALOG_FILE)
    print(">>> Catalog saved successfully!")

    print("\n" + "=" * 60)
    print("              CLASSIFICATION DISTRIBUTION")
    print("=" * 60)
    print(f"Romantasy 'Yes' Count: {classified_yes}")
    print(f"Romantasy 'No' Count:  {classified_no}")
    print(f"Total Rows Processed:  {ws.max_row - 1}")
    print("-" * 60)
    print("Subgenre Breakdown:")
    for sub, count in sorted(counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {sub:40} : {count} books")
    print("=" * 60 + "\n")

    # Auto-open on Windows
    if os.name == 'nt':
        print("  [System] Auto-opening sheet for review...")
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    run_classification()
