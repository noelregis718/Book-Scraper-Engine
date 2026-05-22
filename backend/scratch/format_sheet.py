import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def format_excel(file_path, output_path):
    print(f"Loading {file_path}...")
    df = pd.read_excel(file_path)
    
    # 1. Drop fully empty Unnamed columns
    unnamed_cols = [c for c in df.columns if 'Unnamed:' in str(c)]
    empty_unnamed = [c for c in unnamed_cols if df[c].isnull().all()]
    df = df.drop(columns=empty_unnamed)
    
    # 2. Reorder columns logically
    priority_cols = [
        'Subgenre', 'Title', 'Author', 'URL', 'Rating', 'No. of Ratings', 
        'Customer Reviews', 'Goodreads Rating', 'Goodreads No. of Ratings',
        'Publisher', 'Publication Date', 'Language', 'Format', 'Print Length'
    ]
    
    # Filter priority columns that actually exist in df
    existing_priority = [c for c in priority_cols if c in df.columns]
    other_cols = [c for c in df.columns if c not in existing_priority]
    
    new_order = existing_priority + other_cols
    df = df[new_order]
    
    print(f"Saving temporary formatted file...")
    df.to_excel(output_path, index=False)
    
    # 3. Apply OpenPyXL styling
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format Headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = border
        
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze Panes (First row)
    ws.freeze_panes = "A2"
    
    # Auto-adjust column widths and align cells
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        
        # Determine max length for width
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            
            # Apply border to all cells
            cell.border = border
            if cell.row > 1:
                # Wrap text for long columns like Synopsis or URL
                if ws.cell(row=1, column=cell.column).value in ['Synopsis/Summary', 'URL', 'Rationale']:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = Alignment(vertical="center")

        adjusted_width = min(max_length + 2, 50) # Cap width at 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Formatting complete! Saved to {output_path}")

if __name__ == "__main__":
    # Target the new sheet directly
    target_file = "Horror_-_Amazon_Keyword_Crawl (3).xlsx"
    
    if os.path.exists(target_file):
        # We use a temporary name then rename to safely overwrite
        temp_file = "temp_formatted.xlsx"
        format_excel(target_file, temp_file)
        
        # Replace the original with the formatted one
        if os.path.exists(temp_file):
            os.replace(temp_file, target_file)
            print(f"Success: The new sheet '{target_file}' has been updated with formatting.")
    else:
        print(f"File {target_file} not found.")
