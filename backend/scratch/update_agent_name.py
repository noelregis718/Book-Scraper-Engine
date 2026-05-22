import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATALOG_FILE = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching seymour style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def update_agent_name():
    print("=" * 60)
    print("      SEYMOUR AGENCY: UPDATE AGENT NAME")
    print("=" * 60)
    print(f"Spreadsheet: {CATALOG_FILE}\n")

    if not os.path.exists(CATALOG_FILE):
        print(f"[Error] Catalog spreadsheet {CATALOG_FILE} not found!")
        return

    wb = load_workbook(CATALOG_FILE)
    ws = wb.active

    agent_name = "Nicole Resciniti"
    updated_count = 0

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=11, value=agent_name)
        updated_count += 1

    print(">>> Re-applying premium styling and saving workbook...")
    format_excel_sheet(ws)
    wb.save(CATALOG_FILE)
    print(">>> Catalog saved successfully!")

    print("\n" + "=" * 60)
    print(f"Agent Name Set:   {agent_name}")
    print(f"Rows Updated:      {updated_count}")
    print("=" * 60 + "\n")

    # Auto-open on Windows
    if os.name == 'nt':
        print("  [System] Auto-opening sheet for review...")
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    update_agent_name()
