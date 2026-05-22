import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATALOG_FILE = r"E:\Internship\PocketFM\Kensington_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def create_kensington_catalog():
    print("=" * 60)
    print("      KENSINGTON AGENCY: CATALOG CREATION")
    print("=" * 60)
    print(f"Creating empty catalog at: {CATALOG_FILE}\n")

    # Create new workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Kensington Media Catalog"

    # Set up basic parameters (Grid lines, etc.)
    ws.views.sheetView[0].showGridLines = True

    # Styling Elements
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    border_side = Side(border_style="thin", color="D3D3D3")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Write Headers
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # Standard column widths initialization
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 22  # Standard visual separation
        
    # Specific width overrides for longer columns
    ws.column_dimensions["A"].width = 28  # Name of Series
    ws.column_dimensions["B"].width = 24  # Author Name
    ws.column_dimensions["D"].width = 30  # Goodreads link
    ws.column_dimensions["H"].width = 40  # Synopsis

    # Freeze header pane
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # Save
    wb.save(CATALOG_FILE)
    print(">>> Kensington Media Catalog created and styled successfully!")

    # Auto-open on Windows
    if os.name == 'nt':
        print("  [System] Auto-opening sheet for review...")
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    create_kensington_catalog()
