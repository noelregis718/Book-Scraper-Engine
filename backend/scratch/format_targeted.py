from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

OUTPUT_FILE = r"E:\Internship\PocketFM\Jacob_Peppers_to_End_Enrichment.xlsx"

def apply_formatting(path):
    if not os.path.exists(path):
        print(f"Error: {path} not found.")
        return
        
    print(f"Applying professional formatting to {path}...")
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        # 1. Header Styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # 2. Zebra Striping & Content Alignment
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
        
        # 3. Freeze Panes (Lock Header)
        ws.freeze_panes = "A2"
        
        # 4. Column Auto-Width
        # Define some base widths
        col_widths = {
            'Author Name': 30,
            'Author Email ID': 35,
            'Author Contact Form - Website': 50,
            'Agency Email ID': 35
        }
        
        for col in ws.columns:
            column_name = col[0].value
            column_letter = col[0].column_letter
            width = col_widths.get(column_name, 25)
            ws.column_dimensions[column_letter].width = width
            
        # 5. Auto-Filter
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(path)
        print("Formatting applied successfully.")
    except Exception as e:
        print(f"Error applying formatting: {e}")

if __name__ == "__main__":
    apply_formatting(OUTPUT_FILE)
    if os.name == 'nt':
        os.startfile(OUTPUT_FILE)
