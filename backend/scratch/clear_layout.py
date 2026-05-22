from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def apply_clear_first_line_layout(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Applying clear first-line layout for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Force a slightly taller "thin" height (20) and DISABLE wrap_text to prevent vertical overlap
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20 # Increased from 15 to 20 for readability
        
        for cell in ws[row_idx]:
            # wrap_text=False ensures only the first line shows and it stays clear
            cell.alignment = Alignment(wrap_text=False, vertical='center', horizontal='left')
            cell.border = thin_border

    # 2. Widen the key columns so the first line is actually readable
    col_widths = {
        'Subgenre': 20,
        'Title': 60,         # Widened significantly
        'URL': 40,
        'Author': 25,
        'Synopsis/Summary': 30,
        'Goodread Link': 50,  # Widened significantly
        'Series Link': 50     # Widened significantly
    }

    for idx, col in enumerate(ws.columns):
        header_val = col[0].value
        column_letter = get_column_letter(idx + 1)
        
        width = col_widths.get(header_val, 15)
        ws.column_dimensions[column_letter].width = width

    wb.save(file_path)
    print("Clear layout applied! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    apply_clear_first_line_layout(target)
