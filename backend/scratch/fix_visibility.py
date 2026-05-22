from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, Color
from openpyxl.utils import get_column_letter
import os

def fix_invisible_text(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Fixing invisible text for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Header Style (Row 1)
    header_font = Font(bold=True, color="FFFFFF")
    
    # 2. Data Style (Row 2 onwards)
    data_font = Font(bold=False, color="000000") # Forced Black
    
    for row_idx, row in enumerate(ws.iter_rows()):
        # Keep our fixed height
        ws.row_dimensions[row_idx + 1].height = 18
        
        for cell in row:
            if row_idx == 0:
                cell.font = header_font
            else:
                cell.font = data_font
            
            # Keep our strict boundary layout
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = thin_border

    wb.save(file_path)
    print("Visibility fixed! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    fix_invisible_text(target)
