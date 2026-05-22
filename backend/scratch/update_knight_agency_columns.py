import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def update_excel_columns(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Updating {file_path}...")
    df = pd.read_excel(file_path)
    
    if "Is it Romantasy ?" not in df.columns:
        # Find the index of "Romantasy Sub-Genre of series"
        cols = list(df.columns)
        if "Romantasy Sub-Genre of series" in cols:
            idx = cols.index("Romantasy Sub-Genre of series")
            # Insert before it
            df.insert(idx, "Is it Romantasy ?", df["Romantasy Sub-Genre of series"].apply(lambda x: "Yes" if str(x) != "N/A" and str(x) != "nan" else "No"))
            print(f"Added 'Is it Romantasy ?' column at index {idx}")
        else:
            # Append if not found
            df["Is it Romantasy ?"] = "No"
            print("Added 'Is it Romantasy ?' column at the end")
            
        df.to_excel(file_path, index=False)
        
        # Restyle
        wb = load_workbook(file_path)
        ws = wb.active
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
        column_widths = {"A": 30, "B": 25, "C": 20, "D": 35, "E": 15, "F": 12, "G": 12, "H": 60, "I": 18, "J": 30, "K": 25}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(file_path)
        print("Styling applied.")
    else:
        print("Column 'Is it Romantasy ?' already exists.")

if __name__ == "__main__":
    update_excel_columns(r"E:\Internship\PocketFM\Knight Agency.xlsx")
