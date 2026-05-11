import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_styled_template(filename):
    columns = [
        "Name of Series", 
        "Author Name", 
        "Publisher", 
        "GoodReads series link",
        "Number of PRIMARY books in the series", 
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1", 
        "Synopsis (if available)",
        "Romantasy = Yes or No?", 
        "Romantasy Sub-Genre of series", 
        "Name of agent"
    ]
    
    # Create empty DataFrame with headers
    df = pd.DataFrame(columns=columns)
    
    # Save to Excel
    df.to_excel(filename, index=False)
    
    # Apply styling
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    ws = wb.active
    
    # Style definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Style Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust Column Widths
    column_widths = {
        "A": 30, # Name of Series
        "B": 25, # Author Name
        "C": 20, # Publisher
        "D": 35, # GoodReads link
        "E": 15, # Num Books
        "F": 12, # Rating
        "G": 12, # Num Ratings
        "H": 60, # Synopsis
        "I": 22, # Romantasy ?
        "J": 30, # Sub-Genre
        "K": 25  # Agent
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(filename)
    print(f"Successfully created styled template: {filename}")

if __name__ == "__main__":
    target_path = r"e:\Internship\PocketFM\Agency_Scraping_Template.xlsx"
    create_styled_template(target_path)
