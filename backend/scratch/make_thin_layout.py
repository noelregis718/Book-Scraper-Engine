from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def make_thin_layout(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Applying thin layout for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Reset alignment to NO WRAPPING and set standard row height
    for row_idx, row in enumerate(ws.iter_rows()):
        # Set a fixed, thin row height (15 is standard Excel height)
        ws.row_dimensions[row_idx + 1].height = 15
        for cell in row:
            # Disable wrap text to keep rows thin
            cell.alignment = Alignment(wrap_text=False, vertical='center', horizontal='left')
            cell.border = thin_border

    # 2. Adjust column widths slightly to ensure headers fit
    for col in ws.columns:
        column_letter = get_column_letter(col[0].column)
        header_val = col[0].value
        
        # Use header length as base width, but don't make it too wide
        if header_val:
            width = min(max(len(str(header_val)) + 5, 12), 30)
            ws.column_dimensions[column_letter].width = width

    wb.save(file_path)
    print("Thin layout applied! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    make_thin_layout(target)
