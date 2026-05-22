import os
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOCAL_HTML_FILE = r"E:\Internship\PocketFM\Authors - Kensington Publishing.html"
CATALOG_FILE = r"E:\Internship\PocketFM\Kensington_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching kensington style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def parse_local_html():
    print("=" * 60)
    print("      KENSINGTON AGENCY: LOCAL HTML PARSER")
    print("=" * 60)
    print(f"Reading HTML from: {LOCAL_HTML_FILE}\n")

    if not os.path.exists(LOCAL_HTML_FILE):
        print(f"[Error] Local HTML file {LOCAL_HTML_FILE} does not exist yet!")
        print("Please save the webpage as HTML first.")
        return

    with open(LOCAL_HTML_FILE, "r", encoding="utf-8", errors="ignore") as f:
        html_content = f.read()

    # Parse all anchor links using regex to bypass BeautifulSoup dependencies
    # Matching href containing /author/ or /authors/
    link_pattern = re.compile(r'<a\s+[^>]*href=["\']([^"\']*(?:/author/|/authors/)[^"\']*)["\'][^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL)
    matches = link_pattern.findall(html_content)

    author_names = []
    blacklist = ["author", "authors", "home", "books", "contact", "about", "submission", "faq", "terms", "privacy", "help", "search"]

    for href, text in matches:
        # Strip HTML tags inside link text if any
        clean_text = re.sub(r'<[^>]+>', '', text).strip()
        # Clean whitespaces
        clean_text = " ".join(clean_text.split())
        
        if clean_text and len(clean_text) > 2 and len(clean_text) < 50:
            text_lower = clean_text.lower()
            if not any(b_item in text_lower for b_item in blacklist):
                if clean_text not in author_names:
                    author_names.append(clean_text)

    if not author_names:
        print("[Warning] No authors found matching link patterns. Trying broader regex search...")
        # Fallback regex search for author links
        fallback_pattern = re.compile(r'href=["\'][^"\']*/author/[^"\']*["\'][^>]*>([^<]+)', re.IGNORECASE)
        fallback_matches = fallback_pattern.findall(html_content)
        for name in fallback_matches:
            clean_text = name.strip()
            if clean_text and len(clean_text) > 2 and len(clean_text) < 50:
                text_lower = clean_text.lower()
                if not any(b_item in text_lower for b_item in blacklist):
                    if clean_text not in author_names:
                        author_names.append(clean_text)

    print(f"[System] Found {len(author_names)} unique author names in HTML file.")
    if not author_names:
        print("[Error] No author names could be parsed from the file. Please verify it's the correct authors page!")
        return

    # Sort alphabetically
    author_names = sorted(author_names)

    # Load and populate spreadsheet
    print(f"\n>>> Writing {len(author_names)} authors to: {CATALOG_FILE}...")
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active

    # Clear existing data rows (except header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for idx, author in enumerate(author_names, 2):
        ws.cell(row=idx, column=1, value="N/A")            # Name of Series
        ws.cell(row=idx, column=2, value=author)           # Author Name
        ws.cell(row=idx, column=3, value="Kensington")     # Publisher
        ws.cell(row=idx, column=4, value="N/A")            # GoodReads link
        ws.cell(row=idx, column=5, value=1)                # Books count
        ws.cell(row=idx, column=6, value="N/A")            # Rating
        ws.cell(row=idx, column=7, value="N/A")            # Ratings #
        ws.cell(row=idx, column=8, value="N/A")            # Synopsis
        ws.cell(row=idx, column=9, value="No")             # Romantasy Yes/No
        ws.cell(row=idx, column=10, value="N/A")           # Romantasy Subgenre
        ws.cell(row=idx, column=11, value="N/A")           # Agent Name

    print(">>> Applying styling and saving workbook...")
    format_excel_sheet(ws)
    wb.save(CATALOG_FILE)
    print(">>> Catalog updated successfully!")

    print("\n" + "=" * 60)
    print("                 PARSING COMPLETE")
    print("=" * 60)
    print(f"Total Authors Saved:   {len(author_names)}")
    print(f"Output Catalog:        {CATALOG_FILE}")
    print("=" * 60 + "\n")

    # Auto-open on Windows
    if os.name == 'nt':
        print("  [System] Auto-opening sheet for review...")
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    parse_local_html()
