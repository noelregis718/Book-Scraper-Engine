from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def apply_clean_cutoff_layout(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Applying clean cutoff layout for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Apply fixed height and TOP-aligned wrapping
    # Height of 18 is perfect for 1 line of text in standard 11pt font
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18
        
        for cell in ws[row_idx]:
            # wrap_text=True + fixed height = No spillover
            # vertical='top' = First line is clear and readable
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = thin_border

    # 2. Set column widths
    col_widths = {
        'Subgenre': 18,
        'Title': 45,
        'URL': 35,
        'Author': 22,
        'Synopsis/Summary': 30,
        'Goodread Link': 40,
        'Series Link': 40
    }

    for idx, col in enumerate(ws.columns):
        header_val = col[0].value
        column_letter = get_column_letter(idx + 1)
        width = col_widths.get(header_val, 15)
        ws.column_dimensions[column_letter].width = width

    wb.save(file_path)
    print("Clean cutoff layout applied! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    apply_clean_cutoff_layout(target)
