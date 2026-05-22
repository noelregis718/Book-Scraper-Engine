import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = r"E:\Internship\PocketFM\gernert_books_full_to_end_updated.xlsx"
OUTPUT_FILE = r"E:\Internship\PocketFM\Gernert_Media_Catalog.xlsx"

def clean_author_name(name):
    if not isinstance(name, str):
        return "Unknown"
    cleaned = name.strip()
    # Remove leading "By " case-insensitive
    cleaned = re.sub(r'(?i)^by\s+', '', cleaned)
    # Replace multiple spaces with a single space
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()

def create_gernert_catalog():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: Source file {INPUT_FILE} not found!")
        return

    print(f"Loading {INPUT_FILE}...")
    df_src = pd.read_excel(INPUT_FILE)
    
    # Standard 11 columns requested
    columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]
    
    # Construct empty records mapped to new columns
    rows = []
    for idx, row in df_src.iterrows():
        book_title = str(row.get("Book Title", "N/A")).strip()
        contributor = row.get("Contributor(s)", "Unknown")
        author_cleaned = clean_author_name(contributor)
        
        new_row = {
            "Name of Series": book_title,
            "Author Name": author_cleaned,
            "Publisher": "The Gernert Company",
            "GoodReads series link": "N/A",
            "Number of PRIMARY books in the series": "N/A",
            "Rating (out of 5) of Primary Book 1": "N/A",
            "Ratings (#) of Primary Book 1": "N/A",
            "Synopsis (if available)": "N/A",
            "Romantasy = Yes or No?": "N/A",
            "Romantasy Sub-Genre of series": "N/A",
            "Name of agent": "David Gernert"
        }
        rows.append(new_row)
        
    df_out = pd.DataFrame(rows, columns=columns)
    
    # Save the catalog
    print(f"Saving temporary structured spreadsheet to {OUTPUT_FILE}...")
    df_out.to_excel(OUTPUT_FILE, index=False)
    
    # --- Apply Premium Styling ---
    print(f"Applying professional styling to {OUTPUT_FILE}...")
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    # Header styling (Navy Blue, white text)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Border definition
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"),
        right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"),
        bottom=Side(style='thin', color="BFBFBF")
    )
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Header format
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Body rows formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10, color="000000")
            
            # Alignments per column type
            if cell.column in [1, 2, 4, 8]:  # Title, Author, Goodreads link, Synopsis
                if cell.column == 8: # Synopsis has wrapping
                    cell.alignment = wrap_align
                else:
                    cell.alignment = left_align
            else:
                cell.alignment = center_align
                
    # Column width specifications
    column_widths = {
        1: 35,  # Name of Series
        2: 25,  # Author Name
        3: 15,  # Publisher
        4: 40,  # GoodReads series link
        5: 15,  # Number of PRIMARY books in the series
        6: 15,  # Rating (out of 5) of Primary Book 1
        7: 15,  # Ratings (#) of Primary Book 1
        8: 60,  # Synopsis (if available)
        9: 22,  # Romantasy = Yes or No?
        10: 30, # Romantasy Sub-Genre of series
        11: 20  # Name of agent
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Freeze the top header row
    ws.freeze_panes = "A2"
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save the polished workbook
    wb.save(OUTPUT_FILE)
    print(f"Success! Beautifully styled 11-column catalog generated: {OUTPUT_FILE}")
    
    if os.name == 'nt':
        os.startfile(OUTPUT_FILE)

if __name__ == "__main__":
    create_gernert_catalog()
