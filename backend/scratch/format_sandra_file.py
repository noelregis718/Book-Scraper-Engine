import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

file_path = r"e:\Internship\PocketFM\Sandra Djikstra.xlsx"

if not os.path.exists(file_path):
    print(f"Error: {file_path} not found.")
    exit()

print(f"Applying professional formatting to {file_path}...")

# Load the workbook and select the active sheet
wb = load_workbook(file_path)
ws = wb.active

# Define Styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# 1. Style Headers
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 2. Style Data Rows (Wrapping and Alignment)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

# 3. Adjust Column Widths
column_widths = {
    "A": 35, # Name of Series
    "B": 25, # Author Name
    "C": 20, # Publisher
    "D": 40, # GoodReads series link
    "E": 15, # Number of PRIMARY books
    "F": 12, # Rating
    "G": 12, # Ratings (#)
    "H": 70, # Synopsis
    "I": 18, # Is it Romantasy ?
    "J": 30, # Romantasy Sub-Genre
    "K": 25  # Name of agent
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# 4. Freeze Panes (Header row)
ws.freeze_panes = "A2"

# Save the workbook
wb.save(file_path)
print("Formatting complete!")
