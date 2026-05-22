from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def fix_layout(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Fixing layout for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Apply wrapping and alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            # Enable wrap text and top-left alignment
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = thin_border

    # 2. Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        # Check header length
        header_val = col[0].value
        if header_val:
            max_length = len(str(header_val))
            
        # Check first 50 rows for data length
        for cell in col[:50]:
            if cell.value:
                # For very long text (like synopsis), we don't want the column to be miles wide
                # We'll cap the width and let Wrap Text handle the rest
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        
        # Reasonable width bounds
        adjusted_width = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
    print("Layout fixed! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    fix_layout(target)
