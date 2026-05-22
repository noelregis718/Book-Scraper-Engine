import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def apply_ai_mapping(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Applying AI Judgment to {file_path}...")
    df = pd.read_excel(file_path)

    # Dictionary of AI classifications for every title in the sheet
    mapping = {
        "HER CAMERON BODYGUARD": ("No", "N/A"),
        "TROUBLE IN WEST TEXAS": ("No", "N/A"),
        "COLTON ON GUARD": ("No", "N/A"),
        "EXPOSING LETHAL SECRETS": ("No", "N/A"),
        "HER CAMERON DEFENDER": ("No", "N/A"),
        "Saved": ("No", "N/A"),
        "COLTON’S ULTIMATE TEST": ("No", "N/A"),
        "REFUGE UP IN FLAMES": ("No", "N/A"),
        "Protected": ("No", "N/A"),
        "EVERYDAY WITCH’S BOOK OF DEITIES: ANCIENT GODS FOR MODERN PAGANS": ("No", "N/A"),
        "PARANORMAL PAYBACK edited": ("Yes", "Paranormal Romance"),
        "HEIR TO THE ITALIAN ALTAR": ("No", "N/A"),
        "WE BURNED SO BRIGHT": ("No", "N/A"),
        "DEATH AT A HIGHLAND WEDDING": ("No", "N/A"),
        "THE GOBLIN IN THE SINK DRAIN": ("Yes", "Urban / Contemporary Fantasy Romance"),
        "THE VAMPIRE IN THE POTTING SHED": ("Yes", "Urban / Contemporary Fantasy Romance"),
        "MORE LIKE ENEMIGAS": ("No", "N/A"),
        "RISE OF THE CELESTIALS": ("Yes", "High Fantasy Court Adventure"),
        "STAR SHIPPED": ("No", "N/A"),
        "A DEADLY INHERITANCE": ("Yes", "Dark Academia Romantasy"),
        "JITTERBUG": ("No", "N/A"),
        "LOVE A COMEBACK": ("No", "N/A"),
        "TRAILBREAKER": ("No", "N/A"),
        "OUT OF THE LOOP": ("Yes", "Urban / Contemporary Fantasy Romance"),
        "FIRST SIGN OF DANGER": ("No", "N/A"),
        "OWN YOUR TIME": ("No", "N/A"),
        "SUCH A PERFECT FAMILY": ("No", "N/A"),
        "THE MERMAID IN THE SHOT GLASS": ("Yes", "Urban / Contemporary Fantasy Romance"),
        "THE SHARK HOUSE": ("No", "N/A"),
        "JUNKYARD RIDERS": ("Yes", "Werewolf / Shifter Romance"),
        "KINGDOM OF TODAY": ("Yes", "High Fantasy Court Adventure"),
        "THE BURNING QUEEN": ("Yes", "High Fantasy Court Adventure"),
        "EMBER ETERNAL": ("Yes", "Paranormal Romance"),
        "THE PLACE WHERE THEY BURIED YOUR HEART": ("No", "N/A"),
        "THE THIN BLUE LEY-LINE": ("Yes", "Urban / Contemporary Fantasy Romance"),
        "THE HAUNTING OF PAYNE’S HOLLOW": ("No", "N/A"),
        "THE GUEST IN ROOM 120": ("No", "N/A"),
        "EVERYTHING ABOUT YOU": ("No", "N/A"),
        "THE ENDURING UNIVERSE": ("Yes", "High Fantasy Court Adventure"),
        "WHO WILL YOU SAVE?": ("No", "N/A"),
        "THE UNDERACHIEVER’S GUIDE TO LOVE AND SAVING THE WORLD": ("Yes", "High Fantasy Court Adventure"),
        "WITCH OF THE WOLVES": ("Yes", "Werewolf / Shifter Romance"),
        "WICKEDLY EVER AFTER": ("Yes", "Mythology, Legend & Fairy Tale Retelling"),
        "HEAD WITCH IN CHARGE": ("Yes", "Paranormal Romance"),
        "KINGDOM OF TOMORROW": ("Yes", "High Fantasy Court Adventure"),
        "A SEASON OF SECOND CHANCES": ("No", "N/A"),
        "JOY MOODY IS OUT OF TIME": ("No", "N/A"),
        "FOR THE RECORD": ("No", "N/A"),
        "BONBONS AND BURGLARS": ("No", "N/A"),
        "THE WELL": ("No", "N/A"),
        "A MASTERY OF MONSTERS": ("Yes", "Monster Romance (Non-Shifter)"),
        "FRIENDS TO LOVERS": ("No", "N/A"),
        "ATONEMENT SKY": ("Yes", "Paranormal Romance"),
        "EVERY STEP SHE TAKES": ("No", "N/A"),
        "WE WHO HUNT ALEXANDERS": ("Yes", "Gothic Dark Romantasy"),
        "TEN INCARNATIONS OF REBELLION": ("Yes", "Korean Romance Fantasy / Isekai"),
        "WRITING MR. WRONG": ("No", "N/A"),
        "CHECKERED HEARTS": ("No", "N/A"),
        "THE SECRET LIBRARY OF HANNA REEVES": ("No", "N/A"),
        "IF I TOLD YOU, I’D HAVE TO KISS YOU": ("No", "N/A"),
        "I THINK I’M IN LOVE WITH AN ALIEN": ("Yes", "Monster Romance (Non-Shifter)"),
        "ARCHANGEL’S ASCENSION": ("Yes", "Paranormal Romance"),
        "DAMNED": ("Yes", "Paranormal Romance"),
        "THE SHERIFF NEXT DOOR": ("No", "N/A"),
        "DIDN’T YOU USE TO BE QUEENIE B": ("No", "N/A"),
        "TRUST ME ON THIS": ("No", "N/A"),
        "Reunited": ("No", "N/A"),
        "IN IT TO WIN IT": ("No", "N/A"),
        "JUNKYARD WAR": ("Yes", "Werewolf / Shifter Romance"),
        "SECRET OF THE LOST PEARLS": ("No", "N/A"),
        "A VALENTINE FOR CHRISTMAS": ("No", "N/A"),
        "END OF THE DAY": ("No", "N/A"),
        "WOLFSONG": ("Yes", "Werewolf / Shifter Romance"),
        "THE FINAL TRIAL: ROYAL GUIDE TO MONSTER SLAYING": ("Yes", "High-Stakes Games & Deadly Trials")
    }

    # Apply the mapping
    def get_romantasy_status(row):
        title = str(row['Name of Series']).strip()
        # Direct match or partial match for titles like "Saved the Texas Cowboy"
        for key in mapping:
            if key in title or title in key:
                return mapping[key][0]
        return "No"

    def get_subgenre_status(row):
        title = str(row['Name of Series']).strip()
        for key in mapping:
            if key in title or title in key:
                return mapping[key][1]
        return "N/A"

    df['Is it Romantasy ?'] = df.apply(get_romantasy_status, axis=1)
    df['Romantasy Sub-Genre of series'] = df.apply(get_subgenre_status, axis=1)

    df.to_excel(file_path, index=False)
    
    # Restyle
    wb = load_workbook(file_path)
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
    column_widths = {"A": 30, "B": 25, "C": 20, "D": 35, "E": 15, "F": 12, "G": 12, "H": 60, "I": 18, "J": 30, "K": 25}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        
    wb.save(file_path)
    print("AI-Based mapping and styling complete.")

if __name__ == "__main__":
    apply_ai_mapping(r"E:\Internship\PocketFM\Knight Agency.xlsx")
