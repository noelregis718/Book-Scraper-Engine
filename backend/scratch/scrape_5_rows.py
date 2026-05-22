import asyncio
import os
import sys
from openpyxl import load_workbook

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from goodreads_scraper import GoodreadsScraper
from romantasy_analyzer import identify_romantasy_subgenre
from format_1488_catalog import format_excel_sheet
from playwright.async_api import async_playwright

CATALOG_FILE = r"E:\Internship\PocketFM\kensington_authors_MASTER_1_to_1488.xlsx"

async def process_book(row_idx, series_name, author_name, context, gr_scraper, ws, sem):
    async with sem:
        print(f"[Row {row_idx}] Searching for: '{series_name}' by '{author_name}'...")
        try:
            gr_data = await gr_scraper.scrape_goodreads_data(context, str(series_name), str(author_name))
            
            if gr_data:
                series_link = gr_data.get("GoodReads_Series_URL")
                book_link = gr_data.get("GoodReads_Book_URL")
                final_link = series_link if series_link and series_link != "N/A" else (book_link or "N/A")
                
                synopsis = gr_data.get("Description", "N/A")
                rom_sub = identify_romantasy_subgenre(synopsis, gr_data.get("Genre", ""))
                
                ws.cell(row=row_idx, column=4, value=final_link)
                ws.cell(row=row_idx, column=5, value=gr_data.get("Num_Primary_Books", 1) if final_link != "N/A" else "N/A")
                ws.cell(row=row_idx, column=6, value=gr_data.get("GoodReads_Rating", "N/A"))
                ws.cell(row=row_idx, column=7, value=gr_data.get("GoodReads_Rating_Count", "N/A"))
                ws.cell(row=row_idx, column=8, value=synopsis)
                ws.cell(row=row_idx, column=9, value="Yes" if rom_sub != "N/A" else "No")
                ws.cell(row=row_idx, column=10, value=rom_sub)
                
                print(f"    [Success] '{series_name}' mapped to URL!")
            else:
                print(f"    [Skip] Could not locate exact match for '{series_name}'.")
                
        except Exception as e:
            print(f"    [Error] Processing '{series_name}': {e}")

async def run_5_rows():
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active
    
    # Rows 1697 to 1746 inclusive (next 50 rows)
    target_rows = list(range(1697, 1747))
    
    targets = []
    for r in target_rows:
        title = ws.cell(row=r, column=1).value
        author = ws.cell(row=r, column=2).value
        link = ws.cell(row=r, column=4).value
        
        # We will process it regardless of link status to strictly obey the user's explicit row range command,
        # but only if title and author exist.
        if title and author:
            targets.append((r, str(title).strip(), str(author).strip()))
            
    print(f"Targeting {len(targets)} exact rows: {target_rows}")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        gr_scraper = GoodreadsScraper(headless=False)
        sem = asyncio.Semaphore(5)
        
        tasks = [process_book(r[0], r[1], r[2], context, gr_scraper, ws, sem) for r in targets]
        await asyncio.gather(*tasks)
            
        await browser.close()
        
    format_excel_sheet(ws)
    wb.save(CATALOG_FILE)
    print("Done and saved.")
    if os.name == 'nt':
        os.startfile(CATALOG_FILE)
    
if __name__ == '__main__':
    asyncio.run(run_5_rows())
