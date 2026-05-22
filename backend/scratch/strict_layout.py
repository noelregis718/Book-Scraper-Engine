from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def apply_strict_boundary_layout(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    print(f"Applying strict boundary layout for {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Apply fixed height and internal wrapping (trick to prevent spillover)
    for row_idx in range(1, ws.max_row + 1):
        # Force a thin height for every row
        ws.row_dimensions[row_idx].height = 15
        
        for cell in ws[row_idx]:
            # Setting wrap_text=True with a FIXED row height prevents horizontal spillover
            # while keeping the row thin.
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            cell.border = thin_border

    # 2. Set column widths to match the image style (wider for titles/links)
    col_widths = {
        'Subgenre': 20,
        'Title': 40,
        'URL': 30,
        'Author': 25,
        'Synopsis/Summary': 50,
        'Goodread Link': 35,
        'Series Link': 35
    }

    for idx, col in enumerate(ws.columns):
        header_val = col[0].value
        column_letter = get_column_letter(idx + 1)
        
        # Use predefined width or default to 15
        width = col_widths.get(header_val, 15)
        ws.column_dimensions[column_letter].width = width

    wb.save(file_path)
    print("Strict layout applied! Opening file...")
    if os.name == 'nt':
        os.startfile(file_path)

if __name__ == "__main__":
    target = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"
    apply_strict_boundary_layout(target)
