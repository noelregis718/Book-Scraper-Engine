import asyncio
import json
import os
import pandas as pd
import sys
from playwright.async_api import async_playwright
import easyocr
from goodreads_scraper import GoodreadsScraper
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
BASE_URL = "https://sbrmedia.com/authors/"
OUTPUT_FILE = r"E:\Internship\PocketFM\SBR Media.xlsx"
STATE_FILE = r"E:\Internship\PocketFM\backend\sbr_state.json"
TEMP_IMG_DIR = r"E:\Internship\PocketFM\backend\scratch\ocr_temp"
HEADLESS = False

if not os.path.exists(TEMP_IMG_DIR):
    os.makedirs(TEMP_IMG_DIR)

# Initialize EasyOCR once
print("[System] Loading OCR Engine...", flush=True)
READER = easyocr.Reader(['en'], gpu=False)

def format_excel(file_path):
    if not os.path.exists(file_path): return
    try:
        df = pd.read_excel(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SBR Media Catalog')
            ws = writer.sheets['SBR Media Catalog']
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = "A2"
    except: pass

async def get_all_authors(page):
    print(f"[SBR Media] Scanning authors...", flush=True)
    await page.goto(BASE_URL, wait_until="domcontentloaded")
    authors = await page.evaluate("""() => {
        return Array.from(document.querySelectorAll('a'))
            .filter(a => a.href.includes('/authors/') && a.href !== 'https://sbrmedia.com/authors/' && a.innerText.trim().length > 2)
            .map(a => ({name: a.innerText.trim(), url: a.href}));
    }""")
    return authors

async def scrape_author_visual(page, author_name, author_url):
    print(f"  [Scraping] {author_name} (Visual OCR)...", flush=True)
    try:
        await page.goto(author_url, wait_until="networkidle", timeout=60000)
    except:
        return []

    # 1. Check if the book carousel even exists
    carousel_selector = ".elementor-image-carousel-wrapper, .elementor-swiper"
    carousel = await page.query_selector(carousel_selector)
    if not carousel:
        print(f"    [Skip] No book carousel found for {author_name}.", flush=True)
        return []

    # 2. Interact with carousel to reveal all books
    next_btn_selector = ".elementor-swiper-button-next"
    titles = set()
    
    for i in range(1): # Only check the first slide
        # ONLY grab images inside the carousel
        visible_images = await carousel.query_selector_all("img")
        for idx, img in enumerate(visible_images):
            if await img.is_visible():
                box = await img.bounding_box()
                # Book covers are usually TALL (height > width)
                if box and box['width'] > 40 and box['height'] > box['width'] * 1.2:
                    img_path = os.path.join(TEMP_IMG_DIR, f"single_test.png")
                    try:
                        await img.screenshot(path=img_path)
                        results = READER.readtext(img_path)
                        cover_text = " ".join([res[1] for res in results if res[2] > 0.3])
                        if len(cover_text) > 4:
                            clean = cover_text.replace(author_name, "").strip()
                            print(f"      [Single Test] Visually Detected: {clean}")
                            titles.add(clean)
                            # STOP after finding the first valid book
                            return list(titles)
                    except: pass
        
        next_btn = await page.query_selector(next_btn_selector)
        if next_btn and await next_btn.is_visible():
            await next_btn.click()
            await asyncio.sleep(1.2)
        else:
            break
            
    print(f"    -> Extracted {len(titles)} titles visually.", flush=True)
    return list(titles)

async def run_visual_mission(limit=5):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context()
        page = await context.new_page()
        
        gr_scraper = GoodreadsScraper(headless=HEADLESS)
        authors = await get_all_authors(page)
        
        # Limit to first 5 for testing as requested
        target_authors = authors[:limit]
        print(f"--- Starting Visual Mission for first {limit} authors ---", flush=True)
        
        all_results = []
        
        for author in target_authors:
            author_name = author['name']
            author_url = author['url']
            
            book_titles = await scrape_author_visual(page, author_name, author_url)
            
            for title in book_titles:
                print(f"    [Goodreads] Searching: {title}...", flush=True)
                try:
                    gr_data = await gr_scraper.scrape_goodreads_data(context, title, author_name)
                    if gr_data:
                        all_results.append({
                            'Name of Series': gr_data.get('Book_Title', title),
                            'Author Name': author_name,
                            'Publisher': 'SBR Media',
                            'GoodReads series link': gr_data.get('GoodReads_Series_URL', 'N/A'),
                            'Rating (out of 5) of Primary Book 1': gr_data.get('Book1_Rating', 'N/A'),
                            'Ratings (#) of Primary Book 1': gr_data.get('Book1_Num_Ratings', 'N/A'),
                            'Synopsis (if available)': gr_data.get('Description', 'N/A'),
                            'Is it Romantasy ?': gr_data.get('Romantasy_Subgenre', 'No'),
                            'Name of agent': 'SBR Media'
                        })
                except: pass
        
        if all_results:
            df_new = pd.DataFrame(all_results)
            if os.path.exists(OUTPUT_FILE):
                df_old = pd.read_excel(OUTPUT_FILE)
                df_final = pd.concat([df_old, df_new], ignore_index=True).drop_duplicates(subset=['Name of Series'], keep='last')
            else:
                df_final = df_new
            
            df_final.to_excel(OUTPUT_FILE, index=False)
            format_excel(OUTPUT_FILE)
            print(f"Mission Batch Complete. Saved to {OUTPUT_FILE}", flush=True)
            os.startfile(OUTPUT_FILE)
            
        await browser.close()

if __name__ == "__main__":
    asyncio.run(run_visual_mission(limit=5))
