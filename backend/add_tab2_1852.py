import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def add_tab2_and_style():
    file_path = "e:/Internship/PocketFM/1852 Media.xlsx"
    print(f"Loading {file_path} into pandas...")
    df = pd.read_excel(file_path, sheet_name=0, engine="openpyxl")
    
    # Filter for > 4 primary books
    df['books_numeric'] = pd.to_numeric(df['Number of PRIMARY books in the series'], errors='coerce')
    filtered_df = df[df['books_numeric'] > 4].copy()
    filtered_df.drop(columns=['books_numeric'], inplace=True)
    
    print(f"Filtered down to {len(filtered_df)} rows with > 4 primary books.")
    
    print("Opening with openpyxl...")
    wb = openpyxl.load_workbook(file_path)
    
    # Remove the tab if it already exists
    if "Series > 4" in wb.sheetnames:
        del wb["Series > 4"]
        
    ws_new = wb.create_sheet("Series > 4")
    
    # Write dataframe to the new sheet
    for r in dataframe_to_rows(filtered_df, index=False, header=True):
        ws_new.append(r)
        
    # Apply premium styles to the new sheet
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") 
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") 
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), 
                         right=Side(style='thin', color='E2E8F0'), 
                         top=Side(style='thin', color='E2E8F0'), 
                         bottom=Side(style='thin', color='E2E8F0'))

    ws_new.freeze_panes = "A2"

    for row_idx, row in enumerate(ws_new.iter_rows()):
        is_header = (row_idx == 0)
        fill_color = header_fill if is_header else (zebra_fill if row_idx % 2 == 1 else white_fill)
        
        for cell in row:
            if is_header:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
            cell.fill = fill_color
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws_new.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) < 150:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 45)
        ws_new.column_dimensions[col_letter].width = adjusted_width

    print("Saving changes...")
    wb.save(file_path)
    print("Success! Second tab created and styled.")

if __name__ == "__main__":
    add_tab2_and_style()
