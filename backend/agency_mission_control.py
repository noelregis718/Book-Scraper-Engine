import asyncio
import os
import re
import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Import existing scrapers
try:
    from goodreads_scraper import GoodreadsScraper
    from scraper import AuthorScraper, clean_text
except ImportError:
    def clean_text(text):
        return str(text).strip() if text else "N/A"
    class GoodreadsScraper:
        def __init__(self, **kwargs): pass
        async def scrape_goodreads_data(self, *args, **kwargs): return {}
    class AuthorScraper:
        def __init__(self, **kwargs): pass
        async def find_author_details(self, *args, **kwargs): return {}

import sys
import json

try:
    from ai_classifier import identify_subgenre
except ImportError:
    def identify_subgenre(synopsis, tags): return "N/A"

# --- UNIVERSAL CONFIGURATION ---
AGENCY_NAME = "Knight Agency"
if len(sys.argv) > 1:
    AGENCY_NAME = sys.argv[1]

# Dynamic Pathing
SAVE_FILE = rf"E:\Internship\PocketFM\{AGENCY_NAME}.xlsx"
MAX_CONCURRENT_TABS = 15

# Default URL for fallback
DEFAULT_URL = "https://knightagency.net/ourbooks/?product_cat=romantic-suspense"
if len(sys.argv) > 2:
    DEFAULT_URL = sys.argv[2]

def clean_name(name):
    """Removes platform fluff like (Audiobook), (paperback) from names."""
    if not name: return "N/A"
    return re.sub(r'\s*\(.*?\)', '', name).strip()

# --- STATE MANAGEMENT ---
STATE_FILE = r"e:\Internship\PocketFM\backend\agency_mission_state.json"
STATE_KEY = AGENCY_NAME
if len(sys.argv) > 3:
    STATE_KEY = sys.argv[3]

def load_agency_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                content = f.read().strip()
                if not content: return {"last_page": 1}
                state = json.loads(content)
                return state.get(STATE_KEY, {"last_page": 1})
        except:
            return {"last_page": 1}
    return {"last_page": 1}

def save_agency_state(page_num):
    state = {}
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                state = json.load(f)
        except: pass
    state[STATE_KEY] = {"last_page": page_num}
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=4)

class AgencyMissionControl:
    def __init__(self, headless=True):
        self.headless = headless
        self.gr_scraper = GoodreadsScraper(headless=headless)
        self.author_scraper = AuthorScraper(headless=headless)
        self.semaphore = asyncio.Semaphore(MAX_CONCURRENT_TABS)
        self.seen_books = set()
        self.ensure_schema()
        self._load_existing_books()

    def ensure_schema(self):
        """Ensures the Excel file matches the 11-column schema."""
        if not os.path.exists(SAVE_FILE): return
        
        try:
            df = pd.read_excel(SAVE_FILE)
            if "Is it Romantasy ?" not in df.columns:
                print(f"  [System] Updating schema for {SAVE_FILE}...")
                columns = [
                    "Name of Series", "Author Name", "Publisher", "GoodReads series link",
                    "Number of PRIMARY books in the series", "Rating (out of 5) of Primary Book 1",
                    "Ratings (#) of Primary Book 1", "Synopsis (if available)",
                    "Is it Romantasy ?", "Romantasy Sub-Genre of series", "Name of agent"
                ]
                
                # Insert the column if missing
                if "Romantasy Sub-Genre of series" in df.columns:
                    idx = list(df.columns).index("Romantasy Sub-Genre of series")
                    df.insert(idx, "Is it Romantasy ?", df["Romantasy Sub-Genre of series"].apply(
                        lambda x: "Yes" if str(x) != "N/A" and str(x) != "nan" else "No"
                    ))
                
                # Reindex to ensure all columns exist and are in order
                df = df.reindex(columns=columns)
                df.to_excel(SAVE_FILE, index=False)
                self.style_excel()
                print("  [System] Schema update complete.")
        except Exception as e:
            print(f"  [Warning] Schema check failed: {e}")

    def _load_existing_books(self):
        """Loads already scraped books to prevent duplicates."""
        if os.path.exists(SAVE_FILE):
            try:
                df = pd.read_excel(SAVE_FILE)
                for _, row in df.iterrows():
                    key = f"{str(row['Name of Series']).strip()}|{str(row['Author Name']).strip()}".lower()
                    self.seen_books.add(key)
                print(f"  [System] Loaded {len(self.seen_books)} existing books from {SAVE_FILE}.")
            except Exception:
                pass

    async def run_mission(self, context, start_url, target_count=50):
        """Scrapes the catalog and enriches data, resuming from state."""
    async def run_mission(self, context, url, target_count=50):
        """Standard mission logic for any agency catalog URL."""
        new_books_gathered = 0
        state = load_agency_state()
        current_page = state.get("last_page", 1)
        
        try:
            page = await context.new_page()
            
            while new_books_gathered < target_count:
                # Build paginated URL
                target_url = url
                if "dijkstraagency.com" in url:
                    # Dijkstra uses a single long page; skip pagination logic
                    pass
                elif current_page > 1:
                    if "?" in url:
                        target_url = url.replace("?", f"page/{current_page}/?")
                    else:
                        target_url = f"{url.rstrip('/')}/page/{current_page}/"
                
                print(f"    Scanning Page {current_page}: {target_url}", flush=True)
                
                try:
                    await page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
                    if await page.query_selector('text="Wait until you are past the block"'):
                        print("\n" + "!" * 60)
                        print("  [BLOCK DETECTED] Cloudflare/Antibot active.")
                        print("  Please solve the CAPTCHA in the browser window.")
                        print("  The script will resume automatically once you are past the block.")
                        print("!" * 60 + "\n")
                        await page.wait_for_selector('.product, #searchform, .next', timeout=300000)
                        print("  [OK] Block cleared! Resuming...")
                except Exception as e:
                    print(f"    [Warning] Could not load Page {current_page}: {e}")
                    await asyncio.sleep(5)
                    
                await asyncio.sleep(2)
                
                # Scroll loop to ensure all lazy-loaded products are revealed
                print(f"    [Discovery] Scrolling Page {current_page} deeply to reveal all titles...", flush=True)
                for i in range(10):
                    await page.evaluate("window.scrollBy(0, 1000)")
                    await asyncio.sleep(1.2)
                
                await page.wait_for_timeout(3000)

                page_leads = []
                books_elements = []
                if "dijkstraagency.com" in url:
                    # Specialized logic for Dijkstra Agency (Single Page, Link-based)
                    print(f"    [Discovery] Scanning Dijkstra Agency structure...")
                    wraps = await page.query_selector_all('.books-by-subject-wrap')
                    for wrap in wraps:
                        try:
                            # Title Extraction
                            title_el = await wrap.query_selector('.book_title_list')
                            if not title_el:
                                title_el = await wrap.query_selector('a[href*="book-page.php"]')
                            
                            if not title_el: continue
                            title = (await title_el.inner_text()).strip()
                            if not title: continue
                            
                            # Author Extraction
                            author_el = await wrap.query_selector('a[href*="author-page.php"]')
                            author = "Unknown"
                            if author_el:
                                author = (await author_el.inner_text()).strip()
                            else:
                                # Fallback: check p tags for "By"
                                ps = await wrap.query_selector_all('p')
                                for p in ps:
                                    p_text = await p.inner_text()
                                    if "By" in p_text:
                                        author = p_text.replace("By", "").strip()
                                        break
                            
                            lead_key = f"{title}|{author}".lower()
                            if lead_key not in self.seen_books:
                                page_leads.append({"Name of Series": title, "Author Name": author})
                        except Exception as e:
                            print(f"      [Warning] Error parsing Dijkstra wrap: {e}")
                            continue
                else:
                    # Standard logic for WordPress/WooCommerce agencies (Knight Agency, etc.)
                    books_elements = await page.query_selector_all('.product')
                    if not books_elements:
                        print(f"    [End] No products found on page {current_page}.")
                        break
                    
                    for el in books_elements:
                        try:
                            title_el = await el.query_selector('h4.mfn-woo-product-title, h2, h3, .woocommerce-loop-product__title')
                            if title_el:
                                full_text = await title_el.inner_text()
                                if " by " in full_text:
                                    parts = full_text.split(" by ")
                                    title = clean_name(parts[0])
                                    author = clean_name(re.split(r'\$', parts[1])[0].split('\n')[0])
                                    lead_key = f"{title}|{author}".lower()
                                    if lead_key not in self.seen_books:
                                        page_leads.append({"Name of Series": title, "Author Name": author})
                                else:
                                    title = clean_name(full_text)
                                    # Author extraction
                                    author_el = await el.query_selector('.mfn-woo-product-author, .author, .woocommerce-loop-product__author, .product-author')
                                    author_name = "Unknown"
                                    if author_el:
                                        author_name = (await author_el.inner_text()).strip()
                                    else:
                                        # Fallback: check if author is in the title or a separate span
                                        fallback_author = await el.query_selector('span.author, .desc-author')
                                        if fallback_author:
                                            author_name = (await fallback_author.inner_text()).strip()
                                    author = clean_name(author_name)
                                    lead_key = f"{title}|{author}".lower()
                                    if lead_key not in self.seen_books:
                                        page_leads.append({"Name of Series": title, "Author Name": author})
                        except: continue

                if not page_leads:
                    print(f"    -> Page {current_page}: Found {len(books_elements)} total books. (0 NEW). Advancing...")
                    current_page += 1
                    save_agency_state(current_page)
                    continue

                print(f"    -> Page {current_page}: Found {len(books_elements)} total books. {len(page_leads)} are NEW.", flush=True)

                # Enrichment Phase for this page
                batch_results = []
                for i in range(0, len(page_leads), MAX_CONCURRENT_TABS):
                    if new_books_gathered >= target_count: break
                    
                    batch = page_leads[i : i + MAX_CONCURRENT_TABS]
                    tasks = [self.process_lead(context, lead, new_books_gathered + j + 1, target_count) for j, lead in enumerate(batch)]
                    results = await asyncio.gather(*tasks)
                    
                    valid = [r for r in results if r]
                    batch_results.extend(valid)
                    new_books_gathered += len(valid)

                if batch_results:
                    self.save_to_excel(batch_results)
                    self.seen_books.update([f"{r['Name of Series']}|{r['Author Name']}".lower() for r in batch_results])
                
                if new_books_gathered >= target_count:
                    print(f"\n[OK] Reached target batch of {target_count} books.")
                    break

                if "dijkstraagency.com" in url:
                    # Dijkstra is usually all on one page
                    print(f"\n[Mission] Dijkstra Agency scanning complete.")
                    mission_active = False 
                    break

                current_page += 1
                save_agency_state(current_page)

            await page.close()
            return new_books_gathered

        except Exception as e:
            print(f"[CRITICAL ERROR] Mission failed: {e}")
            return new_books_gathered

    async def process_lead(self, context, lead, index, total):
        async with self.semaphore:
            try:
                print(f"      [{index}/{total}] Enriching: {lead['Name of Series']}...")
                gr_data = await self.gr_scraper.scrape_goodreads_data(
                    context, lead['Name of Series'], lead['Author Name']
                )
                
                # FALLBACK: If Goodreads fails, we still record the book with N/A fields
                if not gr_data:
                    print(f"      [Notice] Goodreads missed: {lead['Name of Series']}. Recording basic info only.")
                    return {
                        "Name of Series": lead['Name of Series'],
                        "Author Name": lead['Author Name'],
                        "Publisher": "Various / Knight Agency",
                        "GoodReads series link": "N/A",
                        "Number of PRIMARY books in the series": "N/A",
                        "Rating (out of 5) of Primary Book 1": "N/A",
                        "Ratings (#) of Primary Book 1": "N/A",
                        "Synopsis (if available)": "N/A",
                        "Is it Romantasy ?": "N/A",
                        "Romantasy Sub-Genre of series": "N/A",
                        "Name of agent": "Knight Agency Representative"
                    }
                
                synopsis = gr_data.get("Description", "N/A")
                tags = [gr_data.get("Genre", ""), gr_data.get("Sub_Genre", "")]
                matched_genre = identify_subgenre(synopsis, tags)
                
                return {
                    "Name of Series": lead['Name of Series'],
                    "Author Name": lead['Author Name'],
                    "Publisher": gr_data.get("Publisher", "Various / Knight Agency"),
                    "GoodReads series link": gr_data.get("GoodReads_Series_URL", "N/A"),
                    "Number of PRIMARY books in the series": gr_data.get("Num_Primary_Books", "N/A"),
                    "Rating (out of 5) of Primary Book 1": gr_data.get("Book1_Rating", "N/A"),
                    "Ratings (#) of Primary Book 1": gr_data.get("Book1_Num_Ratings", "N/A"),
                    "Synopsis (if available)": synopsis[:1000] if synopsis != "N/A" else "N/A",
                    "Is it Romantasy ?": "Yes" if matched_genre != "N/A" else "No",
                    "Romantasy Sub-Genre of series": matched_genre,
                    "Name of agent": "Knight Agency Representative"
                }
            except Exception as e:
                print(f"      [Error] Failed to process {lead['Name of Series']}: {e}")
                return None

    def save_to_excel(self, data):
        """Incremental save to Excel with lock protection."""
        if not data: return
        import time
        
        columns = [
            "Name of Series", "Author Name", "Publisher", "GoodReads series link",
            "Number of PRIMARY books in the series", "Rating (out of 5) of Primary Book 1",
            "Ratings (#) of Primary Book 1", "Synopsis (if available)",
            "Is it Romantasy ?", "Romantasy Sub-Genre of series", "Name of agent"
        ]
        
        df_new = pd.DataFrame(data)
        df_new = df_new.reindex(columns=columns)
        
        max_retries = 10
        for attempt in range(max_retries):
            try:
                if not os.path.exists(SAVE_FILE):
                    df_new.to_excel(SAVE_FILE, index=False, header=True)
                else:
                    existing_df = pd.read_excel(SAVE_FILE)
                    combined_df = pd.concat([existing_df, df_new], ignore_index=True).drop_duplicates(subset=['Name of Series', 'Author Name'], keep='last')
                    combined_df = combined_df.reindex(columns=columns)
                    combined_df.to_excel(SAVE_FILE, index=False, header=True)
                
                self.style_excel()
                return # Success
            except PermissionError:
                print(f"\n[!!!] PERMISSION DENIED: Please CLOSE '{os.path.basename(SAVE_FILE)}' so I can save! (Attempt {attempt+1}/{max_retries})")
                if attempt < max_retries - 1:
                    time.sleep(10)
                else:
                    recovery_path = SAVE_FILE.replace(".xlsx", f"_RECOVERY_{int(time.time())}.xlsx")
                    print(f"[Critical] Saving to recovery file: {recovery_path}")
                    df_new.to_excel(recovery_path, index=False)
            except Exception as e:
                print(f"    [Warning] Save failed: {e}")
                break

    def style_excel(self):
        try:
            if not os.path.exists(SAVE_FILE): return
            wb = load_workbook(SAVE_FILE)
            if not wb: return
            ws = wb.active
            if not ws: return
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Style Headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Style Data Rows (Wrapping and Alignment)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Adjust Column Widths
            column_widths = {
                "A": 30, # Name of Series
                "B": 25, # Author Name
                "C": 20, # Publisher
                "D": 35, # GoodReads link
                "E": 15, # Num Books
                "F": 12, # Rating
                "G": 12, # Num Ratings
                "H": 60, # Synopsis
                "I": 18, # Is it Romantasy ?
                "J": 30, # Sub-Genre
                "K": 25  # Agent
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
                
            wb.save(SAVE_FILE)
        except Exception as e:
            print(f"    [Warning] Styling failed: {e}")

async def main():
    control = AgencyMissionControl(headless=False)
    async with async_playwright() as p:
        # Launch using actual Chrome channel for better bypass
        browser = await p.chromium.launch(
            headless=False,
            channel="chrome" 
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
            device_scale_factor=1,
            is_mobile=False,
            has_touch=False,
            locale="en-US",
            timezone_id="America/New_York"
        )
        
        print(f"\n>>> [UNIVERSAL MISSION] Agency: {AGENCY_NAME} (State Key: {STATE_KEY})", flush=True)
        print(f">>> [TARGET URL] {DEFAULT_URL}\n", flush=True)
        
        # Industrial Looping: Process in batches of 50 until no more books are found
        mission_active = True
        total_session_new = 0
        
        while mission_active:
            print(f"\n--- [Industrial Batch Start] Target: 50 NEW books ---")
            new_gathered = await control.run_mission(context, DEFAULT_URL, target_count=50)
            total_session_new += new_gathered
            
            # Final Deduplication & Styling per batch
            if os.path.exists(SAVE_FILE):
                max_retries = 5
                for attempt in range(max_retries):
                    try:
                        final_df = pd.read_excel(SAVE_FILE)
                        final_df.drop_duplicates(subset=['Name of Series', 'Author Name'], keep='first', inplace=True)
                        final_df.to_excel(SAVE_FILE, index=False)
                        control.style_excel()
                        
                        # Auto-open the Excel sheet as requested after each batch
                        if os.name == 'nt': 
                            print(f"  [Mission] Opening updated results for batch...")
                            os.startfile(SAVE_FILE)
                        break
                    except PermissionError:
                        print(f"  [Warning] Could not finalize Excel (Permission Denied). Retrying in 10s... ({attempt+1}/{max_retries})")
                        await asyncio.sleep(10)
                    except Exception as e:
                        print(f"  [Warning] Finalization failed: {e}")
                        break
            
            if new_gathered < 1:
                # Check if we actually reached the end or just no new books on these pages
                # If run_mission returns 0, it means it either hit the end of pages or target was 0
                print("\n[Mission] No more NEW books found or end of catalog reached.")
                mission_active = False
            else:
                print(f"  [Mission] Batch complete ({new_gathered} NEW). Continuing mission...")
                # Small cool-down between batches
                await asyncio.sleep(5)

        await browser.close()
        print(f"\n{'='*60}")
        print(f"MISSION ACCOMPLISHED: {total_session_new} New Titles added in this session.")
        print(f"Final Data: {os.path.abspath(SAVE_FILE)}")
        print(f"{'='*60}\n")

if __name__ == "__main__":
    asyncio.run(main())
