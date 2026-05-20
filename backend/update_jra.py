import pandas as pd
from openpyxl import load_workbook

def update_jra_excel(input_file):
    # Load the workbook and select the active worksheet
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Get column indices
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    publisher_col = headers.get("Publisher")
    agent_col = headers.get("Name of agent")
    
    if publisher_col and agent_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=publisher_col, value="Jane Rotrosen Agency")
            ws.cell(row=row, column=agent_col, value="Meg Ruley")
            
    wb.save(input_file)

if __name__ == "__main__":
    update_jra_excel('JRA_Bestsellers_Complete.xlsx')
    print("Update complete.")
