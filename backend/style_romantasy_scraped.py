import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

file_path = r'e:\Internship\PocketFM\Romantasy_v2_Scraped.xlsx'
if not os.path.exists(file_path):
    print(f"File {file_path} not found!")
    exit(1)

print("Loading workbook...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define styles
header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                     right=Side(style='thin', color='BFBFBF'), 
                     top=Side(style='thin', color='BFBFBF'), 
                     bottom=Side(style='thin', color='BFBFBF'))

hyperlink_font = Font(color="0563C1", underline="single", name='Calibri', size=11)
regular_font = Font(name='Calibri', size=11)

print("Formatting header...")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = thin_border

print("Formatting rows and columns...")
column_widths = {
    'A': 50, # GR Book 1 link
    'B': 25, # Agency (if)
    'C': 50, # GR Series Link
    'D': 25, # No. of books in the series
    'E': 15, # Page count
}

for i, col in enumerate(range(1, ws.max_column + 1)):
    letter = get_column_letter(col)
    if letter in column_widths:
        ws.column_dimensions[letter].width = column_widths[letter]

print("Applying data styles...")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment_left_top
        cell.border = thin_border
        
        # Format hyperlinks in Column A (1) and Column C (3)
        if cell.column in [1, 3] and cell.value and str(cell.value).startswith('http'):
            cell.font = hyperlink_font
            cell.hyperlink = cell.value
        else:
            cell.font = regular_font

# Freeze header
ws.freeze_panes = 'A2'

print("Saving workbook...")
wb.save(file_path)
print("Done styling!")
