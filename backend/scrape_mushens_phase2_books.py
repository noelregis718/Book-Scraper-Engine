"""
PHASE 2 of the Mushens full-roster pipeline.

Reads the "All Authors" sheet (117 authors), and for each author:
  1. Searches Goodreads
  2. Navigates to that author's profile page
  3. Pulls their first 3 books (title + link)
  4. Visits each of those 3 book pages, extracts metadata
  5. Appends a new row to the main "Mushens Entertainment" sheet — deduped
     by (title, author) so existing rows aren't duplicated
  6. Updates the author's row in "All Authors" with their Goodreads URL + a
     "done N/3" or "not found" status

Concurrency: 10 author tabs at a time inside one shared browser context.
"""

import asyncio
import io
import json
import os
import re
import shutil
import subprocess
import sys

import openpyxl
from playwright.async_api import async_playwright

# Force stdout to UTF-8 so non-ASCII author names don't crash the script on cp1252 consoles
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper, clean_text
from format_mushens import format_mushens


BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "Mushens_Entertainment_Bestsellers.xlsx")
MAIN_SHEET    = "Mushens Entertainment"
AUTHORS_SHEET = "All Authors"

BATCH_SIZE       = 20   # process 20 authors per batch
CONCURRENCY      = 20   # all 20 in the batch run concurrently
BATCH_PAUSE_SEC  = 30   # pause between batches so Goodreads doesn't rate-limit
BOOKS_PER_AUTHOR = 3


# ----------------------------------------------------------------- helpers
def normalize(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def load_state(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)

    ws_main = wb[MAIN_SHEET]
    existing_keys = set()
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        title = row[0] if len(row) > 0 else None
        author = row[1] if len(row) > 1 else None
        if title:
            existing_keys.add((normalize(title), normalize(author)))

    ws_auth = wb[AUTHORS_SHEET]
    authors = []
    for r in range(2, ws_auth.max_row + 1):
        name = ws_auth.cell(row=r, column=2).value
        status = str(ws_auth.cell(row=r, column=5).value or "").lower()
        # Re-process anything that isn't fully done (done X/Y means it succeeded)
        if name and not status.startswith("done "):
            authors.append({"row": r, "name": str(name).strip()})
    return wb, existing_keys, authors


# ----------------------------------------------------------------- scraping
async def find_author_page(page, author_name):
    """Navigate to a Goodreads author profile page; return its URL or None.
    Tries 3 strategies in order:
      1. People-specific search (search_type=people)
      2. General search, scan all /author/show/ links on the page
      3. First book result -> open it -> grab the author link from the book page
    """
    q = author_name.replace(" ", "+")

    async def _scan_for_author_links():
        anchors = await page.query_selector_all('a[href*="/author/show/"]')
        for a in anchors[:10]:
            try:
                href = await a.evaluate("el => el.href")
                if "/author/show/" in href:
                    return href
            except Exception:
                continue
        return None

    # Strategy 1: people search
    try:
        await page.goto(f"https://www.goodreads.com/search?q={q}&search_type=people",
                        wait_until="domcontentloaded", timeout=45000)
        await asyncio.sleep(2.0)
        found = await _scan_for_author_links()
        if found:
            return found
    except Exception:
        pass

    # Strategy 2: general search
    try:
        await page.goto(f"https://www.goodreads.com/search?q={q}",
                        wait_until="domcontentloaded", timeout=45000)
        await asyncio.sleep(2.0)
        found = await _scan_for_author_links()
        if found:
            return found
    except Exception:
        pass

    # Strategy 3: open the first book result and pull the author link from there
    try:
        first_book = await page.query_selector(
            'a.bookTitle, [data-testid="bookTitle"] a, a[href*="/book/show/"]')
        if first_book:
            book_url = await first_book.evaluate("el => el.href")
            if "/book/show/" in book_url:
                await page.goto(book_url, wait_until="domcontentloaded", timeout=45000)
                await asyncio.sleep(1.5)
                author_el = await page.query_selector(
                    'a[href*="/author/show/"], .authorName, [data-testid="authorName"]')
                if author_el:
                    author_url = await author_el.evaluate("el => el.href")
                    if "/author/show/" in author_url:
                        return author_url
    except Exception:
        pass

    return None


async def list_first_books(page, author_url, max_books):
    """Open author page; return list of {'title','link'} for first N distinct books."""
    await page.goto(author_url, wait_until="domcontentloaded", timeout=45000)
    await asyncio.sleep(2.0)

    anchors = await page.query_selector_all('a.bookTitle, [data-testid="bookTitle"] a, a[href*="/book/show/"]')
    seen = set()
    results = []
    for el in anchors:
        try:
            href = await el.evaluate("el => el.href")
            if "/book/show/" not in href:
                continue
            t = clean_text(await el.inner_text())
            if not t:
                continue
            if href in seen:
                continue
            seen.add(href)
            results.append({"title": t, "link": href})
            if len(results) >= max_books:
                break
        except Exception:
            continue
    return results


async def extract_book(page, book_url):
    """Visit a book page and return scraped fields."""
    await page.goto(book_url, wait_until="domcontentloaded", timeout=60000)
    await asyncio.sleep(0.6)

    # Genres
    genres = []
    gel = await page.query_selector_all('[data-testid="genresList"] .Button__labelItem, .BookPageMetadataSection__genre a')
    for g in gel:
        try:
            t = clean_text(await g.inner_text())
            if t and t not in genres:
                genres.append(t)
        except Exception:
            continue
    is_romantasy = "Yes" if any("romantasy" in x.lower() for x in genres) else "No"
    sub_genre = genres[1] if len(genres) > 1 else (genres[0] if genres else "N/A")

    # Rating / Count via JSON-LD
    avg_rating = "N/A"
    rating_count = "N/A"
    try:
        ld_el = await page.query_selector('script[type="application/ld+json"]')
        if ld_el:
            ld = json.loads(await ld_el.inner_text())
            if isinstance(ld, list):
                ld = ld[0]
            avg_rating   = str(ld.get("aggregateRating", {}).get("ratingValue",  "N/A"))
            rating_count = str(ld.get("aggregateRating", {}).get("ratingCount", "N/A"))
    except Exception:
        pass

    description = "N/A"
    desc_el = await page.query_selector('[data-testid="description"] .Formatted, .readable')
    if desc_el:
        description = clean_text(await desc_el.inner_text())

    series_url = "N/A"
    series_link = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a')
    if series_link:
        try:
            series_url = await series_link.evaluate("el => el.href")
        except Exception:
            pass

    num_primary = "1"
    if series_url != "N/A":
        try:
            await page.goto(series_url, wait_until="domcontentloaded", timeout=60000)
            content = await page.content()
            m = re.search(r"(\d+)\s+primary\s+works", content, re.IGNORECASE)
            if m:
                num_primary = m.group(1)
        except Exception:
            pass

    final_link = series_url if series_url != "N/A" else book_url

    return {
        "link":         final_link,
        "num_primary":  num_primary,
        "rating":       avg_rating,
        "ratings_n":    rating_count,
        "synopsis":     description,
        "romantasy":    is_romantasy,
        "sub_genre":    sub_genre,
    }


# ----------------------------------------------------------------- per-author worker
async def process_author(context, wb, author, semaphore, existing_keys, lock, xlsx_path):
    async with semaphore:
        name = author["name"]
        row_idx = author["row"]
        print(f"  [Author] {name}")

        page = await context.new_page()
        try:
            await page.set_extra_http_headers({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            })

            author_url = await find_author_page(page, name)
            if not author_url:
                async with lock:
                    wb[AUTHORS_SHEET].cell(row=row_idx, column=5).value = "not found"
                    try:
                        wb.save(xlsx_path)
                    except Exception as e:
                        print(f"    [Save Error] {e}")
                print(f"    [Skip] no author page for {name}")
                return

            async with lock:
                wb[AUTHORS_SHEET].cell(row=row_idx, column=4).value = author_url

            books = await list_first_books(page, author_url, BOOKS_PER_AUTHOR)
            if not books:
                async with lock:
                    wb[AUTHORS_SHEET].cell(row=row_idx, column=5).value = "no books"
                    try:
                        wb.save(xlsx_path)
                    except Exception as e:
                        print(f"    [Save Error] {e}")
                return

            added = 0
            for b in books:
                title = b["title"]
                key = (normalize(title), normalize(name))
                if key in existing_keys:
                    print(f"    [Dup] '{title}' already in main sheet")
                    continue

                try:
                    details = await extract_book(page, b["link"])
                except Exception as e:
                    print(f"    [Book Error] '{title}': {e}")
                    continue

                async with lock:
                    if key in existing_keys:
                        continue
                    existing_keys.add(key)

                    ws_main = wb[MAIN_SHEET]
                    ws_main.append([
                        title,
                        name,
                        "Mushens Entertainment",
                        details["link"],
                        details["num_primary"],
                        details["rating"],
                        details["ratings_n"],
                        details["synopsis"],
                        details["romantasy"],
                        details["sub_genre"],
                        "Juliet Mushens",
                    ])
                    added += 1
                    try:
                        wb.save(xlsx_path)
                    except Exception as e:
                        print(f"    [Save Error] {e}")

                print(f"    [OK] '{title}' -> {details['link']}")

            async with lock:
                wb[AUTHORS_SHEET].cell(row=row_idx, column=5).value = f"done {added}/{len(books)}"
                try:
                    wb.save(xlsx_path)
                except Exception as e:
                    print(f"    [Save Error] {e}")

        except Exception as e:
            print(f"  [Error] {name}: {e}")
            async with lock:
                try:
                    wb[AUTHORS_SHEET].cell(row=row_idx, column=5).value = f"error: {str(e)[:40]}"
                    wb.save(xlsx_path)
                except Exception:
                    pass
        finally:
            try:
                await page.close()
            except Exception:
                pass


# ----------------------------------------------------------------- main
async def run():
    if not os.path.exists(XLSX):
        print(f"Workbook not found: {XLSX}")
        return

    wb, existing_keys, authors = load_state(XLSX)
    print(f"Existing (title, author) pairs in main sheet: {len(existing_keys)}")
    print(f"Authors to process: {len(authors)}\n")
    if not authors:
        print("Nothing to do.")
        return

    # Chunk into batches of BATCH_SIZE
    batches = [authors[i:i + BATCH_SIZE] for i in range(0, len(authors), BATCH_SIZE)]
    print(f"Will process {len(batches)} batches of up to {BATCH_SIZE} authors each "
          f"(concurrency={CONCURRENCY} tabs per batch, {BATCH_PAUSE_SEC}s pause between batches).\n")

    lock      = asyncio.Lock()
    scraper   = GoodreadsScraper(headless=False)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()

        print("[System] Logging in to Goodreads...")
        login_page = await context.new_page()
        await scraper.login_to_goodreads(login_page)
        await login_page.close()
        print(f"[System] Login done.\n")

        for batch_idx, batch in enumerate(batches, start=1):
            print(f"\n========== BATCH {batch_idx}/{len(batches)}  "
                  f"({len(batch)} authors) ==========")
            semaphore = asyncio.Semaphore(CONCURRENCY)
            tasks = [
                process_author(context, wb, a, semaphore, existing_keys, lock, XLSX)
                for a in batch
            ]
            await asyncio.gather(*tasks)
            print(f"========== BATCH {batch_idx} done ==========")

            # Snapshot copy + open it for the user (keeps live file unlocked)
            snapshot = XLSX.replace(".xlsx", f"_batch{batch_idx}.xlsx")
            try:
                shutil.copy2(XLSX, snapshot)
                subprocess.Popen(["cmd", "/c", "start", "", snapshot], shell=False)
                print(f"[View] Snapshot opened: {os.path.basename(snapshot)}")
            except Exception as e:
                print(f"[View] Snapshot error: {e}")

            if batch_idx < len(batches):
                print(f"[Pause] Sleeping {BATCH_PAUSE_SEC}s before next batch...")
                await asyncio.sleep(BATCH_PAUSE_SEC)

        try:
            await browser.close()
        except Exception:
            pass

    # Re-apply main-sheet styling (preserves the All Authors tab now)
    print("\nReapplying Mushens 11-column styling...")
    format_mushens(XLSX, XLSX)
    print("Done.")


if __name__ == "__main__":
    asyncio.run(run())
