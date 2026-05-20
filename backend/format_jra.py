import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def format_jra_excel(input_file, output_file):
    # Read the data, skipping the first two title rows
    df = pd.read_excel(input_file, skiprows=2)
    
    # Columns in original: ['#', 'Author', 'Book Title (Current Bestseller)', 'Also By (Other Books Listed)', 'Genre']
    
    # Create the new dataframe with 11 columns
    new_columns = [
        "Name of Series",
        "Author Name",
        "Publisher",
        "GoodReads series link",
        "Number of PRIMARY books in the series",
        "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1",
        "Synopsis (if available)",
        "Romantasy = Yes or No?",
        "Romantasy Sub-Genre of series",
        "Name of agent"
    ]
    
    new_df = pd.DataFrame(columns=new_columns)
    
    # Map the existing data
    new_df["Name of Series"] = df["Book Title (Current Bestseller)"]
    new_df["Author Name"] = df["Author"]
    new_df["Name of agent"] = "Jane Rotrosen Agency"
    
    # Optionally, we can set Romantasy = Yes or No based on Genre
    def check_romantasy(genre):
        if not isinstance(genre, str): return ""
        g = genre.lower()
        if "paranormal" in g or "fantasy" in g or "romantasy" in g:
            return "Yes"
        return "No"
    
    new_df["Romantasy = Yes or No?"] = df["Genre"].apply(check_romantasy)
    new_df["Romantasy Sub-Genre of series"] = df["Genre"]
    
    # Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "JRA Bestsellers"
    
    for r in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(r)
        
    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Headers
    for col in range(1, len(new_columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # Format Data
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(new_columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border
            
    # Set Column Widths
    widths = {
        "A": 30,  # Name of Series
        "B": 20,  # Author Name
        "C": 20,  # Publisher
        "D": 35,  # GoodReads series link
        "E": 15,  # Number of PRIMARY books
        "F": 15,  # Rating
        "G": 15,  # Ratings (#)
        "H": 40,  # Synopsis
        "I": 15,  # Romantasy
        "J": 20,  # Sub-Genre
        "K": 25   # Name of agent
    }
    
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(output_file)

if __name__ == "__main__":
    format_jra_excel('JRA_Bestsellers_Complete.xlsx', 'JRA_Bestsellers_Formatted.xlsx')
    print("Formatting complete. Saved to JRA_Bestsellers_Formatted.xlsx")
