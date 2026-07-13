import asyncio
import os
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = '../New_Agency_Template.xlsx'

async def fetch_link(context, title, author, sem, row_idx, ws):
    async with sem:
        page = await context.new_page()
        try:
            query = f'{title} {author} site:goodreads.com/book/show'
            url = f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}"
            
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            
            # DuckDuckGo HTML result links
            links = await page.query_selector_all('a.result__url')
            found_link = None
            if links:
                for link in links:
                    text = (await link.inner_text()).strip()
                    if "goodreads.com/book/show" in text:
                        # Construct proper https url from DDG display text
                        clean_url = "https://" + text.replace(' ', '')
                        found_link = clean_url
                        break
                        
            if not found_link:
                # Try getting href from result snippet directly
                a_tags = await page.query_selector_all('a.result__snippet')
                for a in a_tags:
                    href = await a.get_attribute('href')
                    if href and "goodreads.com" in href:
                        # DDG wraps hrefs, we need to extract the actual url
                        import urllib.parse
                        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                        if 'uddg' in parsed and "goodreads.com/book/show" in parsed['uddg'][0]:
                            found_link = parsed['uddg'][0]
                            break

            if found_link:
                print(f"  -> [Row {row_idx}] Success: {found_link}")
                ws.cell(row=row_idx, column=4).value = found_link
            else:
                print(f"  -> [Row {row_idx}] Failed.")
                
        except Exception as e:
            print(f"  -> [Row {row_idx}] Error: {e}")
        finally:
            await page.close()

async def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    tasks_data = []
    
    # 4 is GoodReads series link column (D)
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        author = ws.cell(row=row, column=2).value
        link = ws.cell(row=row, column=4).value
        
        if title and str(title).strip() and str(title).strip() != "N/A":
            if not link or str(link).strip() in ["N/A", "nan", "Not Found", "Error"]:
                tasks_data.append({
                    'row': row,
                    'title': str(title).strip(),
                    'author': str(author).strip() if author else ""
                })
                
    print(f"Found {len(tasks_data)} books still missing links.")
    if not tasks_data:
        return
        
    sem = asyncio.Semaphore(10) # 10 tabs concurrently since DDG is very lenient
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        
        tasks = []
        for t in tasks_data:
            tasks.append(fetch_link(context, t['title'], t['author'], sem, t['row'], ws))
            
        print("--- Starting DDG link extraction ---")
        await asyncio.gather(*tasks)
        await browser.close()
        
    wb.save(EXCEL_FILE)
    print("Excel file updated successfully with recovered links!")

if __name__ == "__main__":
    asyncio.run(main())
