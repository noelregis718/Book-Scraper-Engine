import asyncio
import json
import os
import pandas as pd
import re
from playwright.async_api import async_playwright
from goodreads_scraper import GoodreadsScraper
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
BASE_URL = "https://awfulagent.com/jabclients/"
OUTPUT_FILE = r"E:\Internship\PocketFM\awful agents.xlsx"
STATE_FILE = r"E:\Internship\PocketFM\backend\jabberwocky_state.json"
HEADLESS = False  # User wants to see it

def format_excel(file_path):
    """Applies professional formatting to the Excel file."""
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        print(f"  [Error] Cannot format {file_path}: File is empty or missing.")
        return

    print(f"  [System] Applying professional formatting to {os.path.basename(file_path)}...")
    try:
        # Use a temporary file for formatting to prevent corruption
        temp_path = file_path + ".tmp"
        df = pd.read_excel(file_path, engine='openpyxl')
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Jabberwocky Catalog')
        
        workbook = writer.book
        worksheet = writer.sheets['Jabberwocky Catalog']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply to headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        # Column widths and alignment
        column_widths = [30, 25, 15, 45, 15, 15, 15, 60, 15, 20, 15]
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            
        # Apply alignment to all cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left_align if cell.column in [1, 2, 4, 8] else center_align
                cell.border = border

        worksheet.freeze_panes = "A2"
        writer.close()
        
        # Atomic swap
        if os.path.exists(file_path): os.remove(file_path)
        os.rename(temp_path, file_path)
        print(f"  [System] Formatting complete.")
    except Exception as e:
        print(f"  [Error] Formatting failed: {e}")

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f:
            return json.load(f)
    return {"last_author_index": 0}

def save_state(index):
    with open(STATE_FILE, 'w') as f:
        json.dump({"last_author_index": index}, f)

async def get_author_links(page):
    """Scrapes all author links from the main client list."""
    print(f"[Jabberwocky] Scanning main client list: {BASE_URL}")
    await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60000)
    
    # Authors are in a grid/list
    # The correct selector based on inspection is a.authorlink
    links = await page.query_selector_all('a.authorlink')
    author_list = []
    for link in links:
        href = await link.evaluate("el => el.href")
        name = await link.evaluate("el => el.innerText")
        if '/jabclients/' in href and href != BASE_URL:
            if name and len(name.strip().split()) >= 2:
                author_list.append({"name": name.strip(), "url": href})
    
    print(f"[Jabberwocky] Found {len(author_list)} authors.")
    return author_list

async def scrape_bibliography(page, author_name):
    """Scrapes book titles from the bibliography section of an author profile."""
    print(f"  [Scraping] {author_name} bibliography...")
    books = []
    
    # The bibliography on Jabberwocky is strictly inside <div class="book-title">
    elements = await page.query_selector_all('div.book-title')
    
    for el in elements:
        text = (await el.evaluate("el => el.innerText")).strip()
        if not text: continue
        
        # Additional safety check to avoid any stray UI text
        if any(k in text.lower() for k in ["back to top", "rights", "social media", "menu", "search", "all rights reserved"]):
            continue
            
        books.append(text)
            
    # Deduplicate titles
    unique_books = list(dict.fromkeys(books))
    if unique_books:
        print(f"    -> Found {len(unique_books)} book candidates.")
    return unique_books

async def run_mission():
    print("Starting Jabberwocky Deep-Scrape Mission...")
    
    if os.path.exists(OUTPUT_FILE):
        df = pd.read_excel(OUTPUT_FILE)
    else:
        columns = ['Name of Series', 'Author Name', 'Publisher', 'GoodReads series link', 
                   'Number of PRIMARY books in the series', 'Rating (out of 5) of Primary Book 1', 
                   'Ratings (#) of Primary Book 1', 'Synopsis (if available)', 'Is it Romantasy ?', 
                   'Romantasy Sub-Genre of series', 'Name of agent']
        df = pd.DataFrame(columns=columns)

    seen_titles = set(df['Name of Series'].astype(str).tolist())
    gr_scraper = GoodreadsScraper(headless=HEADLESS)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context()
        page = await context.new_page()

        authors = await get_author_links(page)
        state = load_state()
        start_idx = state["last_author_index"]
        
        if start_idx > 0:
            print(f"--- [Resume] Starting from Author Index {start_idx} ---")

        # Batching: 15 authors at a time
        BATCH_SIZE = 15
        for i in range(start_idx, len(authors), BATCH_SIZE):
            batch = authors[i:i+BATCH_SIZE]
            print(f"\n--- [Batch Start] Processing Authors {i+1} to {min(i+BATCH_SIZE, len(authors))} ---")
            
            for author in batch:
                author_name = author['name']
                author_url = author['url']
                
                try:
                    await page.goto(author_url, wait_until="domcontentloaded", timeout=60000)
                    titles = await scrape_bibliography(page, author_name)
                    
                    # Parallel Book Processing: 8 books at a time
                    BOOK_BATCH_SIZE = 8
                    for j in range(0, len(titles), BOOK_BATCH_SIZE):
                        book_batch = titles[j:j+BOOK_BATCH_SIZE]
                        tasks = []
                        
                        for title in book_batch:
                            clean_title = re.sub(r'\(\d{4}\)', '', title).strip()
                            if clean_title in seen_titles: continue
                            
                            print(f"    [Processing] {clean_title}...")
                            tasks.append(gr_scraper.scrape_goodreads_data(context, clean_title, author_name))
                        
                        results = await asyncio.gather(*tasks, return_exceptions=True)
                        
                        for gr_data in results:
                            if isinstance(gr_data, Exception):
                                print(f"      [Error] Book processing failed: {gr_data}")
                                continue
                                
                            if gr_data:
                                # FIX: Hardened Author Name and Goodreads link capture
                                final_author = gr_data.get('Author_Found', 'Unknown')
                                if final_author == 'Unknown':
                                    final_author = author_name # Fallback to Agency author name
                                    
                                new_row = {
                                    'Name of Series': gr_data.get('Book_Title', 'Unknown'), # Use GR title if possible
                                    'Author Name': final_author,
                                    'Publisher': 'Jabberwocky',
                                    'GoodReads series link': gr_data.get('GoodReads_Series_URL', gr_data.get('GoodReads_Book_URL', 'N/A')),
                                    'Number of PRIMARY books in the series': gr_data.get('Num_Primary_Books', '1'),
                                    'Rating (out of 5) of Primary Book 1': gr_data.get('Book1_Rating', 'N/A'),
                                    'Ratings (#) of Primary Book 1': gr_data.get('Book1_Num_Ratings', 'N/A'),
                                    'Synopsis (if available)': gr_data.get('Description', 'N/A'),
                                    'Is it Romantasy ?': gr_data.get('Romantasy_Subgenre', 'No'),
                                    'Romantasy Sub-Genre of series': gr_data.get('Genre', 'N/A'),
                                    'Name of agent': 'Jabberwocky'
                                }
                                # Ensure we don't save duplicates if gather returned multiple hits
                                series_name = new_row['Name of Series']
                                if series_name not in seen_titles:
                                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                                    seen_titles.add(series_name)
                except Exception as e:
                    print(f"  [Error] Failed to process {author_name}: {e}")

            # Save after each batch of 15 authors
            df.to_excel(OUTPUT_FILE, index=False)
            format_excel(OUTPUT_FILE)
            save_state(i + BATCH_SIZE)
            
            print(f"--- [Batch Complete] Saved and Formatted. Total books: {len(df)} ---")
            
            # AUTO-OPEN after first batch (if this was the first start)
            if i == start_idx:
                print("  [System] Auto-opening file for your review...")
                os.startfile(OUTPUT_FILE)
            
            # Short pause between batches
            await asyncio.sleep(5)

        # Final Save
        df.to_excel(OUTPUT_FILE, index=False)
        print(f"Mission Complete. Total books captured: {len(df)}")
        await browser.close()

if __name__ == "__main__":
    asyncio.run(run_mission())
