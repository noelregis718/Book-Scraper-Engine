import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

file_path = 'New_Agency_Template.xlsx'

columns = [
    'Name of Series', 'Author Name', 'Publisher', 'GoodReads series link',
    'Number of PRIMARY books in the series', 'Rating (out of 5) of Primary Book 1',
    'Ratings (#) of Primary Book 1', 'Synopsis (if available)', 'Romantasy = Yes or No?',
    'Romantasy Sub-Genre of series', 'Name of agent in the main folder'
]

try:
    # 1. Create an empty DataFrame with just these headers
    df = pd.DataFrame(columns=columns)
    df.to_excel(file_path, index=False)
    
    # 2. Add professional styling so it matches your other sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D3D3D3'), 
                         right=Side(style='thin', color='D3D3D3'), 
                         top=Side(style='thin', color='D3D3D3'), 
                         bottom=Side(style='thin', color='D3D3D3'))
                         
    # Apply styling to the header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    ws.freeze_panes = 'A2'
    
    # Set proper default widths for each column so the template is ready to use
    widths = {
        'A': 35, 'B': 25, 'C': 25, 'D': 40, 'E': 15, 'F': 15, 
        'G': 15, 'H': 60, 'I': 20, 'J': 30, 'K': 30
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save(file_path)
    print(f"Successfully created a styled empty template: {file_path}")

except Exception as e:
    print(f"Error: {e}")
