import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

FILE_1488 = r"E:\Internship\PocketFM\kensington_authors_MASTER_1_to_1488.xlsx"
FILE_SCRAPED = r"E:\Internship\PocketFM\Kensington_Media_Catalog.xlsx"
OUTPUT_FILE = r"E:\Internship\PocketFM\kensington_authors_MASTER_1_to_1488.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching kensington style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def merge_catalogs():
    print(f">>> Reading Massive 1488 File: {FILE_1488}")
    df_1488 = pd.read_excel(FILE_1488)
    print(f"    Loaded {len(df_1488)} rows.")

    print(f">>> Reading Scraped File: {FILE_SCRAPED}")
    df_scraped = pd.read_excel(FILE_SCRAPED)
    print(f"    Loaded {len(df_scraped)} rows.")

    # Standardize column names to be absolutely sure they match the template
    # We will map whatever columns they have to the HEADERS if they loosely match, 
    # or just force them. The easiest is if they already match.
    
    # We enforce HEADERS exactly.
    df_1488.columns = [c.strip() for c in df_1488.columns]
    df_scraped.columns = [c.strip() for c in df_scraped.columns]

    print(">>> Merging datasets...")
    # Concatenate vertically: 1488 file on top, scraped file on bottom
    merged_df = pd.concat([df_1488, df_scraped], ignore_index=True)
    
    # Fill NaNs with "N/A" for consistency
    merged_df.fillna("N/A", inplace=True)
    
    # Ensure all 11 columns exist in the output in exact order
    for h in HEADERS:
        if h not in merged_df.columns:
            merged_df[h] = "N/A"
            
    merged_df = merged_df[HEADERS]
    
    print(f">>> Total Rows after merge: {len(merged_df)}")
    
    print(f">>> Saving raw merged data to {OUTPUT_FILE}...")
    merged_df.to_excel(OUTPUT_FILE, index=False)
    
    print(">>> Applying premium 11-column formatting to the massive sheet...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    format_excel_sheet(ws)
    
    wb.save(OUTPUT_FILE)
    print(">>> Formatting complete!")
    
    if os.name == 'nt':
        os.startfile(OUTPUT_FILE)

if __name__ == "__main__":
    merge_catalogs()
