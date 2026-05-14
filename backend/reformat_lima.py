import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH = r"E:\Internship\PocketFM\Lima Agency.xlsx"

def format_lima_excel(file_path):
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return

    print(f"Formatting {file_path}...")
    
    # Reload with openpyxl to apply styles
    wb = load_workbook(file_path)
    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
        ws.title = 'Lima Agency Catalog'
    else:
        ws = wb.active
        ws.title = 'Lima Agency Catalog'

    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid") # Lima Purple
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
    # Column Widths
    column_widths = [40, 25, 15, 45, 15, 15, 15, 65, 15, 25, 15]
    for i, width in enumerate(column_widths):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = width
        
    # Body Formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            # Align long text columns to left, others to center
            if cell.column in [1, 2, 4, 8, 10]: # Series, Author, Link, Synopsis, Sub-genre
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            cell.border = border

    ws.freeze_panes = "A2"
    wb.save(file_path)
    print("Formatting Complete!")

if __name__ == "__main__":
    format_lima_excel(FILE_PATH)
