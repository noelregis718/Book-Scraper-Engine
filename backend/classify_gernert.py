import openpyxl
import os
import re

CATALOG_FILE = r"E:\Internship\PocketFM\Gernert_Media_Catalog.xlsx"

# The 12 Romantasy Categories & their respective keywords
TAXONOMY = {
    "Werewolf / Shifter Romance": [
        "werewolf", "shifter", "wolf shifter", "bear shifter", "dragon shifter", 
        "pack alpha", "omegaverse", "true mate", "shifter romance", "alpha", 
        "pack", "omega", "luna", "werewolves", "shifters", "fated mate", "mate bond"
    ],
    "Monster Romance (Non-Shifter)": [
        "monster romance", "tentacle", "orc", "kraken", "minotaur", "beastman", 
        "non-human lover", "monster", "demons", "alien", "beast", "demon", "monster lover"
    ],
    "Mythology, Legend & Fairy Tale Retelling": [
        "greek myth", "norse myth", "retelling", "fairy tale retelling", 
        "gods and monsters", "mortal and god", "mythology", "greek gods", 
        "hades", "persephone", "fairy tale", "legend", "myths", "folklore"
    ],
    "War College / Military Academy": [
        "war college", "military academy", "dragon rider", "beast bond", 
        "combat training", "lethal training", "training", "fourth wing", 
        "basgiath", "war academy", "military school"
    ],
    "High-Stakes Games & Deadly Trials": [
        "deadly trials", "magical tournament", "deadly games", "magical contest", 
        "trial of the", "trials", "competition", "tournament", "survival", 
        "hunger games style", "arena", "lethal trials", "deadly competition"
    ],
    "Dark Academia Romantasy": [
        "magical academy", "secret society", "magic school", "university of magic", 
        "scholar of the", "campus of magic", "university", "secret societies", 
        "library", "dark academia", "campus", "scholarly", "boarding school"
    ],
    "Gothic Dark Romantasy": [
        "gothic romance", "dark magic", "haunted castle", "vampiric", "macabre", 
        "blood magic", "deathly curse", "gothic mystery", "gothic", "haunted", 
        "manor", "dark romance", "gloomy", "castle", "shadows"
    ],
    "Korean Romance Fantasy / Isekai": [
        "isekai", "reincarnation", "villainess", "empress", "saintess", 
        "transmigration", "isekai romance", "reincarnated", "otome", "korean", 
        "manhwa style", "webtoon style"
    ],
    "Cozy / Cottagecore": [
        "cozy fantasy", "cottagecore", "magical bakery", "small town magic", 
        "low-stakes fantasy", "cozy", "small town", "low stakes", "wholesome", 
        "bakery", "tea shop", "cottage"
    ],
    "Paranormal Romance": [
        "vampire romance", "demon lover", "succubus", "paranormal mystery", 
        "coven of witches", "warlock", "vampire", "ghost", "witch", 
        "paranormal", "psychic", "medium", "vampires", "witches", "ghosts", 
        "coven", "necromancer"
    ],
    "High Fantasy Court Adventure": [
        "fae", "elven", "magical kingdom", "royal court", "fantasy empire", 
        "epic quest", "magic world", "crown prince", "queen of", "king of", 
        "throne of", "heir to the", "castle of", "court", "throne", 
        "kingdom", "royalty", "epic fantasy", "high fantasy", "princess", 
        "prince", "emperor", "duchy", "lord of"
    ],
    "Urban / Contemporary Fantasy Romance": [
        "urban fantasy", "hidden magic", "magic in the city", "secret supernatural", 
        "hidden world", "modern day", "city", "contemporary fantasy"
    ]
}

def identify_subgenre(synopsis, series_name):
    if not isinstance(synopsis, str):
        synopsis = ""
    if not isinstance(series_name, str):
        series_name = ""
        
    text = f"{series_name} {synopsis}".lower()
    
    # Check each category sequentially
    for genre, keywords in TAXONOMY.items():
        for kw in keywords:
            # We look for substring matches for flexible and robust detection
            if kw.lower() in text:
                return genre
                
    return "N/A"

def run_classification():
    if not os.path.exists(CATALOG_FILE):
        print(f"Error: Catalog file not found at {CATALOG_FILE}")
        return

    print(">>> Loading workbook to classify Romantasy columns...")
    wb = openpyxl.load_workbook(CATALOG_FILE, data_only=False)
    ws = wb.active

    # Find column indexes
    # A (1) = Series Name
    # H (8) = Synopsis
    # I (9) = Romantasy Yes/No
    # J (10) = Romantasy Sub-genre
    
    total_rows = ws.max_row
    print(f"Total rows in sheet: {total_rows}")

    classified_count = 0
    for row_idx in range(2, total_rows + 1):
        series_name = ws.cell(row=row_idx, column=1).value
        synopsis = ws.cell(row=row_idx, column=8).value
        
        # Resolve values
        series_name_str = str(series_name) if series_name is not None else ""
        synopsis_str = str(synopsis) if synopsis is not None else ""
        
        subgenre = identify_subgenre(synopsis_str, series_name_str)
        
        if subgenre != "N/A":
            ws.cell(row=row_idx, column=9, value="Yes")
            ws.cell(row=row_idx, column=10, value=subgenre)
            classified_count += 1
            print(f"  [Row {row_idx}] '{series_name_str[:30]}' Classified as: {subgenre}")
        else:
            ws.cell(row=row_idx, column=9, value="No")
            ws.cell(row=row_idx, column=10, value="N/A")

    print(f"\nSaving updated sheet to {CATALOG_FILE}...")
    wb.save(CATALOG_FILE)
    print(f"Success! Classified {classified_count} / {total_rows - 1} series as Romantasy!")

if __name__ == "__main__":
    run_classification()
