import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = r'e:\Internship\PocketFM\Books_Scraping_Template.xlsx'
print("Loading workbook...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 1. Replicate "The Unter Agency" Header Styling
header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 2. Replicate "The Unter Agency" Data Styling
data_font = Font(name='Calibri', size=11, color="000000")
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
hyperlink_font = Font(name='Calibri', size=11, color="0563C1", underline="single")

thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                     right=Side(style='thin', color='BFBFBF'), 
                     top=Side(style='thin', color='BFBFBF'), 
                     bottom=Side(style='thin', color='BFBFBF'))

print("Formatting header...")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.row_dimensions[1].height = 40  # Give header some breathing room

print("Formatting data rows...")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Hyperlinks for column D (GoodReads series link)
        if cell.column == 4 and cell.value and str(cell.value).startswith('http'):
            cell.font = hyperlink_font
            cell.hyperlink = cell.value
        else:
            cell.font = data_font

# Try to clear specific row heights so Excel auto-fits the wrapped text
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = None

print("Applying column widths...")
column_widths = {
    'A': 40, # Name of Series
    'B': 25, # Author Name
    'C': 20, # Publisher
    'D': 55, # GoodReads series link
    'E': 15, # Number of PRIMARY books in the series
    'F': 15, # Rating (out of 5) of Primary Book 1
    'G': 15, # Ratings (#) of Primary Book 1
    'H': 65, # Synopsis (if available)
    'I': 20, # Romantasy = Yes or No?
    'J': 35, # Romantasy Sub-Genre of series
    'K': 20  # Name of agent
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

ws.freeze_panes = 'A2'

print("Saving workbook...")
wb.save(file_path)
print("Done!")
