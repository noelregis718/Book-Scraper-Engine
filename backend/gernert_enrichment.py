import asyncio
import os
import re
import sys
import pandas as pd
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure backend directory is in path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

# --- CONFIGURATION ---
CATALOG_FILE = r"E:\Internship\PocketFM\Gernert_Media_Catalog.xlsx"
START_ROW = 512    # Next data row in Excel
END_ROW = 551     # Next 40 rows in Excel (inclusive)
HEADLESS = False  # Set to False so the user can watch the browser run

def format_excel_sheet(ws):
    """Applies professional styling to the worksheet in-place."""
    # Header: Navy Blue Fill, White Bold Text
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Border definition
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"),
        right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"),
        bottom=Side(style='thin', color="BFBFBF")
    )
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Style Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style Body Rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10, color="000000")
            
            # Alignments per column type
            if cell.column in [1, 2, 4, 8]:  # Name of Series, Author Name, GoodReads link, Synopsis
                if cell.column == 8: # Synopsis wraps
                    cell.alignment = wrap_align
                else:
                    cell.alignment = left_align
            else:
                cell.alignment = center_align
                
    # Column width specifications
    column_widths = {
        1: 35,  # Name of Series
        2: 25,  # Author Name
        3: 15,  # Publisher
        4: 40,  # GoodReads series link
        5: 15,  # Number of PRIMARY books in the series
        6: 15,  # Rating (out of 5) of Primary Book 1
        7: 15,  # Ratings (#) of Primary Book 1
        8: 60,  # Synopsis (if available)
        9: 22,  # Romantasy = Yes or No?
        10: 30, # Romantasy Sub-Genre of series
        11: 20  # Name of agent
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Freeze the top header row
    ws.freeze_panes = "A2"
    
    # Enable sorting/filtering
    ws.auto_filter.ref = ws.dimensions

async def run_enrichment():
    print("=" * 60)
    print("      GERNERT COMPANY DEEP ENRICHMENT PIPELINE")
    print("=" * 60)
    print(f"Targeting Rows: {START_ROW} to {END_ROW} inclusive.")
    print(f"Spreadsheet: {CATALOG_FILE}\n")

    if not os.path.exists(CATALOG_FILE):
        print(f"[Error] Catalog spreadsheet {CATALOG_FILE} not found!")
        return

    # Initialize Goodreads Scraper
    gr_scraper = GoodreadsScraper(headless=HEADLESS)
    
    async with async_playwright() as p:
        print("[System] Launching Playwright browser...")
        browser = await p.chromium.launch(headless=HEADLESS)
        # Create a persistent profile or context to prevent immediate blocking
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        # Load workbook
        wb = load_workbook(CATALOG_FILE)
        ws = wb.active
        
        success_count = 0
        failure_count = 0
        
        for r in range(START_ROW, END_ROW + 1):
            print(f"\n--- [Row {r} / {END_ROW}] ---")
            
            # Read Book Title & Author Name from sheet (1-based indices)
            book_title = ws.cell(row=r, column=1).value
            author_name = ws.cell(row=r, column=2).value
            
            if not book_title or str(book_title).strip() == "":
                print("  [Warning] Empty book title. Skipping row.")
                continue
                
            book_title = str(book_title).strip()
            author_name = str(author_name).strip() if author_name else "Unknown"
            
            print(f"  [Scraping] Title: '{book_title}' | Author: '{author_name}'")
            
            try:
                # Query Goodreads search & extract details
                gr_data = await gr_scraper.scrape_goodreads_data(context, book_title, author_name)
                
                if gr_data:
                    print("  [Success] Retrieved data from Goodreads!")
                    
                    # Map to the 11 columns
                    # Col 3: Publisher is already populated as "The Gernert Company"
                    # Col 11: Name of Agent is already populated as "David Gernert"
                    
                    # Col 4: GoodReads series link
                    series_url = gr_data.get("GoodReads_Series_URL", "N/A")
                    if series_url == "N/A" or not series_url:
                        series_url = gr_data.get("GoodReads_Book_URL", "N/A")
                    ws.cell(row=r, column=4, value=series_url)
                    
                    # Col 5: Number of PRIMARY books in the series
                    ws.cell(row=r, column=5, value=gr_data.get("Num_Primary_Books", "1"))
                    
                    # Col 6: Rating (out of 5) of Primary Book 1
                    ws.cell(row=r, column=6, value=gr_data.get("Book1_Rating", "N/A"))
                    
                    # Col 7: Ratings (#) of Primary Book 1
                    ws.cell(row=r, column=7, value=gr_data.get("Book1_Num_Ratings", "N/A"))
                    
                    # Col 8: Synopsis (if available)
                    ws.cell(row=r, column=8, value=gr_data.get("Description", "N/A"))
                    
                    # Col 9: Romantasy = Yes or No?
                    ws.cell(row=r, column=9, value=gr_data.get("Romantasy_Subgenre", "No"))
                    
                    # Col 10: Romantasy Sub-Genre of series
                    ws.cell(row=r, column=10, value=gr_data.get("Sub_Genre", "N/A"))
                    
                    success_count += 1
                else:
                    print("  [Failure] Goodreads search returned no details.")
                    # Set defaults or leave as N/A
                    ws.cell(row=r, column=4, value="N/A")
                    ws.cell(row=r, column=5, value="1")
                    ws.cell(row=r, column=6, value="N/A")
                    ws.cell(row=r, column=7, value="N/A")
                    ws.cell(row=r, column=8, value="N/A")
                    ws.cell(row=r, column=9, value="No")
                    ws.cell(row=r, column=10, value="N/A")
                    failure_count += 1
                    
            except Exception as e:
                print(f"  [Error] Scraping failed for row {r}: {e}")
                # Set fallbacks so the row is not completely empty
                ws.cell(row=r, column=4, value="N/A")
                ws.cell(row=r, column=5, value="1")
                ws.cell(row=r, column=6, value="N/A")
                ws.cell(row=r, column=7, value="N/A")
                ws.cell(row=r, column=8, value="N/A")
                ws.cell(row=r, column=9, value="No")
                ws.cell(row=r, column=10, value="N/A")
                failure_count += 1
                
            # Apply styling and save workbook incrementally after each row
            print("  [System] Saving progress and applying styling...")
            try:
                format_excel_sheet(ws)
                wb.save(CATALOG_FILE)
                print("  [System] Progress saved successfully!")
            except Exception as save_err:
                print(f"  [Error] Could not save progress: {save_err}. Is the file open in Excel?")
                
            # Pause between requests to prevent Goodreads rate-limiting
            await asyncio.sleep(3)
            
        await browser.close()
        print("\n" + "=" * 60)
        print("                 ENRICHMENT BATCH COMPLETE")
        print("=" * 60)
        print(f"Total Processed Rows: {END_ROW - START_ROW + 1}")
        print(f"Successful Hits:     {success_count}")
        print(f"Failed/Skipped:      {failure_count}")
        print(f"Output Catalog:      {CATALOG_FILE}")
        print("=" * 60 + "\n")
        
        # Auto-open on Windows
        if os.name == 'nt':
            print("  [System] Auto-opening sheet for review...")
            os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    asyncio.run(run_enrichment())
