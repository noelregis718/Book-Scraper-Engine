import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_sleek_styling(file_path):
    print(f"Loading {file_path} for styling...")
    wb = load_workbook(file_path)
    
    # Premium Colors (Slate theme)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
    header_font = Font(color="F8FAFC", bold=True, name="Calibri", size=11) # Slate 50
    zebra_light = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), 
        right=Side(style='thin', color='CBD5E1'), 
        top=Side(style='thin', color='CBD5E1'), 
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for sheet_name in wb.sheetnames:
        print(f"Styling sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Freeze the top row so scrolling is easy
        ws.freeze_panes = "A2"
        
        # Apply Header Styles
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Column Widths logic
        for col_idx in range(1, ws.max_column + 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            header_val = str(ws.cell(row=1, column=col_idx).value).lower()
            
            # Make summary/description columns very wide so text wraps nicely
            if "summary" in header_val or "description" in header_val or "synopsis" in header_val:
                ws.column_dimensions[column_letter].width = 65
            elif "url" in header_val or "link" in header_val:
                ws.column_dimensions[column_letter].width = 35
            else:
                ws.column_dimensions[column_letter].width = 25
                
        # Apply uniform white background, borders, and wrap text (Auto-fits row heights!)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = zebra_light
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    print("Saving the master workbook...")
    wb.save(file_path)
    print("Sleek premium styling successfully applied to ALL sheets in the workbook!")

if __name__ == "__main__":
    apply_sleek_styling("e:/Internship/PocketFM/PocketFM_CT_Analysis_Master.xlsx")
