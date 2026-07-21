import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_sleek_styling(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Premium Colors
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
    header_font = Font(color="F8FAFC", bold=True, name="Calibri", size=11) # Slate 50
    
    zebra_dark = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate 100
    zebra_light = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), 
        right=Side(style='thin', color='CBD5E1'), 
        top=Side(style='thin', color='CBD5E1'), 
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Freeze the top row so scrolling is easy
    ws.freeze_panes = "A2"
    
    # Apply Header Styles
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Fixed Column Widths
    for col_idx in range(1, ws.max_column + 1):
        column = ws.cell(row=1, column=col_idx).column_letter
        if ws.cell(row=1, column=col_idx).value == "Podium Summary / Description":
            ws.column_dimensions[column].width = 65
        else:
            ws.column_dimensions[column].width = 25
            
    # Apply uniform white background & cell formatting (Excel will auto-fit row height)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = zebra_light  # All rows pure white
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(file_path)
    print("Sleek premium styling successfully applied!")

if __name__ == "__main__":
    apply_sleek_styling("e:/Internship/PocketFM/podium_data.xlsx")
