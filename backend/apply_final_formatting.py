import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
FILE_NAME = "Horror_Amazon_Keyword_Crawl_Goodreads_Links_Checked.xlsx"

def apply_clean_formatting():
    if not os.path.exists(FILE_NAME):
        print(f"File {FILE_NAME} not found.")
        return

    print(f"Applying 'Clean Cutoff' sizing and formatting to {FILE_NAME}...")
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # 1. Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_top_aligned = Alignment(horizontal="left", vertical="top", wrap_text=True)
    standard_aligned = Alignment(horizontal="left", vertical="center")

    # 2. Iterate columns for sizing and alignment
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        header_name = col[0].value

        for cell in col:
            # Apply border to every cell
            cell.border = thin_border
            
            # Content-based width detection
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

            # Alignment logic
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_aligned
            else:
                # Wrap text for long content
                if header_name in ['URL', 'Synopsis/Summary', 'Rationale', 'Goodread Link', 'Series Link', 'Amazon URL']:
                    cell.alignment = left_top_aligned
                else:
                    cell.alignment = standard_aligned

        # Set column width (Fixed for links/summaries, Auto for others)
        if header_name in ['URL', 'Amazon URL', 'Goodread Link', 'Series Link']:
            adjusted_width = 25
        elif header_name == 'Synopsis/Summary':
            adjusted_width = 40
        else:
            adjusted_width = min(max_length + 2, 30)
            
        ws.column_dimensions[col_letter].width = adjusted_width

    # 3. Row heights (Header=25, Data=18)
    ws.row_dimensions[1].height = 25
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # 4. Freeze top row
    ws.freeze_panes = "A2"

    wb.save(FILE_NAME)
    print("Formatting and sizing applied successfully!")
    
    if os.name == 'nt':
        os.startfile(FILE_NAME)

if __name__ == "__main__":
    apply_clean_formatting()
