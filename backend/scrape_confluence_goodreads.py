import asyncio
import os
import re
import json
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from playwright.async_api import async_playwright

EXCEL_FILE = '../New_Agency_Template.xlsx'
PROFILE_DIR = '../playwright_goodreads_profile'

async def scrape_author(context, author_name, sem, results_dict):
    async with sem:
        page = await context.new_page()
        print(f"[{author_name}] Starting scrape...")
        books_data = []
        try:
            # 1. Search the author
            search_url = f"https://www.goodreads.com/search?q={author_name.replace(' ', '+')}"
            await page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(2)
            
            # Find the author profile link
            author_link = await page.query_selector('a[href*="/author/show/"]')
            if not author_link:
                author_link = await page.query_selector('.authorName, .authorName__container a')

            if not author_link:
                print(f"[{author_name}] Could not find author profile.")
                results_dict[author_name] = []
                return
                
            author_url = await author_link.evaluate("el => el.href")
            await page.goto(author_url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(2)
            
            # 2. On author page, books are sorted by popularity by default
            # Grab the top 2 books
            rows = await page.query_selector_all('tr[itemtype="http://schema.org/Book"]')
            
            book_links = []
            for row in rows[:2]:
                title_el = await row.query_selector('a.bookTitle')
                if title_el:
                    link = await title_el.evaluate("el => el.href")
                    if link not in book_links:
                        book_links.append(link)
                        
            if not book_links:
                print(f"[{author_name}] Found no books on profile.")
                results_dict[author_name] = []
                return
                
            print(f"[{author_name}] Found {len(book_links)} top books. Scraping details...")
            
            # 3. For each book, visit the page and extract details
            for idx, link in enumerate(book_links):
                try:
                    await page.goto(link, wait_until="domcontentloaded", timeout=60000)
                    await asyncio.sleep(1.5)
                    
                    title = "N/A"
                    title_el = await page.query_selector('[data-testid="bookTitle"], #bookTitle')
                    if title_el:
                        title = (await title_el.inner_text()).strip()
                        
                    rating = "N/A"
                    ratings_count = "N/A"
                    rating_el = await page.query_selector('.RatingStatistics__rating')
                    if rating_el:
                        rating = (await rating_el.inner_text()).strip()
                        
                    count_el = await page.query_selector('[data-testid="ratingsCount"]')
                    if count_el:
                        count_text = (await count_el.inner_text()).strip()
                        ratings_count = re.sub(r'[^\d]', '', count_text)
                        
                    if rating == "N/A":
                        ld_el = await page.query_selector('script[type="application/ld+json"]')
                        if ld_el:
                            try:
                                ld_data = json.loads(await ld_el.inner_text())
                                if isinstance(ld_data, list): ld_data = ld_data[0]
                                rating = str(ld_data.get('aggregateRating', {}).get('ratingValue', 'N/A'))
                                ratings_count = str(ld_data.get('aggregateRating', {}).get('ratingCount', 'N/A'))
                            except: pass

                    synopsis = "N/A"
                    desc_el = await page.query_selector('[data-testid="description"] .Formatted, .readable')
                    if desc_el:
                        synopsis = (await desc_el.inner_text()).strip()
                        synopsis = re.sub(r'\s+', ' ', synopsis)
                        
                    series_link = link
                    series_el = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a')
                    if series_el:
                        series_link = await series_el.evaluate("el => el.href")
                        
                    books_data.append({
                        'title': title,
                        'link': link,
                        'rating': rating,
                        'ratings_count': ratings_count,
                        'synopsis': synopsis,
                        'series_link': series_link
                    })
                    print(f"[{author_name}] Extracted Book {idx+1}: {title} ({rating} stars, {ratings_count} ratings)")
                except Exception as e:
                    print(f"[{author_name}] Error extracting book {idx+1}: {e}")
                    
            results_dict[author_name] = books_data
            
        except Exception as e:
            print(f"[{author_name}] General Error: {e}")
            results_dict[author_name] = []
        finally:
            await page.close()


async def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    author_col = 2
    
    # We use dict from keys to deduplicate while preserving order
    authors = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=author_col).value
        if name and str(name).strip() and str(name).strip() != "Author Name":
            name = str(name).strip()
            if name not in authors:
                authors.append(name)
            
    print(f"Found {len(authors)} unique authors to scrape.")
    
    results_dict = {}
    sem = asyncio.Semaphore(8) # 8 tabs concurrently
    
    async with async_playwright() as p:
        print("Launching browser with persistent profile...")
        context = await p.chromium.launch_persistent_context(
            user_data_dir=os.path.abspath(PROFILE_DIR),
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            args=['--disable-blink-features=AutomationControlled']
        )
        
        page = await context.new_page()
        await page.goto("https://www.goodreads.com/")
        print("Waiting 5 seconds for user to verify no captchas on home page...")
        await asyncio.sleep(5)
        await page.close()
        
        print(f"Starting concurrent scrape for {len(authors)} authors (8 tabs max)...")
        tasks = []
        for author in authors:
            tasks.append(scrape_author(context, author, sem, results_dict))
            
        await asyncio.gather(*tasks)
        await context.close()
        
    print("\nScraping complete! Preparing to update Excel sheet...")
    
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Agency Data"
    
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin', color='D3D3D3'), 
                         right=Side(style='thin', color='D3D3D3'), 
                         top=Side(style='thin', color='D3D3D3'), 
                         bottom=Side(style='thin', color='D3D3D3'))
                         
    headers = [
        "Name of Series", "Author Name", "Publisher", "GoodReads series link",
        "Number of PRIMARY books in the series", "Rating (out of 5) of Primary Book 1",
        "Ratings (#) of Primary Book 1", "Synopsis (if available)",
        "Romantasy = Yes or No?", "Romantasy Sub-Genre of series",
        "Name of agent in the main folder"
    ]
    
    ws_new.append(headers)
    for cell in ws_new[1]:
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for author in authors:
        books = results_dict.get(author, [])
        if not books:
            row_data = [None] * 11
            row_data[1] = author
            row_data[10] = "Confluence Literary Agency"
            ws_new.append(row_data)
            for cell in ws_new[ws_new.max_row]:
                cell.alignment = cell_alignment
                cell.border = thin_border
        else:
            for book in books:
                row_data = [
                    book.get('title', 'N/A'),
                    author,
                    None,
                    book.get('series_link', book.get('link', 'N/A')),
                    None,
                    book.get('rating', 'N/A'),
                    book.get('ratings_count', 'N/A'),
                    book.get('synopsis', 'N/A'),
                    None,
                    None,
                    "Confluence Literary Agency"
                ]
                ws_new.append(row_data)
                for cell in ws_new[ws_new.max_row]:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    
    column_widths = {
        'A': 25, 'B': 20, 'C': 15, 'D': 35, 'E': 15,
        'F': 12, 'G': 12, 'H': 50, 'I': 15, 'J': 15, 'K': 25
    }
    for col, width in column_widths.items():
        ws_new.column_dimensions[col].width = width
        
    ws_new.freeze_panes = 'A2'
    
    wb_new.save(EXCEL_FILE)
    print(f"Successfully saved {ws_new.max_row - 1} records to {EXCEL_FILE}")

if __name__ == "__main__":
    asyncio.run(main())
