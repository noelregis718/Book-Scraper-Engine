"""
PHASE 1 of the Mushens full-roster pipeline.

Scrapes every author from https://www.mushens-entertainment.com/all-authors
and writes them to a new "All Authors" sheet inside the existing workbook
Mushens_Entertainment_Bestsellers.xlsx (leaves the main 11-column sheet untouched).

Columns in the new sheet:
  #  |  Author Name  |  Section Letter  |  Goodreads URL  |  Status

Phase 2 (separate script) will read this sheet and pull each author's first 3
books from Goodreads.
"""

import asyncio
import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from playwright.async_api import async_playwright


URL  = "https://www.mushens-entertainment.com/all-authors"
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "Mushens_Entertainment_Bestsellers.xlsx")
SHEET_NAME = "All Authors"


# ---------------------------------------------------------------- parsing
def parse_authors_from_body(body_text):
    """
    The page is plain-text alphabetic. We walk the lines, skipping the
    nav/footer, treating single capital letters as section headers, and
    treating everything else as an author name.
    """
    lines = [ln.strip() for ln in body_text.splitlines()]
    # find the second "Authors" — that's the H1 of the listing, right before "A"
    listing_start = None
    seen_authors_heading = 0
    for i, ln in enumerate(lines):
        if ln == "Authors":
            seen_authors_heading += 1
            if seen_authors_heading == 2:
                listing_start = i + 1
                break
    if listing_start is None:
        # fallback: start after first standalone "A"
        for i, ln in enumerate(lines):
            if ln == "A":
                listing_start = i
                break

    nav_words   = {"About Us", "Authors", "Submissions", "Rights", "News", "Blog",
                   "Skip to Content", ""}
    footer_pat  = re.compile(r"to enquire about sending proofs|mushens entertainment\b|"
                             r"made with squarespace|privacy policy|contact us at info",
                             re.IGNORECASE)

    authors = []
    section = ""
    for ln in lines[listing_start:]:
        if footer_pat.search(ln):
            break
        if not ln or ln in nav_words:
            continue
        # Single uppercase letter = section header
        if re.fullmatch(r"[A-Z]", ln):
            section = ln
            continue
        # Strip stray non-text chars and replace common mojibake apostrophe
        name = ln.replace("’", "'").replace("�", "'").strip()
        # Skip junk lines (headings already filtered)
        if len(name) < 2:
            continue
        # Some lines might still be nav-y if structure shifts
        if name.lower().startswith(("about", "submissions", "rights", "news", "blog")):
            continue
        authors.append({"section": section, "name": name})
    return authors


# ---------------------------------------------------------------- scrape
async def scrape_author_list():
    print(f"Fetching {URL} ...")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context()
        page = await ctx.new_page()
        await page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        await page.goto(URL, wait_until="networkidle", timeout=60000)
        await asyncio.sleep(2)
        body = await page.inner_text("body")
        await browser.close()

    authors = parse_authors_from_body(body)
    # de-dupe in case any name appears twice
    seen = set()
    uniq = []
    for a in authors:
        key = a["name"].lower()
        if key in seen:
            continue
        seen.add(key)
        uniq.append(a)
    return uniq


# ---------------------------------------------------------------- write
def write_authors_to_workbook(authors):
    if not os.path.exists(XLSX):
        raise FileNotFoundError(f"Workbook not found: {XLSX}")

    wb = openpyxl.load_workbook(XLSX)

    # Replace the sheet if it already exists, so re-runs are idempotent
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    headers = ["#", "Author Name", "Section", "Goodreads URL", "Status"]
    ws.append(headers)
    for i, a in enumerate(authors, start=1):
        ws.append([i, a["name"], a["section"], "", "pending"])

    # Styling — match the main sheet's dark blue/white style
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border

    for r in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.fill = white_fill
            c.alignment = align_left
            c.border = border

    widths = {"A": 6, "B": 32, "C": 10, "D": 55, "E": 12}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(XLSX)
    print(f"Wrote {len(authors)} authors to sheet '{SHEET_NAME}' in {XLSX}")


# ---------------------------------------------------------------- main
def main():
    authors = asyncio.run(scrape_author_list())
    print(f"Parsed {len(authors)} unique authors.")
    # Quick sanity print
    for a in authors[:8]:
        print(f"  [{a['section']}] {a['name']}")
    print("  ...")
    for a in authors[-5:]:
        print(f"  [{a['section']}] {a['name']}")
    write_authors_to_workbook(authors)


if __name__ == "__main__":
    main()
