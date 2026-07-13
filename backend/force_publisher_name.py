import openpyxl

excel_file = '../New_Agency_Template.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    publisher_col = 3 # 'Publisher' is the 3rd column (C)
    
    count = 0
    # Start at row 2 to skip the header
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=publisher_col).value = "Confluence Literary Agency"
        count += 1
        
    wb.save(excel_file)
    print(f"Successfully populated 'Confluence Literary Agency' across all {count} data rows in Column C!")

except Exception as e:
    print(f"Error: {e}")
