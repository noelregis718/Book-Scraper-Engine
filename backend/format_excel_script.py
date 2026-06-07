import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

file_path = 'books_from_uploaded_images.xlsx'
df = pd.read_excel(file_path)

if "Book Name" in df.columns:
    if "Name of Series" in df.columns:
        df = df.drop(columns=["Name of Series"])
    df = df.rename(columns={"Book Name": "Name of Series"})

# Save back
df.to_excel(file_path, index=False)

# Re-apply styles
wb = load_workbook(file_path)
ws = wb.active

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_num, column_cells in enumerate(ws.columns, 1):
    header_cell = column_cells[0]
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = header_alignment
    
    max_length = 0
    col_letter = get_column_letter(col_num)
    for cell in column_cells:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    adjusted_width = max(adjusted_width, 15)
    ws.column_dimensions[col_letter].width = adjusted_width

ws.freeze_panes = 'A2'
wb.save(file_path)
print("Updated Book Name to Name of Series.")
