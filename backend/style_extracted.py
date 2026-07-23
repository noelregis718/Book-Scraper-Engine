import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = r"e:\Internship\PocketFM\Romantasy_v2_extracted.xlsx"
print("Loading workbook for styling...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define Styles
header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='BFBFBF'), 
    right=Side(style='thin', color='BFBFBF'), 
    top=Side(style='thin', color='BFBFBF'), 
    bottom=Side(style='thin', color='BFBFBF')
)

print("Applying header styles...")
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = thin_border

print("Applying body styles and formatting hyperlinks...")
# Apply styles to data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment_left_top
        cell.border = thin_border
        
        # Turn URLs into clickable hyperlinks
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
            cell.font = Font(color="0563C1", underline="single", name='Calibri', size=11)
            cell.hyperlink = cell.value

# Set column widths
col_widths = {
    'A': 8,   # S.No
    'B': 30,  # Title
    'C': 25,  # Author
    'D': 35,  # GR Book 1
    'E': 20,  # Agency
    'F': 35,  # GR Series
    'G': 12,  # Books in series
    'H': 12,  # Page count
    'R': 12,  # GR ratings
    'T': 12,  # MG
}

for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Set a default width for the rest
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    if col_letter not in col_widths:
        ws.column_dimensions[col_letter].width = 20

# Freeze panes below the header
ws.freeze_panes = 'A2'

print("Saving workbook...")
wb.save(file_path)
print("Workbook completely styled!")
