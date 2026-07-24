import asyncio
import re
import os
import openpyxl
import json
from playwright.async_api import async_playwright
from goodreads_scraper import GoodreadsScraper

INPUT_FILE = r"e:\Internship\PocketFM\All-Genre Licensing Tracker.xlsx"
STATE_FILE = r"e:\Internship\PocketFM\backend\scraper_state.json"

class BlockedError(Exception):
    pass

async def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

async def save_state(state, lock):
    async with lock:
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=4)

async def solve_captcha_if_present(page, url=""):
    try:
        page_text = await page.evaluate("document.body ? document.body.innerText.toLowerCase() : ''")
        if "unexpected error" in page_text or "403 forbidden" in page_text or ("captcha" in page_text and not await page.query_selector('.bookTitle')):
            if await page.query_selector('#captcha-image, .captcha, iframe[src*="captcha"]'):
                print(f"    [!!!] CAPTCHA detected! Please solve it manually.")
                try:
                    await page.wait_for_selector('a.bookTitle, [data-testid="pagesFormat"], .listWithDividers__item, h1', timeout=120000)
                    print(f"    [Success] CAPTCHA solved.")
                    return
                except:
                    print(f"    [Timeout] CAPTCHA wait timeout. Blocked.")
                    raise BlockedError("CAPTCHA timeout")
            else:
                # It's an unexpected error page or silent block
                print(f"    [!!!] Unexpected Goodreads error page detected. Blocked.")
                raise BlockedError("Unexpected error page")
                
    except BlockedError:
        raise
    except Exception:
        pass

async def search_book_on_goodreads(page, title, author):
    search_term = f"{title} {author}".strip()
    if not search_term:
        return None
        
    search_url = f"https://www.goodreads.com/search?q={search_term.replace(' ', '+')}"
    print(f"    Searching Goodreads: {search_url}")
    try:
        await page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(2)
        await solve_captcha_if_present(page, search_url)
        
        first_result = await page.query_selector('a.bookTitle')
        if first_result:
            b_url = await first_result.evaluate("el => el.href")
            if b_url.startswith('/'):
                b_url = f"https://www.goodreads.com{b_url}"
            return b_url
    except BlockedError:
        raise
    except Exception as e:
        if "TargetClosedError" in str(type(e)) or "Timeout" in str(e):
            raise BlockedError(f"Blocked during search: {e}")
        print(f"    Error searching for {search_term}: {e}")
    return None

async def get_book_details(page, url):
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(1.5)
        await solve_captcha_if_present(page, url)
        
        content = await page.content()
        pages = 0
        match = re.search(r'"numPages":(\d+)', content)
        if match:
            pages = int(match.group(1))
        else:
            match = re.search(r'>(\d+)\s*pages', content, re.IGNORECASE)
            if match:
                pages = int(match.group(1))
                
        agency = ""
        pub_info = await page.query_selector('[data-testid="publicationInfo"]')
        if pub_info:
            pub_text = await pub_info.inner_text()
            if " by " in pub_text:
                agency = pub_text.split(" by ")[-1].strip()
                
        # Also try to find series link if we are on a book page
        series_link = None
        b_series = await page.query_selector('h3.Text__title3 a[href*="/series/"], [data-testid="series"] a, h2 a[href*="/series/"]')
        if b_series:
            series_link = await b_series.evaluate("el => el.href")
            if series_link.startswith('/'):
                series_link = "https://www.goodreads.com" + series_link
                
        return pages, agency, series_link
    except BlockedError:
        raise
    except Exception as e:
        if "TargetClosedError" in str(type(e)) or "Timeout" in str(e):
            raise BlockedError(f"Blocked during book details: {e}")
        print(f"    Error scraping book page {url}: {e}")
        return 0, "", None

async def get_series_pages(page, url, row_idx, global_state, state_lock, global_stop_event):
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(2)
        await solve_captcha_if_present(page, url)
        
        book_links = await page.query_selector_all('.listWithDividers__item a.bookTitle, .listWithDividers__item a[href*="/book/show/"]')
        urls_to_visit = []
        for link in book_links:
            href = await link.evaluate("el => el.href")
            if href.startswith('/'):
                href = f"https://www.goodreads.com{href}"
            urls_to_visit.append(href)
            
        book1_link = ""
        if urls_to_visit:
            book1_link = urls_to_visit[0]
        
        urls_to_visit = list(set(urls_to_visit))
        num_books = len(urls_to_visit)
        print(f"    Found {num_books} books in series.")
        
        if num_books == 0:
            raise BlockedError("0 books found (silent block)")
        
        # Initialize or fetch state
        str_row = str(row_idx)
        async with state_lock:
            if str_row not in global_state:
                global_state[str_row] = {
                    "scraped_urls": [],
                    "total_pages": 0,
                    "agency": "",
                    "book1_link": book1_link,
                    "num_books": num_books
                }
            row_state = global_state[str_row]
            
        agency = row_state["agency"]
        total_pages = row_state["total_pages"]
        
        for b_url in urls_to_visit:
            if global_stop_event.is_set():
                raise BlockedError("Global stop triggered by another tab")
                
            if b_url in row_state["scraped_urls"]:
                print(f"      - Skipping already scraped book ({b_url})")
                continue
                
            pages, bk_agency, _ = await get_book_details(page, b_url)
            print(f"      - {pages} pages ({b_url})")
            
            total_pages += pages
            if not agency and bk_agency:
                agency = bk_agency
                
            # Update state immediately
            async with state_lock:
                global_state[str_row]["scraped_urls"].append(b_url)
                global_state[str_row]["total_pages"] = total_pages
                global_state[str_row]["agency"] = agency
            await save_state(global_state, state_lock)
            
        return num_books, total_pages, row_state.get("book1_link", book1_link), agency
    except BlockedError:
        raise
    except Exception as e:
        raise BlockedError(f"Blocked or error during series scrape: {e}")

async def process_row(row_idx, ws, col_map, context, semaphore, excel_lock, wb, global_state, state_lock, global_stop_event):
    s_no = ws.cell(row=row_idx, column=col_map['S. No.']).value
    title = ws.cell(row=row_idx, column=col_map['Title']).value or ""
    author = ws.cell(row=row_idx, column=col_map['Author Name']).value or ""
    
    series_link = ws.cell(row=row_idx, column=col_map['GR Series Link']).value or ""
    book1_link = ws.cell(row=row_idx, column=col_map['GR Book 1 link']).value or ""
    agency = ws.cell(row=row_idx, column=col_map['Agency (if)']).value or ""
    
    async with semaphore:
        if global_stop_event.is_set():
            return
            
        print(f"\n[Row {row_idx}] (S.No {s_no}) Processing: {title}")
        page = await context.new_page()
        try:
            has_series = isinstance(series_link, str) and series_link.startswith('http')
            
            if not has_series:
                print(f"  [Row {row_idx}] Series link missing. Searching for book...")
                target_book = book1_link if (isinstance(book1_link, str) and book1_link.startswith('http')) else None
                
                if not target_book:
                    target_book = await search_book_on_goodreads(page, str(title), str(author))
                    if target_book:
                        # We don't save to excel yet until everything finishes
                        book1_link = target_book
                            
                if target_book:
                    # Find series link from book page
                    pages, bk_agency, found_series = await get_book_details(page, target_book)
                    if found_series:
                        series_link = found_series
                        has_series = True
                        print(f"  [Row {row_idx}] Found Series Link: {series_link}")
                    else:
                        print(f"  [Row {row_idx}] No Series Link found on book page. Standalone.")
                        async with excel_lock:
                            ws.cell(row=row_idx, column=col_map['No. of books in the series']).value = 1
                            ws.cell(row=row_idx, column=col_map['Page count']).value = pages
                            if bk_agency and not agency:
                                ws.cell(row=row_idx, column=col_map['Agency (if)']).value = bk_agency
                            ws.cell(row=row_idx, column=col_map['GR Book 1 link']).value = target_book
                            wb.save(INPUT_FILE)
                        return
                        
            if has_series:
                print(f"  [Row {row_idx}] Scraping series: {series_link}")
                num_books, total_pages, found_book1, found_agency = await get_series_pages(page, series_link, row_idx, global_state, state_lock, global_stop_event)
                
                # If we get here, series scrape completed fully without BlockedError
                async with excel_lock:
                    ws.cell(row=row_idx, column=col_map['No. of books in the series']).value = num_books
                    ws.cell(row=row_idx, column=col_map['Page count']).value = total_pages
                    ws.cell(row=row_idx, column=col_map['GR Series Link']).value = series_link
                    
                    if (not book1_link or str(book1_link).strip() == "") and found_book1:
                        ws.cell(row=row_idx, column=col_map['GR Book 1 link']).value = found_book1
                        
                    if (not agency or str(agency).strip() == "") and found_agency:
                        ws.cell(row=row_idx, column=col_map['Agency (if)']).value = found_agency
                        
                    wb.save(INPUT_FILE)
                    print(f"  [Row {row_idx}] Saved: {num_books} books, {total_pages} pages")
                
                # Clear state for this row since it's fully complete
                async with state_lock:
                    if str(row_idx) in global_state:
                        del global_state[str(row_idx)]
                await save_state(global_state, state_lock)
                    
        except BlockedError as e:
            print(f"  [Row {row_idx}] BLOCKED by Goodreads! Saving state and halting this row. Reason: {e}")
            global_stop_event.set()
        except Exception as e:
            if "TargetClosedError" in str(type(e)) or "Timeout" in str(e):
                print(f"  [Row {row_idx}] BLOCKED by Goodreads! Saving state and halting this row. Reason: {e}")
                global_stop_event.set()
            else:
                print(f"  [Row {row_idx}] Failed to process: {e}")
        finally:
            await page.close()

async def main():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    if 'Romantasy v2' not in wb.sheetnames:
        print("Sheet 'Romantasy v2' not found!")
        return
    ws = wb['Romantasy v2']
    
    header_row = 2
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            col_map[str(val).strip()] = col_idx
            
    required_cols = ['S. No.', 'Title', 'Author Name', 'GR Book 1 link', 'Agency (if)', 'GR Series Link', 'No. of books in the series', 'Page count']
    for req in required_cols:
        if req not in col_map:
            print(f"Required column '{req}' not found in headers!")
            return
            
    global_state = await load_state()
    state_lock = asyncio.Lock()
    global_stop_event = asyncio.Event()
    
    # Aggressive scraper with 5 tabs concurrently
    semaphore = asyncio.Semaphore(5)
    excel_lock = asyncio.Lock()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        print("Starting Login Phase...")
        login_page = await context.new_page()
        scraper = GoodreadsScraper(headless=False)
        await scraper.login_to_goodreads(login_page)
        await login_page.close()
        print("Login complete. Starting concurrent scrape...")
        
        tasks = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if 285 <= row_idx <= 322:
                books_val = ws.cell(row=row_idx, column=col_map['No. of books in the series']).value
                pages_val = ws.cell(row=row_idx, column=col_map['Page count']).value
                
                if books_val not in [0, "0", 0.0, None, ""] and pages_val not in [0, "0", 0.0, None, ""]:
                    # Skipping row quietly to avoid console spam for thousands of rows
                    continue
                    
                tasks.append(process_row(row_idx, ws, col_map, context, semaphore, excel_lock, wb, global_state, state_lock, global_stop_event))
            
        print(f"Found {len(tasks)} tasks to run.")
        if tasks:
            await asyncio.gather(*tasks)
        else:
            print("No tasks found in the specified range.")
            
        await browser.close()
    
    print(f"Scraping run complete.")

if __name__ == "__main__":
    asyncio.run(main())
