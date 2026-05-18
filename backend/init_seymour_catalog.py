import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

SOURCE_EXCEL = r"E:\Internship\PocketFM\Seymour_Agency_Authors.xlsx"
OUTPUT_EXCEL = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def format_excel_sheet(ws):
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Premium deep-teal branding colors
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Header Styling
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    # Data Rows Styling
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments
            if col_idx in [5, 6, 7]:  # Numbers & Ratings
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:  # Publisher / Yes/No / Sub-genre
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Apply native Excel auto-filters
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def initialize_catalog():
    if not os.path.exists(SOURCE_EXCEL):
        print(f"Error: Scraped authors file not found at {SOURCE_EXCEL}")
        return

    print(">>> Loading scraped authors...")
    wb_src = openpyxl.load_workbook(SOURCE_EXCEL, data_only=True)
    ws_src = wb_src.active
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Seymour Media Catalog"
    
    # Append the headers
    ws_out.append(HEADERS)
    
    # Read rows from source (excluding header)
    count = 0
    for r in range(2, ws_src.max_row + 1):
        author_name = ws_src.cell(row=r, column=1).value
        category = ws_src.cell(row=r, column=2).value
        
        if not author_name:
            continue
            
        # Map values to the 11 columns
        row_data = [
            "",                        # Name of Series (empty for input)
            author_name,               # Author Name
            "The Seymour Agency",       # Publisher (Default)
            "N/A",                     # GoodReads series link
            "N/A",                     # Number of PRIMARY books in the series
            "N/A",                     # Rating (out of 5) of Primary Book 1
            "N/A",                     # Ratings (#) of Primary Book 1
            "N/A",                     # Synopsis (if available)
            "N/A",                     # Romantasy = Yes or No?
            "N/A",                     # Romantasy Sub-Genre of series
            "N/A"                      # Name of agent (or N/A)
        ]
        ws_out.append(row_data)
        count += 1
        
    # Apply master styling
    format_excel_sheet(ws_out)
    
    # Save Catalog
    print(f">>> Saving initialized 11-column catalog to {OUTPUT_EXCEL}...")
    wb_out.save(OUTPUT_EXCEL)
    print(f"Success! Initialized {count} authors in the 11-column master format!")

if __name__ == "__main__":
    initialize_catalog()
