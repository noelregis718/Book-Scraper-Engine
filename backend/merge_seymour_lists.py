import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from copy import copy

CATALOG_FILE = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"
BOOKS_FILE = r"E:\Internship\PocketFM\seymour_agency_books.xlsx"

def duplicate_row(ws, source_row_idx, target_row_idx):
    ws.insert_rows(target_row_idx)
    # Copy values and styling from source to target
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)
        
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            
    # Set the height for the new row
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height

def merge_lists():
    if not os.path.exists(CATALOG_FILE):
        print(f"Error: Master catalog file not found at {CATALOG_FILE}")
        return
    if not os.path.exists(BOOKS_FILE):
        print(f"Error: Featured books file not found at {BOOKS_FILE}")
        return

    print(">>> Loading featured books list...")
    wb_books = openpyxl.load_workbook(BOOKS_FILE, data_only=True)
    ws_books = wb_books.active
    
    # Parse the 24 books (Rows 5 to 28)
    books_by_author = {}
    for r in range(5, 29):
        author_name = ws_books.cell(row=r, column=2).value
        book_title = ws_books.cell(row=r, column=3).value
        
        if not author_name or not book_title:
            continue
            
        author_clean = str(author_name).strip().lower()
        if author_clean not in books_by_author:
            books_by_author[author_clean] = []
        books_by_author[author_clean].append(str(book_title).strip())
        
    print(f"Loaded books for {len(books_by_author)} unique featured authors.")

    print(">>> Loading master catalog...")
    wb_cat = openpyxl.load_workbook(CATALOG_FILE, data_only=False)
    ws_cat = wb_cat.active
    
    initial_rows = ws_cat.max_row
    print(f"Initial catalog rows: {initial_rows}")
    
    updated_count = 0
    duplicate_count = 0
    
    # Loop backwards to handle dynamically inserted rows safely
    for row_idx in range(initial_rows, 1, -1):
        author_name = ws_cat.cell(row=row_idx, column=2).value
        if not author_name:
            continue
            
        author_clean = str(author_name).strip().lower()
        if author_clean in books_by_author:
            books = books_by_author[author_clean]
            
            # 1. Update the first book in-place
            ws_cat.cell(row=row_idx, column=1, value=books[0])
            updated_count += 1
            print(f"  [Match] Set '{author_name}' Series to '{books[0]}'")
            
            # 2. If there are multiple books, insert duplicate rows directly below
            for i in range(1, len(books)):
                target_row = row_idx + i
                duplicate_row(ws_cat, row_idx, target_row)
                ws_cat.cell(row=target_row, column=1, value=books[i])
                duplicate_count += 1
                print(f"  [Duplicate] Added row for '{author_name}' with Series '{books[i]}'")

    # Refresh auto-filters over the new final range
    final_rows = ws_cat.max_row
    ws_cat.auto_filter.ref = f"A1:K{final_rows}"
    
    print(f"\n>>> Saving updated master catalog to {CATALOG_FILE}...")
    wb_cat.save(CATALOG_FILE)
    print(f"Success! Updated {updated_count} authors and added {duplicate_count} duplicate rows.")
    print(f"Final catalog rows: {final_rows}")

if __name__ == "__main__":
    merge_lists()
