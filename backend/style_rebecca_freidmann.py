import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def apply_styling(file_path):
    print(f"Loading {file_path} for styling...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Freeze the top row
    ws.freeze_panes = "A2"
    
    # Define styles - New color for Rebecca Freidmann (Deep Purple)
    header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Column widths based on 11 column format
    col_widths = {
        'Name of Series': 40,
        'Author Name': 25,
        'Publisher': 15,
        'GoodReads series link': 40,
        'Number of PRIMARY books in the series': 15,
        'Rating (out of 5) of Primary Book 1': 12,
        'Ratings (#) of Primary Book 1': 15,
        'Synopsis (if available)': 65,
        'Romantasy = Yes or No?': 15,
        'Romantasy Sub-Genre of series': 25,
        'Name of agent': 15,
        'Name of Agent': 15
    }
    
    print("Applying styles to columns and cells...")
    # Apply widths and basic cell styling
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value
        
        # Set column width
        width = col_widths.get(header_val, 20)
        ws.column_dimensions[col_letter].width = width
        
        # Style cells in column
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = thin_border
            
            # Header specific styling
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print("Calculating row heights...")
    # Auto-fit row heights based on content length
    for row_num in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = ws.cell(row=1, column=col_idx).value
            cell = ws.cell(row=row_num, column=col_idx)
            val = str(cell.value) if cell.value is not None else ""
            
            col_width = col_widths.get(header_val, 20)
            chars_per_line = max(int(col_width * 1.1), 10)
            
            newline_count = 1
            for line in val.split('\n'):
                wrapped = max(1, -(-len(line) // chars_per_line)) # Ceiling division
                newline_count += wrapped
                
            max_lines = max(max_lines, newline_count)
            
        ws.row_dimensions[row_num].height = max(min(max_lines * 14, 300), 20)

    print(f"Saving styled workbook to {file_path}...")
    wb.save(file_path)
    print("Styling Done!")

if __name__ == "__main__":
    filepath = r"E:\Internship\PocketFM\rebecca_freidmann_authors.xlsx"
    apply_styling(filepath)
