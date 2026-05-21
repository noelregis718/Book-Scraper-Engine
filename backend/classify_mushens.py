"""
Re-classifies the Mushens main sheet:
  - Column I  (Romantasy = Yes or No?)
  - Column J  (Romantasy Sub-Genre of series) — ONLY one of the 12 allowed sub-genres

Strategy: AUTHOR WHITELIST. Romantasy is a specific genre (fantasy where romance is
central). Mushens reps mostly thriller / historical / contemporary / mystery authors,
with only a handful of true romantasy/fantasy writers. The whitelist below names
authors whose books reliably belong in romantasy/fantasy buckets; everyone else
defaults to No.

Sub-genre picker: word-boundary regex against title + synopsis, in priority order
(more specific buckets before broad ones).
"""

import os
import re
import openpyxl
from format_mushens import format_mushens

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "Mushens_Entertainment_Bestsellers.xlsx")
MAIN_SHEET = "Mushens Entertainment"


# Lowercase substring match against the Author Name cell.
# Value = sub-genre to assign (or "auto" to derive from title/synopsis).
ROMANTASY_AUTHORS = {
    "saara el-arifi":     "auto",
    "hannah kaner":       "auto",
    "andrea stewart":     "auto",
    "taran matharu":      "auto",
    "l.r. lam":           "auto",
    "laura lam":          "auto",
    "elizabeth may":      "auto",
    "jennifer saint":     "Mythology, Legend & Fairy Tale Retelling",
    "peter newman":       "auto",
    "jen williams":       "auto",
    "inbali iserles":     "auto",
    "amy mcculloch":      "auto",
    "lou morgan":         "auto",
    "liz de jager":       "auto",
    "kate dylan":         "auto",
}


# Priority-ordered sub-genre keyword map (word-boundary matched)
KEYWORD_MAP = [
    ("Mythology, Legend & Fairy Tale Retelling", [
        r"retelling", r"retold", r"mythology", r"greek myth", r"norse myth",
        r"fairy tale", r"fairytale", r"arthurian", r"olympian",
        r"ariadne", r"atalanta", r"elektra", r"electra", r"hera", r"medusa",
        r"circe", r"persephone", r"hades", r"orpheus",
        r"boudica", r"boudicca", r"cleopatra",
    ]),
    ("War College / Military Academy", [
        r"war college", r"military academy", r"dragon rider", r"fourth wing",
        r"basgiath", r"flight school", r"combat training", r"lethal training",
        r"summoner", r"rider academy", r"bonded dragon", r"warhammer",
    ]),
    ("High-Stakes Games & Deadly Trials", [
        r"deadly trial", r"magical tournament", r"deadly games", r"magical contest",
        r"battle arena", r"death match", r"blood game", r"hunger games",
        r"killing game", r"lethal contest",
    ]),
    ("Dark Academia Romantasy", [
        r"dark academia", r"magical academy", r"magic school", r"secret society",
        r"university of magic", r"arcane academy", r"cursed school",
    ]),
    ("Werewolf / Shifter Romance", [
        r"werewolf", r"shifter", r"pack alpha", r"true mate", r"omegaverse",
        r"shapeshifter", r"wolf shifter", r"bear shifter", r"dragon shifter", r"luna",
    ]),
    ("Monster Romance (Non-Shifter)", [
        r"monster romance", r"tentacle", r"orc romance", r"kraken", r"minotaur",
        r"beastman", r"alien romance",
    ]),
    ("Korean Romance Fantasy / Isekai", [
        r"isekai", r"reincarnated", r"villainess", r"transmigrat", r"saintess",
        r"korean fantasy", r"manhwa", r"regression",
    ]),
    ("Gothic Dark Romantasy", [
        r"gothic romance", r"haunted castle", r"vampire lord",
        r"immortal lord", r"blood magic", r"shadow court", r"cursed castle",
        r"dark romantasy",
    ]),
    ("Paranormal Romance", [
        r"paranormal romance", r"demon lover", r"succubus", r"necromancer",
        r"warlock", r"coven of witches", r"vampire romance",
        r"fae romance",
    ]),
    ("Cozy / Cottagecore", [
        r"cozy fantasy", r"cottagecore", r"magical bakery", r"small town magic",
        r"low-stakes fantasy", r"village witch", r"magical inn", r"cosy fantasy",
    ]),
    ("Urban / Contemporary Fantasy Romance", [
        r"urban fantasy", r"hidden magic", r"secret supernatural", r"hidden world",
        r"magic in the city", r"modern witch", r"contemporary fantasy",
    ]),
    ("High Fantasy Court Adventure", [
        r"fae", r"faerie", r"faery", r"elven", r"elves", r"fae court",
        r"high fantasy", r"royal court", r"magical kingdom", r"fantasy empire",
        r"godkiller", r"faebound", r"cursebound", r"bone shard",
        r"final strife", r"battle drum", r"ending fire",
        r"summoner", r"deathbringer", r"tainted khan",
        r"dragonfall", r"dragon", r"sorcerer", r"warlock",
        r"epic fantasy", r"vagrant", r"deathless",
    ]),
]


def _compile(keyword_map):
    compiled = []
    for sub, words in keyword_map:
        patterns = [re.compile(rf"\b{w}\b", re.IGNORECASE) for w in words]
        compiled.append((sub, patterns))
    return compiled


COMPILED_MAP = _compile(KEYWORD_MAP)


def pick_subgenre(text):
    for sub, patterns in COMPILED_MAP:
        if any(p.search(text) for p in patterns):
            return sub
    return "High Fantasy Court Adventure"


def classify(title, author, synopsis):
    auth_lc = (author or "").lower().strip()
    text = f"{title or ''}  {synopsis or ''}".lower()

    hint = None
    for known, sub in ROMANTASY_AUTHORS.items():
        if known in auth_lc:
            hint = sub
            break

    if hint is None:
        return "No", ""

    if hint != "auto":
        return "Yes", hint

    return "Yes", pick_subgenre(text)


def main():
    if not os.path.exists(XLSX):
        print(f"Workbook not found: {XLSX}")
        return

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[MAIN_SHEET]

    yes_count = 0
    sub_counts = {}
    total = 0
    yes_examples = []

    for r in range(2, ws.max_row + 1):
        title    = ws.cell(row=r, column=1).value
        author   = ws.cell(row=r, column=2).value
        synopsis = ws.cell(row=r, column=8).value
        if not title:
            continue
        total += 1

        yes_no, sub = classify(title, author, synopsis)

        ws.cell(row=r, column=9).value  = yes_no
        ws.cell(row=r, column=10).value = sub

        if yes_no == "Yes":
            yes_count += 1
            sub_counts[sub] = sub_counts.get(sub, 0) + 1
            if len(yes_examples) < 25:
                yes_examples.append((title, author, sub))

    wb.save(XLSX)

    print(f"Classified {total} rows. Romantasy=Yes: {yes_count}  No: {total - yes_count}")
    print("\nSub-genre breakdown (Yes rows):")
    for sub, n in sorted(sub_counts.items(), key=lambda x: -x[1]):
        print(f"  {n:3d}  {sub}")
    print("\nSample Yes classifications:")
    for t, a, s in yes_examples:
        print(f"  - {str(t)[:42]:42}  ({str(a)[:22]:22})  -> {s}")

    format_mushens(XLSX, XLSX)
    print("\nStyling re-applied.")


if __name__ == "__main__":
    main()
