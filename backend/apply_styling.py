import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Premium header styling (Dark blue background, white bold text)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    center_aligned_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_text = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply styling to Header (Row 1)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = thin_border
        
    # Apply styling to Data Rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            # Default to left align with wrap for all cells
            cell.alignment = wrap_text
            cell.border = thin_border
            
            # Center align specific numeric or short text columns
            # E (Number of books), F (Rating), G (Ratings #), I (Romantasy Yes/No)
            if col in [5, 6, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust Column Widths Dynamically
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        
        # Determine optimal width based on first few rows to avoid crazy wide columns from synopsis
        for cell in col[:10]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        adjusted_width = (max_length + 2) * 1.2
        
        # Hard limits on width to keep it readable
        if adjusted_width > 50:
            adjusted_width = 50
        elif adjusted_width < 15:
            adjusted_width = 15
            
        ws.column_dimensions[column].width = adjusted_width

    # Explicitly set the Synopsis column (usually H or 8) to be wider for readability
    ws.column_dimensions['H'].width = 80

    # Freeze the top row so headers stay visible on scroll
    ws.freeze_panes = "A2"

    wb.save(file_path)
    print("Styling applied successfully!")

if __name__ == "__main__":
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else 'Standard_11_Column_Format.xlsx'
    style_excel(file_path)
