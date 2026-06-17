import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

EXCEL_FILE = "e:/Internship/PocketFM/books_from_images.xlsx"

def style_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Premium styles
    # Header: Deep sleek slate color with white text
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    # Keep rows white as requested
    alt_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Subtle elegant borders instead of stark black
    subtle_side = Side(style='thin', color="D3D3D3")
    subtle_border = Border(left=subtle_side, right=subtle_side, top=subtle_side, bottom=subtle_side)

    # Apply header styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = subtle_border

    # Apply alternating rows and specific alignments
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_alt = (row_idx % 2 == 0)
        current_fill = alt_row_fill if is_alt else white_fill
        
        for cell in row:
            cell.fill = current_fill
            cell.border = subtle_border
            
            header_val = ws.cell(row=1, column=cell.column).value
            if header_val in ["Synopsis (if available)", "GoodReads series link", "Name of Series", "Author Name"]:
                cell.alignment = left_aligned
            else:
                cell.alignment = center_aligned

    # Auto-adjust column widths with bounds
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        adjusted_width = (max_length + 2)
        if adjusted_width > 55:
            adjusted_width = 55
        elif adjusted_width < 15:
            adjusted_width = 15
            
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
    print(f"Excel file {file_path} styled with Premium format successfully!")

if __name__ == "__main__":
    target_file = sys.argv[1] if len(sys.argv) > 1 else EXCEL_FILE
    style_excel(target_file)
