import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIGURATION ---
OLD_FILE = r"E:\Internship\PocketFM\Knight Agency_OLD.xlsx"
NEW_FILE = r"E:\Internship\PocketFM\Knight Agency.xlsx"

def clean_val(val):
    return str(val).strip().lower() if pd.notna(val) else ""

def style_excel(file_path):
    try:
        if not os.path.exists(file_path): return
        wb = load_workbook(file_path)
        ws = wb.active
        if ws is None:
            raise ValueError("Active worksheet is None")
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Style Headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style Data Rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Adjust Column Widths
        column_widths = {
            "A": 30, "B": 25, "C": 20, "D": 35, "E": 15,
            "F": 12, "G": 12, "H": 60, "I": 18, "J": 30, "K": 25
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        wb.save(file_path)
        print(f"  [Style] Applied formatting to {os.path.basename(file_path)}")
    except Exception as e:
        print(f"  [Warning] Styling failed: {e}")

def merge_files():
    print(f"\n>>> [MERGE MISSION] Synchronizing Unique Titles")
    
    if not os.path.exists(OLD_FILE):
        print(f"  [Error] Old file not found: {OLD_FILE}")
        return
    
    if not os.path.exists(NEW_FILE):
        print(f"  [Error] New file not found: {NEW_FILE}")
        return

    # Load data
    print(f"  [1/4] Loading Excel files...")
    df_old = pd.read_excel(OLD_FILE)
    df_new = pd.read_excel(NEW_FILE)
    
    # Create unique keys for deduplication
    print(f"  [2/4] Identifying unique titles...")
    df_old['temp_key'] = df_old.apply(lambda r: f"{clean_val(r.get('Name of Series', ''))}|{clean_val(r.get('Author Name', ''))}", axis=1)
    df_new['temp_key'] = df_new.apply(lambda r: f"{clean_val(r.get('Name of Series', ''))}|{clean_val(r.get('Author Name', ''))}", axis=1)
    
    existing_keys = set(df_old['temp_key'].unique())
    
    # Find delta
    df_unique_new = df_new[~df_new['temp_key'].isin(existing_keys)].copy()
    
    if df_unique_new.empty:
        print(f"  [Done] No new unique titles found to add. Everything is already in the old file.")
        return

    print(f"  [3/4] Found {len(df_unique_new)} new unique titles to migrate.")
    
    # Clean up temp columns before merging
    df_old.drop(columns=['temp_key'], inplace=True)
    df_unique_new.drop(columns=['temp_key'], inplace=True)
    
    # Merge and Save
    print(f"  [4/4] Updating {os.path.basename(OLD_FILE)}...")
    df_combined = pd.concat([df_old, df_unique_new], ignore_index=True)
    
    # Final write with permission retry
    import time
    max_retries = 5
    for attempt in range(max_retries):
        try:
            df_combined.to_excel(OLD_FILE, index=False)
            style_excel(OLD_FILE)
            print(f"\nSUCCESS: Added {len(df_unique_new)} titles to {OLD_FILE}.")
            print(f"Total titles now in master file: {len(df_combined)}")
            break
        except PermissionError:
            print(f"  [Warning] Permission Denied. Is the Excel file open? Retrying in 5s... ({attempt+1}/{max_retries})")
            time.sleep(5)
        except Exception as e:
            print(f"  [Error] Save failed: {e}")
            break

if __name__ == "__main__":
    merge_files()
