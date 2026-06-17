import openpyxl

path = r"E:\Internship\Pocket FM Documents\The Unter Agency.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

print("Header (A1) Font:", ws['A1'].font.name, ws['A1'].font.size, ws['A1'].font.bold, ws['A1'].font.color)
print("Header (A1) Fill:", ws['A1'].fill.start_color.index if ws['A1'].fill.start_color else "None")
print("Header (A1) Alignment:", ws['A1'].alignment.horizontal, ws['A1'].alignment.vertical, ws['A1'].alignment.wrap_text)

print("\nData (A2) Font:", ws['A2'].font.name, ws['A2'].font.size, ws['A2'].font.bold, ws['A2'].font.color)
print("Data (A2) Alignment:", ws['A2'].alignment.horizontal, ws['A2'].alignment.vertical, ws['A2'].alignment.wrap_text)

print("\nColumn Widths:")
for i in range(1, 10):
    letter = openpyxl.utils.get_column_letter(i)
    print(f"{letter}: {ws.column_dimensions[letter].width}")
