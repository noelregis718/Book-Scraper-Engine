import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def style_excel():
    file_path = "e:/Internship/PocketFM/1852 Media.xlsx"
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 1. Premium Colors and Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Deep Slate
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light Slate
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), 
                         right=Side(style='thin', color='E2E8F0'), 
                         top=Side(style='thin', color='E2E8F0'), 
                         bottom=Side(style='thin', color='E2E8F0'))

    # 2. Freeze Top Row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    # 3. Apply styling row by row
    for row_idx, row in enumerate(ws.iter_rows()):
        is_header = (row_idx == 0)
        fill_color = header_fill if is_header else (zebra_fill if row_idx % 2 == 1 else white_fill)
        
        for cell in row:
            if is_header:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
            cell.fill = fill_color
            cell.border = thin_border

    # 4. Auto-adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Ignore massively long synopsis text for width calculation
                    if len(str(cell.value)) < 150:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Add padding and cap the maximum width at 45 so it doesn't stretch forever
        adjusted_width = min(max_length + 2, 45)
        ws.column_dimensions[col_letter].width = adjusted_width

    print("Saving...")
    wb.save(file_path)
    print("Styling applied successfully!")

if __name__ == "__main__":
    style_excel()
