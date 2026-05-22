import re
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side

def append_authors():
    md_file = r"C:\Users\noelr\.gemini\antigravity-ide\brain\37f9affe-cb68-44e7-a2bf-9f63822bb435\.system_generated\steps\52\content.md"
    excel_file = r"e:\Internship\PocketFM\madwoman_literary_scraped_books.xlsx"
    
    authors = []
    with open(md_file, 'r', encoding='utf-8') as f:
        for line in f:
            if line.startswith("###### "):
                name = line.strip().replace("###### ", "")
                if "[" not in name and "]" not in name:  # filter out the link at the bottom
                    authors.append(name)
                    
    # remove any potential duplicates within the scraped list if necessary, or just append all
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Alternating row fills
    fill_even = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF",  fill_type="solid")
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style='thin', color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_row = ws.max_row + 1
    
    # The columns are 1 to 11. 
    # B is 2 (Author Name), K is 11 (Name of agent)
    
    for idx, author in enumerate(authors):
        current_row = start_row + idx
        
        # Write values
        ws.cell(row=current_row, column=2, value=author)  # Author Name
        ws.cell(row=current_row, column=11, value="Mad Woman Literary") # Name of agent
        
        # Apply styling
        row_fill = fill_even if current_row % 2 == 0 else fill_odd
        for col_idx in range(1, 12):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = row_fill
            cell.alignment = align_top
            cell.border = thin_border
            
    wb.save(excel_file)
    print(f"Appended {len(authors)} authors starting from row {start_row}.")

if __name__ == '__main__':
    append_authors()
