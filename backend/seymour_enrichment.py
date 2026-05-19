import asyncio
import os
import sys
import openpyxl
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure backend directory is in path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

# --- CONFIGURATION ---
CATALOG_FILE = r"E:\Internship\PocketFM\Seymour_Media_Catalog.xlsx"
CONCURRENCY_LIMIT = 5   # Scrape exactly 5 books at a time!
TOTAL_TARGET_BOOKS = 100 # Process exactly the first 100 unscraped books!
HEADLESS = False        # Set to False so you can see the 5 browser tabs working concurrently live!

HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

def safe_int(val):
    if not val or str(val).strip() == "N/A":
        return "N/A"
    try:
        clean_val = str(val).replace(',', '').split('.')[0].strip()
        return int(clean_val)
    except:
        return "N/A"

def safe_float(val):
    if not val or str(val).strip() == "N/A":
        return "N/A"
    try:
        clean_val = str(val).replace(',', '').strip()
        return float(clean_val)
    except:
        return "N/A"

def format_excel_sheet(ws):
    """Applies premium deep-teal styling to the worksheet in-place."""
    ws.views.sheetView[0].showGridLines = True
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            cell.font = data_font
            
            # Alignments matching seymour style
            if col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [3, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

def find_target_enrichment_rows(ws, total_target):
    """Scans the spreadsheet and collects exactly the first `total_target` unscraped book rows."""
    target_rows = []
    has_changes = False
    
    for r in range(2, ws.max_row + 1):
        book_title = ws.cell(row=r, column=1).value
        gr_link = ws.cell(row=r, column=4).value
        
        # Valid book row has a title in Col A
        if book_title and str(book_title).strip() != "" and book_title != "N/A":
            title_str = str(book_title).strip()
            
            # Check if book title starts with a number (digit 0-9)
            if title_str[0].isdigit():
                # Only update if it wasn't already marked as N/A or set to a URL
                if not gr_link or str(gr_link).strip() == "" or gr_link == "N/A":
                    print(f"  [Auto-Filter] Row {r}: Book title '{title_str}' starts with a number. Marking as N/A.")
                    ws.cell(row=r, column=4, value="N/A")
                    ws.cell(row=r, column=5, value=1)
                    ws.cell(row=r, column=6, value="N/A")
                    ws.cell(row=r, column=7, value="N/A")
                    ws.cell(row=r, column=8, value="N/A")
                    ws.cell(row=r, column=9, value="No")
                    ws.cell(row=r, column=10, value="N/A")
                    has_changes = True
                continue
                
            # If it starts with an alphabet, we proceed with checking if it needs scraping
            if not gr_link or str(gr_link).strip() == "" or gr_link == "N/A" or not str(gr_link).strip().startswith("http"):
                author_name = ws.cell(row=r, column=2).value
                target_rows.append((r, title_str, str(author_name or "Unknown").strip()))
                if len(target_rows) >= total_target:
                    break
                    
    return target_rows, has_changes

async def enrich_single_row(sem, context, r, book_title, author_name, gr_scraper, idx, total):
    """Worker task that runs inside the concurrency-limited semaphore."""
    async with sem:
        print(f"[{idx}/{total}] Scraping row {r}: '{book_title}' by '{author_name}'...")
        try:
            gr_data = await gr_scraper.scrape_goodreads_data(context, book_title, author_name)
            if gr_data:
                print(f"[{idx}/{total}] Success! Mapped data for row {r}.")
                return r, gr_data
            else:
                print(f"[{idx}/{total}] Failed: Goodreads returned empty details for row {r}.")
                return r, None
        except Exception as e:
            print(f"[{idx}/{total}] Error: Failed scraping row {r}: {e}")
            return r, None

async def run_enrichment():
    print("=" * 60)
    print("      SEYMOUR AGENCY PARALLEL ENRICHMENT PIPELINE")
    print("=" * 60)
    print(f"Spreadsheet: {CATALOG_FILE}")
    print(f"Target size: {TOTAL_TARGET_BOOKS} books | Concurrency: {CONCURRENCY_LIMIT} tabs\n")

    if not os.path.exists(CATALOG_FILE):
        print(f"[Error] Catalog spreadsheet {CATALOG_FILE} not found!")
        return

    # Load workbook to scan for the target rows
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active

    print("[System] Scanning sheet for target unscraped book rows and filtering digits...")
    target_rows, has_changes = find_target_enrichment_rows(ws, TOTAL_TARGET_BOOKS)
    
    if has_changes:
        print("[System] Saving auto-filtered digit rows...")
        try:
            format_excel_sheet(ws)
            wb.save(CATALOG_FILE)
            print("[System] Spreadsheet saved with filtered rows!")
        except Exception as e:
            print(f"[Error] Could not save spreadsheet after filtering: {e}")
            
    wb.close()

    if not target_rows:
        print("[Success] All books in the spreadsheet are already fully enriched!")
        return

    print(f"[Target Rows] Found {len(target_rows)} rows to enrich: {[r[0] for r in target_rows]}")

    # Initialize Goodreads Scraper
    gr_scraper = GoodreadsScraper(headless=HEADLESS)
    sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
    
    async with async_playwright() as p:
        print("[System] Launching Playwright browser...")
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )
        
        # Prepare parallel async tasks
        tasks = []
        for idx, (r, book_title, author_name) in enumerate(target_rows, 1):
            tasks.append(enrich_single_row(sem, context, r, book_title, author_name, gr_scraper, idx, len(target_rows)))
            
        # Execute in parallel (5 at a time)
        print(f"[Crawl] Gathering results concurrently ({CONCURRENCY_LIMIT} tabs max)...")
        results = await asyncio.gather(*tasks, return_exceptions=True)
        await browser.close()
        
        # Save results sequentially to Excel
        print("\n>>> Writing results to Excel sheet sequentially...")
        wb = load_workbook(CATALOG_FILE)
        ws = wb.active
        
        success_count = 0
        failure_count = 0
        
        for item in results:
            if isinstance(item, tuple) and item[1] is not None:
                r, gr_data = item
                
                # Col 4: GoodReads series link
                series_url = gr_data.get("GoodReads_Series_URL", "N/A")
                if series_url == "N/A" or not series_url:
                    series_url = gr_data.get("GoodReads_Book_URL", "N/A")
                ws.cell(row=r, column=4, value=series_url)
                
                # Col 5: Number of PRIMARY books in the series
                ws.cell(row=r, column=5, value=safe_int(gr_data.get("Num_Primary_Books", "1")))
                
                # Col 6: Rating (out of 5) of Primary Book 1
                ws.cell(row=r, column=6, value=safe_float(gr_data.get("Book1_Rating", "N/A")))
                
                # Col 7: Ratings (#) of Primary Book 1
                ws.cell(row=r, column=7, value=safe_int(gr_data.get("Book1_Num_Ratings", "N/A")))
                
                # Col 8: Synopsis (if available)
                ws.cell(row=r, column=8, value=gr_data.get("Description", "N/A"))
                
                # Col 9: Romantasy = Yes or No?
                ws.cell(row=r, column=9, value=gr_data.get("Romantasy_Subgenre", "No"))
                
                # Col 10: Romantasy Sub-Genre of series
                ws.cell(row=r, column=10, value=gr_data.get("Sub_Genre", "N/A"))
                
                success_count += 1
            else:
                # If scraping failed, set placeholders
                if isinstance(item, tuple):
                    r = item[0]
                    ws.cell(row=r, column=4, value="N/A")
                    ws.cell(row=r, column=5, value=1)
                    ws.cell(row=r, column=6, value="N/A")
                    ws.cell(row=r, column=7, value="N/A")
                    ws.cell(row=r, column=8, value="N/A")
                    ws.cell(row=r, column=9, value="No")
                    ws.cell(row=r, column=10, value="N/A")
                failure_count += 1
                
        # Format and save
        print(">>> Applying styling and saving workbook...")
        format_excel_sheet(ws)
        wb.save(CATALOG_FILE)
        print(">>> Catalog saved successfully!")
        
        print("\n" + "=" * 60)
        print("                 ENRICHMENT BATCH COMPLETE")
        print("=" * 60)
        print(f"Total Attempted:      {len(target_rows)}")
        print(f"Successful Hits:      {success_count}")
        print(f"Failed/Skipped:       {failure_count}")
        print(f"Output Catalog:       {CATALOG_FILE}")
        print("=" * 60 + "\n")
        
        # Auto-open on Windows
        if os.name == 'nt':
            print("  [System] Auto-opening sheet for review...")
            os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    asyncio.run(run_enrichment())
