import openpyxl
import urllib.parse

EXCEL_FILE = '../New_Agency_Template.xlsx'

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

count = 0
for row in range(2, ws.max_row + 1):
    title = ws.cell(row=row, column=1).value
    author = ws.cell(row=row, column=2).value
    link = ws.cell(row=row, column=4).value
    
    if title and str(title).strip() and str(title).strip() != "N/A":
        # Check if the link is missing or failed in previous runs
        if not link or str(link).strip() in ["N/A", "nan", "Not Found", "Error"]:
            # Build a Goodreads Direct Search URL
            query = f"{title} {author}"
            encoded = urllib.parse.quote_plus(query)
            search_url = f"https://www.goodreads.com/search?q={encoded}"
            
            ws.cell(row=row, column=4).value = search_url
            count += 1
            
wb.save(EXCEL_FILE)
print(f"Instantly generated {count} smart search links!")
