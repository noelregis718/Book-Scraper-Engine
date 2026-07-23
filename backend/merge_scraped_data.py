import pandas as pd
import os
import openpyxl

extracted_file = r"e:\Internship\PocketFM\Romantasy_v2_extracted.xlsx"
scraped_file = r"e:\Internship\PocketFM\Romantasy_v2_Scraped.xlsx"

print("Loading data...")
# Both files should have the same row count. Extracted has header on row 2 (index 1).
# Scraped has header on row 1 (index 0) because we saved it with index=False.
df_extracted = pd.read_excel(extracted_file, header=1)
df_scraped = pd.read_excel(scraped_file)

if len(df_extracted) != len(df_scraped):
    print(f"Row count mismatch! Extracted: {len(df_extracted)}, Scraped: {len(df_scraped)}")
    exit(1)

cols_to_update = [
    'GR Book 1 link',
    'Agency (if)',
    'GR Series Link',
    'No. of books in the series',
    'Page count'
]

print("Updating extracted sheet with scraped details...")
for col in cols_to_update:
    if col in df_extracted.columns and col in df_scraped.columns:
        df_extracted[col] = df_scraped[col]

# Save back to extracted file
temp_file = r"e:\Internship\PocketFM\Romantasy_v2_extracted_temp.xlsx"
df_extracted.to_excel(temp_file, index=False)

# Re-apply the empty first row since header=1 meant the actual header was row 2.
print("Formatting to preserve original header position...")
wb = openpyxl.load_workbook(temp_file)
ws = wb.active
ws.insert_rows(1)
wb.save(extracted_file)
wb.close()

# Clean up temp file
os.remove(temp_file)

# Call the fixed styling script on the extracted sheet to make it look nice
# We can use the existing style_excel_fixed.py script which handles the template
import subprocess
try:
    print("Applying styles to the updated extracted sheet...")
    # The style_excel_fixed.py expects 'e:\Internship\PocketFM\Books_Scraping_Template.xlsx'
    # We will just write a quick styling block here instead.
    
    wb = openpyxl.load_workbook(extracted_file)
    ws = wb.active
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                         right=Side(style='thin', color='BFBFBF'), 
                         top=Side(style='thin', color='BFBFBF'), 
                         bottom=Side(style='thin', color='BFBFBF'))
                         
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col) # header is on row 2
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
        
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment_left_top
            cell.border = thin_border
            
            # Highlight links in blue
            if cell.value and str(cell.value).startswith('http'):
                cell.font = Font(color="0563C1", underline="single", name='Calibri', size=11)
                cell.hyperlink = cell.value
                
    ws.freeze_panes = 'A3'
    wb.save(extracted_file)
    print("Styling applied successfully!")
except Exception as e:
    print(f"Error applying styles: {e}")

print(f"Data successfully merged into {extracted_file}")
