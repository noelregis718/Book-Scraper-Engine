import asyncio
import os
import sys
from copy import copy
from openpyxl import load_workbook
import pandas as pd
from playwright.async_api import async_playwright

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper
from romantasy_analyzer import identify_romantasy_subgenre
from format_1488_catalog import format_excel_sheet

CATALOG_FILE = r"E:\Internship\PocketFM\kensington_authors_MASTER_1_to_1488.xlsx"
HEADERS = [
    "Name of Series",
    "Author Name",
    "Publisher",
    "GoodReads series link",
    "Number of PRIMARY books in the series",
    "Rating (out of 5) of Primary Book 1",
    "Ratings (#) of Primary Book 1",
    "Synopsis (if available)",
    "Romantasy = Yes or No?",
    "Romantasy Sub-Genre of series",
    "Name of agent"
]

async def process_book(row_idx, series_name, author_name, context, gr_scraper, ws, sem):
    async with sem:
        print(f"[Row {row_idx}] Searching for: '{series_name}' by '{author_name}'...")
        try:
            gr_data = await gr_scraper.scrape_goodreads_data(context, str(series_name), str(author_name))
            
            if gr_data:
                series_link = gr_data.get("GoodReads_Series_URL")
                book_link = gr_data.get("GoodReads_Book_URL")
                final_link = series_link if series_link and series_link != "N/A" else (book_link or "N/A")
                
                synopsis = gr_data.get("Description", "N/A")
                rom_sub = identify_romantasy_subgenre(synopsis, gr_data.get("Genre", ""))
                
                ws.cell(row=row_idx, column=4, value=final_link)
                ws.cell(row=row_idx, column=5, value=gr_data.get("Num_Primary_Books", 1) if final_link != "N/A" else "N/A")
                ws.cell(row=row_idx, column=6, value=gr_data.get("GoodReads_Rating", "N/A"))
                ws.cell(row=row_idx, column=7, value=gr_data.get("GoodReads_Rating_Count", "N/A"))
                ws.cell(row=row_idx, column=8, value=synopsis)
                ws.cell(row=row_idx, column=9, value="Yes" if rom_sub != "N/A" else "No")
                ws.cell(row=row_idx, column=10, value=rom_sub)
                
                print(f"    [Success] '{series_name}' mapped to URL!")
            else:
                print(f"    [Skip] Could not locate exact match for '{series_name}'.")
                
        except Exception as e:
            print(f"    [Error] Processing '{series_name}': {e}")

async def scrape_kensington_master():
    while True:
        print("\n>>> Loading Master 1488 Catalog...")
        wb = load_workbook(CATALOG_FILE)
        ws = wb.active
        
        # Read via pandas just to find target rows easily
        df = pd.read_excel(CATALOG_FILE)
        
        targets = []
        # Identify the first 100 rows that need scraping
        for idx, row in df.iterrows():
            title = str(row.get("Name of Series", "")).strip()
            author = str(row.get("Author Name", "")).strip()
            link = str(row.get("GoodReads series link", "")).strip()
            
            # Row must have a title, an author, and NO valid link yet
            if title and title.lower() not in ["nan", "n/a", "none"] and author and author.lower() not in ["nan", "n/a", "none"]:
                if not link or link.lower() in ["nan", "n/a", "none"]:
                    targets.append((idx + 2, title, author)) # +2 for 1-based openpyxl and headers
                    
            if len(targets) >= 100:
                break
                
        if not targets:
            print(">>> No unscraped books found! The catalog is fully enriched.")
            break

        print(f"\n>>> Starting Book-First Scraper for the next {len(targets)} books...")
        for t in targets:
            print(f"  Row {t[0]}: '{t[1]}' by {t[2]}")
            
        async with async_playwright() as p:
            print("\n[System] Launching Playwright browser in headed mode...")
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
            )
            
            # We deliberately DO NOT abort CSS so the user sees proper styling
            gr_scraper = GoodreadsScraper(headless=False)
            sem = asyncio.Semaphore(5)
            
            tasks = [process_book(r[0], r[1], r[2], context, gr_scraper, ws, sem) for r in targets]
            await asyncio.gather(*tasks)
                
            await browser.close()

        format_excel_sheet(ws)
        wb.save(CATALOG_FILE)
        print(f"\n>>> Success! {len(targets)} books aggressively enriched.")
        
        if os.name == 'nt':
            os.startfile(CATALOG_FILE)
            
        print("\n>>> Moving to the next batch of 100 automatically in 5 seconds...\n")
        await asyncio.sleep(5)

if __name__ == "__main__":
    asyncio.run(scrape_kensington_master())
