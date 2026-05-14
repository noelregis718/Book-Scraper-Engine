import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH = r"E:\Internship\PocketFM\awful agents.xlsx"

def format_excel(file_path):
    """Applies professional formatting to the Excel file."""
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        print(f"  [Error] Cannot format {file_path}: File is empty or missing.")
        return

    print(f"  [System] Applying professional formatting to {os.path.basename(file_path)}...")
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        
        # Safe save to a temporary file
        temp_path = file_path + ".formatted.xlsx"
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Jabberwocky Catalog')
        
        workbook = writer.book
        worksheet = writer.sheets['Jabberwocky Catalog']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply to headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        # Column widths and alignment
        # Columns: Name of Series, Author Name, Publisher, GoodReads link, Num Books, Rating, Ratings #, Synopsis, Is Romantasy, Sub-Genre, Agent
        column_widths = [35, 25, 15, 45, 12, 12, 12, 65, 15, 25, 15]
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            
        # Apply alignment to all cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # Text columns get left alignment, numeric/status get center
                if cell.column in [1, 2, 4, 8, 10]:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                cell.border = border

        worksheet.freeze_panes = "A2"
        writer.close()
        
        # Replace original with formatted version
        if os.path.exists(file_path):
            os.remove(file_path)
        os.rename(temp_path, file_path)
        print(f"  [System] Formatting complete! File is ready.")
        
    except Exception as e:
        print(f"  [Error] Formatting failed: {e}")

if __name__ == "__main__":
    format_excel(FILE_PATH)
