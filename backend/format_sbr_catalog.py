import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
CATALOG_FILE = "SBR_Media_Catalog_Final.xlsx"

def format_excel():
    if not os.path.exists(CATALOG_FILE):
        print(f"Error: {CATALOG_FILE} not found!")
        return

    print(f"Applying professional styling to {CATALOG_FILE}...")
    
    # Load workbook and worksheet
    wb = load_workbook(CATALOG_FILE)
    ws = wb.active
    
    # --- Styles ---
    # Header: Bold, White Text, Navy Blue Fill
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    
    # Zebra Striping: Removed (Set to Pure White)
    pure_white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Alignment: Center for ratings, Wrap for Synopsis
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    standard_align = Alignment(horizontal='left', vertical='center')
    
    # Border: Thin black line
    thin_side = Side(border_style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # --- Apply Header Styles ---
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # --- Apply Body Styles & Auto-Width ---
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = border
            cell.fill = pure_white_fill
            cell.font = Font(name='Calibri', size=11, color='000000') # Pure Black Text
            
            # Special alignment for Synopsis (Column 8)
            if cell.column == 8:
                cell.alignment = wrap_align
            # Special alignment for Ratings/Counts (Column 5, 6, 7)
            elif cell.column in [5, 6, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = standard_align

    # --- Adjust Column Widths ---
    # Column A: Name of Series (30)
    # Column B: Author Name (25)
    # Column C: Publisher (15)
    # Column D: GoodReads link (30)
    # Column E: Primary Books (15)
    # Column F: Rating (10)
    # Column G: Ratings Count (15)
    # Column H: Synopsis (60)
    # Column I: Romantasy? (15)
    # Column J: Sub-Genre (35)
    # Column K: Agent (15)
    
    widths = {1: 35, 2: 25, 3: 15, 4: 40, 5: 15, 6: 12, 7: 15, 8: 70, 9: 15, 10: 40, 11: 15}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze Top Row & Add Filter ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(CATALOG_FILE)
    print(f"Success! {CATALOG_FILE} is now professionally formatted.")
    if os.name == 'nt':
        os.startfile(CATALOG_FILE)

if __name__ == "__main__":
    format_excel()
