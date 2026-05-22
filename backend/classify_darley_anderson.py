import os
from openpyxl import load_workbook
import sys

KEYWORD_MAP = [
    ("Werewolf / Shifter Romance", ["werewolf", "shifter", "alpha", "pack", "omega", "luna", "wolf", "lycan"]),
    ("Monster Romance (Non-Shifter)", ["monster", "orc", "kraken", "alien", "beast", "creature", "demon lover", "tentacle"]),
    ("Mythology, Legend & Fairy Tale Retelling", ["retelling", "mythology", "greek god", "hades", "persephone", "fairy tale", "legend", "arthurian", "norse", "anansi", "medusa", "circe", "orpheus", "beauty and the beast"]),
    ("War College / Military Academy", ["war college", "military academy", "dragon rider", "fourth wing", "basgiath", "aerial", "flight school", "training camp", "rider", "bonded dragon", "academy"]),
    ("High-Stakes Games & Deadly Trials", ["trial", "deadly game", "tournament", "competition", "survival", "arena", "hunger game", "death match", "blood game", "lethal"]),
    ("Dark Academia Romantasy", ["dark academia", "secret society", "forbidden library", "ancient university", "campus", "scholarly", "cursed school", "arcane academy", "magic school"]),
    ("Gothic Dark Romantasy", ["gothic", "haunted", "manor", "dark romance", "gloomy castle", "shadow court", "cursed castle", "decaying estate", "vampire lord", "immortal lord"]),
    ("Korean Romance Fantasy / Isekai", ["isekai", "reincarnated", "villainess", "otome", "korean", "manhwa", "transmigrated", "possessed", "regression"]),
    ("Cozy / Cottagecore", ["cozy", "cottagecore", "small town magic", "bakery", "low stakes", "wholesome", "botanical", "village witch", "flower shop", "magical inn"]),
    ("Paranormal Romance", ["vampire", "ghost", "witch", "paranormal", "psychic", "medium", "warlock", "necromancer", "haunting", "supernatural romance", "fae romance", "fairy", "magic"]),
    ("High Fantasy Court Adventure", ["court", "throne", "kingdom", "royalty", "fae court", "high fantasy", "crown", "empire", "queen", "king", "prince", "realm", "dragon", "epic fantasy", "war", "political intrigue", "magic system", "sword"]),
    ("Urban / Contemporary Fantasy Romance", ["urban fantasy", "modern day", "contemporary fantasy", "hidden world", "secret magic", "real world", "city magic", "supernatural city"])
]

def classify_subgenre(text):
    text = str(text).lower()
    for subgenre, keywords in KEYWORD_MAP:
        if any(k in text for k in keywords):
            return subgenre
    return None

def update_classifications(file_path):
    print(f"Classifying {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    yes_count = 0
    no_count = 0
    
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        synopsis = ws.cell(row=row, column=8).value
        
        if not title or str(title).strip() == "":
            continue
            
        combined_text = str(title) + " " + str(synopsis)
        
        subgenre_result = classify_subgenre(combined_text)
        
        if subgenre_result is not None:
            ws.cell(row=row, column=9).value = "Yes"
            ws.cell(row=row, column=10).value = subgenre_result
            yes_count += 1
        else:
            ws.cell(row=row, column=9).value = "No"
            ws.cell(row=row, column=10).value = ""
            no_count += 1
            
    wb.save(file_path)
    print(f"Classification complete! Found {yes_count} Romantasy titles and {no_count} Non-Romantasy titles.")
    
    import subprocess
    subprocess.Popen(["start", file_path], shell=True)

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    target = os.path.join(base, "Darley_Anderson_Formatted.xlsx")
    update_classifications(target)
