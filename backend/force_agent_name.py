import openpyxl

excel_file = '../New_Agency_Template.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    agent_col = 11 # 'Name of agent in the main folder' is the 11th column (K)
    
    count = 0
    # Start at row 2 to skip the header
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=agent_col).value = "Sandra Bond"
        count += 1
        
    wb.save(excel_file)
    print(f"Successfully populated 'Sandra Bond' across all {count} data rows in Column K!")

except Exception as e:
    print(f"Error: {e}")
