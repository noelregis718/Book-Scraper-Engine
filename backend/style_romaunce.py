import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def style_romaunce(file_path):
    df = pd.read_excel(file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Romaunce Final"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    columns_count = len(df.columns)
    for col in range(1, columns_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, columns_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border

    widths = {
        "A": 30, "B": 22, "C": 25, "D": 35,
        "E": 15, "F": 15, "G": 15, "H": 50,
        "I": 15, "J": 25, "K": 25
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    print(f"Styling reapplied to {file_path}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    target = os.path.join(base, "Romaunce_Books_Complete.xlsx")
    style_romaunce(target)
