import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

excel_file = '../New_Agency_Template.xlsx'
text_file = 'body_text.txt'

# 1. Extract authors from the text file
with open(text_file, 'r', encoding='utf-8') as f:
    lines = f.readlines()

authors = []
capture = False
for line in lines:
    line = line.strip()
    if line == "Meet Our Authors":
        capture = True
        continue
    if line == "Privacy Policy":
        break
    if capture and line:
        authors.append(line)

print(f"Extracted {len(authors)} authors.")

# 2. Append authors to the styled Excel sheet using openpyxl
try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    author_col = 2  # Column B is Author Name
    agent_col = 11  # Column K is Name of agent
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D3D3D3'), 
                         right=Side(style='thin', color='D3D3D3'), 
                         top=Side(style='thin', color='D3D3D3'), 
                         bottom=Side(style='thin', color='D3D3D3'))

    for author in authors:
        # Construct an empty row of 11 cells
        row_data = [None] * 11
        row_data[author_col - 1] = author
        row_data[agent_col - 1] = "Confluence Literary Agency"
        
        ws.append(row_data)
        
        # Apply styles to the newly appended row
        new_row_idx = ws.max_row
        for cell in ws[new_row_idx]:
            cell.alignment = cell_alignment
            cell.border = thin_border
            
    wb.save(excel_file)
    print("Successfully populated the Excel sheet while preserving professional styling!")

except Exception as e:
    print(f"Error: {e}")
