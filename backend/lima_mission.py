import asyncio
import json
import os
import pandas as pd
import re
import sys
from playwright.async_api import async_playwright
from goodreads_scraper import GoodreadsScraper
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import unquote

if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# --- CONFIGURATION ---
BASE_URL = "https://www.limaagency.se/authors"
OUTPUT_FILE = r"E:\Internship\PocketFM\Lima Agency.xlsx"
STATE_FILE = r"E:\Internship\PocketFM\backend\lima_state_v3.json"
HEADLESS = False 

def format_excel(file_path):
    """Applies professional formatting to the Lima Agency Excel file."""
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        return

    try:
        temp_path = file_path.replace(".xlsx", "_temp.xlsx")
        df = pd.read_excel(file_path, engine='openpyxl')
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Lima Agency Catalog')
        
        workbook = writer.book
        worksheet = writer.sheets['Lima Agency Catalog']
        
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        column_widths = [30, 25, 15, 45, 15, 15, 15, 60, 15, 20, 15]
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left_align if cell.column in [1, 2, 4, 8] else center_align
                cell.border = border

        worksheet.freeze_panes = "A2"
        writer.close()
        
        if os.path.exists(file_path): os.remove(file_path)
        os.rename(temp_path, file_path)
    except Exception as e:
        print(f"  [Error] Formatting failed: {e}", flush=True)

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f:
            try: return json.load(f)
            except: return {"last_author_index": 0}
    return {"last_author_index": 0}

def save_state(index):
    with open(STATE_FILE, 'w') as f:
        json.dump({"last_author_index": index}, f)

async def get_author_list(page):
    """Scrapes author names from Lima Agency site."""
    print(f"[Lima Agency] Scanning author list: {BASE_URL}", flush=True)
    await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60000)
    await asyncio.sleep(5)
    await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    await asyncio.sleep(2)
    
    authors_data = await page.evaluate("""() => {
        const results = [];
        const seen = new Set();
        const allLinks = Array.from(document.querySelectorAll('a'));
        allLinks.forEach(a => {
            const href = a.href;
            const text = a.innerText.trim();
            if (href.includes('/authors/') && (href.includes('/fiction/') || href.includes('/nonfiction/') || href.includes('/seraf/'))) {
                results.push({name: text, url: href});
            }
        });
        return results;
    }""")
    
    unique_authors = []
    seen_urls = set()
    for auth in authors_data:
        href = auth['url']
        if href not in seen_urls:
            name = auth['name']
            if not name or len(name.split()) < 2:
                name = unquote(href.split('/')[-1]).replace('-', ' ').title()
            
            if name and len(name.split()) >= 2 and "Home" not in name and "Authors" not in name:
                unique_authors.append({"name": name, "url": href})
                seen_urls.add(href)
    
    print(f"[Lima Agency] Found {len(unique_authors)} unique authors.", flush=True)
    return unique_authors

async def process_author(author, context, gr_scraper, seen_titles, output_data_list):
    """Searches Goodreads for an author and scrapes their first 5 books."""
    author_name = author['name']
    try:
        print(f"  [Mission] Starting search for author: {author_name}", flush=True)
        # Search author and get first 5 books
        book_titles = await gr_scraper.search_author_books(context, author_name, max_books=5)
        
        if not book_titles:
            print(f"  [Skip] No books found on Goodreads for {author_name}", flush=True)
            return

        for title in book_titles:
            clean_title = re.sub(r'\(.*?\)', '', title).strip()
            if clean_title in seen_titles: continue
            
            print(f"    [Scraping] {clean_title} ({author_name})", flush=True)
            gr_data = await gr_scraper.scrape_goodreads_data(context, clean_title, author_name)
            
            if gr_data:
                # Fallback for author name
                final_author = gr_data.get('Author_Found', 'Unknown')
                if final_author == 'Unknown' or not final_author:
                    final_author = author_name
                    
                new_row = {
                    'Name of Series': gr_data.get('Book_Title', clean_title),
                    'Author Name': final_author,
                    'Publisher': 'Lima Agency',
                    'GoodReads series link': gr_data.get('GoodReads_Series_URL', gr_data.get('GoodReads_Book_URL', 'N/A')),
                    'Number of PRIMARY books in the series': gr_data.get('Num_Primary_Books', '1'),
                    'Rating (out of 5) of Primary Book 1': gr_data.get('Book1_Rating', 'N/A'),
                    'Ratings (#) of Primary Book 1': gr_data.get('Book1_Num_Ratings', 'N/A'),
                    'Synopsis (if available)': gr_data.get('Description', 'N/A'),
                    'Is it Romantasy ?': gr_data.get('Romantasy_Subgenre', 'No'),
                    'Romantasy Sub-Genre of series': gr_data.get('Genre', 'N/A'),
                    'Name of agent': 'Cecilia Lindh Mattsson'
                }
                output_data_list.append(new_row)
                seen_titles.add(clean_title)
                
    except Exception as e:
        print(f"  [Error] Failed processing author {author_name}: {e}", flush=True)

async def run_mission():
    print("Starting Lima Agency Goodreads-Direct Mission...", flush=True)
    
    if os.path.exists(OUTPUT_FILE):
        try: df = pd.read_excel(OUTPUT_FILE)
        except: df = pd.DataFrame()
    else:
        df = pd.DataFrame()

    seen_titles = set(df['Name of Series'].astype(str).tolist()) if not df.empty else set()
    gr_scraper = GoodreadsScraper(headless=HEADLESS)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context()
        page = await context.new_page()

        authors = await get_author_list(page)
        state = load_state()
        start_idx = state.get("last_author_index", 0)
        
        print(f"--- Mission Control: Starting from Author Index {start_idx} ---", flush=True)
        
        # Batch size for authors to process concurrently
        BATCH_SIZE = 5 
        for i in range(start_idx, len(authors), BATCH_SIZE):
            batch = authors[i:i+BATCH_SIZE]
            print(f"\n--- [Batch Start] Processing {len(batch)} Authors Concurrently ---", flush=True)
            
            output_data_list = []
            tasks = [process_author(author, context, gr_scraper, seen_titles, output_data_list) for author in batch]
            await asyncio.gather(*tasks)
            
            # Update DataFrame and save
            if output_data_list:
                new_df = pd.DataFrame(output_data_list)
                if df.empty: df = new_df
                else: df = pd.concat([df, new_df], ignore_index=True)
                
                df.to_excel(OUTPUT_FILE, index=False)
                format_excel(OUTPUT_FILE)
            
            save_state(i + BATCH_SIZE)
            print(f"--- [Batch Complete] Saved. Total records: {len(df)} ---", flush=True)
            
        print(f"Mission Complete. Saved to {OUTPUT_FILE}", flush=True)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(run_mission())
