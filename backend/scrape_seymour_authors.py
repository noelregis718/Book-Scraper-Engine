import asyncio
import os
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://theseymouragency.com/authors/"
OUTPUT_EXCEL = r"E:\Internship\PocketFM\Seymour_Agency_Authors.xlsx"

async def scrape_authors():
    print("============================================================")
    print("             SEYMOUR AGENCY AUTHOR SCRAPER                  ")
    print("============================================================")
    
    print(">>> Launching headed browser to bypass bot protection...")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = await context.new_page()
        
        print(f">>> Navigating to: {URL}")
        try:
            await page.goto(URL, wait_until="domcontentloaded", timeout=60000)
            print(">>> Waiting 10 seconds for authors to fully render...")
            await page.wait_for_timeout(10000)
            
            # Extract page content
            content = await page.content()
            
        except Exception as e:
            print(f"Error loading page: {e}")
            await browser.close()
            return []
            
        await browser.close()
        
    print(">>> Parsing HTML content...")
    soup = BeautifulSoup(content, "html.parser")
    author_cards = soup.find_all(class_="xixs-archive-author")
    print(f"Found {len(author_cards)} author cards in DOM.")
    
    authors_data = []
    for card in author_cards:
        # Extract author name
        name_tag = card.find(class_="xixs-archive-author-name")
        name = name_tag.text.strip() if name_tag else card.get("data-name", "").strip()
        
        # Extract Profile URL
        a_tag = card.find("a")
        profile_url = a_tag["href"].strip() if a_tag and "href" in a_tag.attrs else "N/A"
        
        # Extract Headshot Image URL
        img_tag = card.find("img")
        headshot_url = img_tag["src"].strip() if img_tag and "src" in img_tag.attrs else "N/A"
        
        # Extract Category
        cat_tag = card.find(class_="xixs-archive-author-category")
        category = cat_tag.text.strip() if cat_tag else "N/A"
        
        if name:
            authors_data.append({
                "Name": name,
                "Category": category,
                "Profile URL": profile_url,
                "Headshot URL": headshot_url
            })
            
    print(f"Successfully extracted {len(authors_data)} authors.")
    return authors_data

def format_excel_sheet(ws):
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Premium deep-teal branding colors
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name="Segoe UI", size=10)
    link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Header Styling
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = data_border
        
    # Data Rows Styling
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = data_border
            
            if col_idx in [3, 4] and cell.value and cell.value != "N/A":
                # Style links beautifully
                cell.font = link_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = data_font
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # Cap link string length for sizing estimation to avoid overly wide columns
            if len(val_str) > 40:
                val_str = val_str[:40]
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Apply native Excel auto-filters
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

def save_to_excel(authors_data):
    if not authors_data:
        print("Error: No data to save.")
        return
        
    print(f">>> Creating Excel file at: {OUTPUT_EXCEL}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seymour Agency Authors"
    
    # Write headers first
    headers = [
        "Author Name",
        "Category",
        "Profile URL",
        "Headshot Image URL"
    ]
    ws.append(headers)
    
    # Write data rows
    for row_data in authors_data:
        ws.append([
            row_data["Name"],
            row_data["Category"],
            row_data["Profile URL"],
            row_data["Headshot URL"]
        ])
        
    # Format and Style
    format_excel_sheet(ws)
    
    # Save Workbook
    wb.save(OUTPUT_EXCEL)
    print(">>> Excel catalog saved and formatted successfully!")

def main():
    authors_data = asyncio.run(scrape_authors())
    if authors_data:
        save_to_excel(authors_data)
        
        # Auto-open the completed sheet for preview
        try:
            os.startfile(OUTPUT_EXCEL)
            print(">>> Auto-opened sheet for review!")
        except Exception as e:
            print(f"Could not open file: {e}")
    else:
        print("Scraping failed or yielded no results.")

if __name__ == "__main__":
    main()
