import os
import sys
import re
from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from apply_jra_style import apply_styling

FILE_PATH = r"e:\Internship\PocketFM\LDLA_Combined.xlsx"

subgenres = {
    'High Fantasy Court Adventure': ['court', 'kingdom', 'throne', 'prince', 'princess', 'king', 'queen', 'crown', 'empire', 'realm', 'royal', 'palace'],
    'Gothic Dark Romantasy': ['gothic', 'dark', 'shadow', 'curse', 'blood', 'demon', 'death', 'macabre', 'haunted', 'sinister'],
    'Dark Academia Romantasy': ['academy', 'university', 'college', 'professor', 'student', 'dark academia', 'school', 'scholars'],
    'Monster Romance (Non-Shifter)': ['monster', 'creature', 'orc', 'goblin', 'alien', 'tentacle', 'demon', 'gargoyle'],
    'Werewolf / Shifter Romance': ['shifter', 'wolf', 'werewolf', 'pack', 'alpha', 'mate', 'omega', 'bear', 'lion', 'dragon shifter'],
    'High-Stakes Games & Deadly Trials': ['trial', 'game', 'tournament', 'survive', 'deadly', 'competition', 'arena', 'hunger games'],
    'Mythology, Legend & Fairy Tale Retelling': ['myth', 'legend', 'retelling', 'fairy tale', 'god', 'goddess', 'hades', 'persephone', 'olympus', 'greek', 'beauty and the beast'],
    'War College / Military Academy': ['war college', 'military academy', 'cadet', 'rider', 'dragon rider', 'conscription', 'squadron', 'rebellion'],
    'Korean Romance Fantasy / Isekai': ['isekai', 'reincarnated', 'villainess', 'manhwa', 'korean', 'duke', 'emperor', 'transmigrated'],
    'Paranormal Romance': ['vampire', 'ghost', 'paranormal', 'angel', 'witch', 'supernatural', 'immortal'],
    'Cozy / Cottagecore': ['cozy', 'cottage', 'tea', 'bakery', 'cafe', 'inn', 'tavern', 'low stakes', 'heartwarming', 'small town'],
    'Urban / Contemporary Fantasy Romance': ['urban', 'city', 'modern', 'detective', 'agency', 'hidden world', 'secret society', 'contemporary']
}

fantasy_keywords = ['magic', 'fae', 'dragon', 'vampire', 'shifter', 'witch', 'fantasy', 'kingdom', 'curse', 'demon', 'elf', 'realm', 'immortal', 'gods', 'myth', 'monster', 'supernatural', 'werewolf', 'faerie', 'spell', 'sword', 'sorcery', 'paranormal', 'portal', 'otherworldly', 'warlock', 'mage', 'beast']

def classify(synopsis, title):
    text = f"{str(synopsis)} {str(title)}".lower()
    
    # Check if it's romantasy
    is_romantasy = any(k in text for k in fantasy_keywords)
    
    if not is_romantasy:
        return 'No', 'N/A'
        
    best_genre = 'Urban / Contemporary Fantasy Romance' # default
    max_matches = 0
    
    for genre, kw_list in subgenres.items():
        matches = sum(1 for k in kw_list if re.search(r'\b' + k + r'\b', text))
        if matches > max_matches:
            max_matches = matches
            best_genre = genre
            
    # If it's romantasy but no specific keywords match, default
    if max_matches == 0:
        if any(k in text for k in ['court', 'kingdom', 'realm', 'sword', 'empire', 'throne']):
            best_genre = 'High Fantasy Court Adventure'
        else:
            best_genre = 'Paranormal Romance'
            
    return 'Yes', best_genre

def safely_classify_sheet():
    if not os.path.exists(FILE_PATH):
        print(f"Error: {FILE_PATH} not found!")
        return

    print(f"Safely opening {FILE_PATH} without overwriting data...")
    # Use openpyxl to modify cells strictly in place (NO pandas DataFrame overwrite)
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    # Find the column indices we need (1-based)
    header = {cell.value: cell.column for cell in ws[1]}
    
    synopsis_col = header.get('Synopsis (if available)')
    title_col = header.get('Name of Series')
    rom_yes_no_col = header.get('Romantasy = Yes or No?')
    rom_subgenre_col = header.get('Romantasy Sub-Genre of series')

    if not all([synopsis_col, title_col, rom_yes_no_col, rom_subgenre_col]):
        print("Error: Missing required columns in the sheet!")
        return

    print("Analyzing synopses and updating cells...")
    count = 0
    
    # Process row by row safely
    for row_idx in range(2, ws.max_row + 1):
        syn_val = ws.cell(row=row_idx, column=synopsis_col).value or ""
        title_val = ws.cell(row=row_idx, column=title_col).value or ""
        
        # Skip fully empty rows
        if not syn_val and not title_val:
            continue
            
        y_n, genre = classify(syn_val, title_val)
        
        # Write ONLY to the two targeted cells
        ws.cell(row=row_idx, column=rom_yes_no_col).value = y_n
        ws.cell(row=row_idx, column=rom_subgenre_col).value = genre
        
        if y_n == 'Yes':
            count += 1

    print("Surgically saving exact cell modifications...")
    wb.save(FILE_PATH)
    
    print("Re-applying JRA styling safely...")
    apply_styling(FILE_PATH)
    
    print(f"Classification Complete! Identified {count} books as Romantasy.")

if __name__ == '__main__':
    safely_classify_sheet()
