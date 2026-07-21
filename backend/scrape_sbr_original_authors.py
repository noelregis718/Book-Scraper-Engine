import asyncio
import os
import sys
import pandas as pd
from playwright.async_api import async_playwright
import traceback
import re

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from goodreads_scraper import GoodreadsScraper

EXCEL_FILE = r"E:\Internship\PocketFM\SBR Media.xlsx"
MAX_CONCURRENT = 8

def map_genres(tags):
    tags = [str(t).lower() for t in tags if str(t)]
    is_romantasy = any('romantasy' in t for t in tags)
    is_fantasy = any('fantasy' in t for t in tags)
    is_romance = any('romance' in t for t in tags)
    is_crime = any('crime' in t or 'thriller' in t or 'mystery' in t for t in tags)
    
    if is_romantasy: return 'Romantasy'
    elif is_crime: return 'Crime Thriller'
    elif is_romance and is_fantasy: return 'Romantasy'
    elif is_romance: return 'Romance Drama'
    elif is_fantasy: return 'Fantasy'
    return 'Unknown'

async def get_top_2_new_books(context, author_name, existing_titles, existing_links):
    page = await context.new_page()
    top_books = []
    try:
        search_url = f"https://www.goodreads.com/search?q={author_name.replace(' ', '+')}"
        await page.set_extra_http_headers({"User-Agent": "Mozilla/5.0"})
        await page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
        await asyncio.sleep(2.5)
        
        author_link = await page.query_selector('a[href*="/author/show/"]')
        if not author_link: author_link = await page.query_selector('.authorName, .authorName__container a')
        if not author_link: 
            await page.close()
            return []
            
        author_url = await author_link.evaluate("el => el.href")
        await page.goto(author_url, wait_until="domcontentloaded", timeout=45000)
        await asyncio.sleep(2.5)
        
        book_els = await page.query_selector_all('tr[itemtype="http://schema.org/Book"]')
        book_candidates = []
        
        for row in book_els[:20]:
            title_el = await row.query_selector('a.bookTitle')
            rating_el = await row.query_selector('.minirating')
            if title_el and rating_el:
                title = (await title_el.inner_text()).strip()
                link = await title_el.evaluate("el => el.href")
                r_text = await rating_el.inner_text()
                r_match = re.search(r'([\d.]+)\s*avg', r_text.lower())
                rating = float(r_match.group(1)) if r_match else 0.0
                
                # Check if we already have it
                clean_title = title.lower()
                clean_link = link.split('?')[0].lower()
                
                already_have = False
                for ex_t in existing_titles:
                    if ex_t in clean_title or clean_title in ex_t:
                        already_have = True
                        break
                for ex_l in existing_links:
                    if ex_l in clean_link:
                        already_have = True
                        break
                        
                if not already_have and link not in [b['link'] for b in book_candidates]:
                    book_candidates.append({'title': title, 'link': link, 'rating': rating})
        
        if book_candidates:
            book_candidates.sort(key=lambda x: x['rating'], reverse=True)
            top_books = book_candidates[:2]
    except Exception as e:
        print(f"    [Goodreads] Author scrape error: {e}")
    finally:
        await page.close()
        
    return top_books

async def process_author(context, scraper, idx, author_name, df, existing_titles, existing_links, semaphore, results_list):
    async with semaphore:
        print(f"[{idx}] Searching for 2 NEW Top books for Author: {author_name}...")
        try:
            top_books = await get_top_2_new_books(context, author_name, existing_titles, existing_links)
            
            if not top_books:
                print(f"[{idx}] No new books found for {author_name}.")
                return
                
            for b in top_books:
                print(f"[{idx}] Found new top book: {b['title']} (Rating: {b['rating']})")
                data = await scraper.scrape_goodreads_data(context, title=b['title'], author=author_name, existing_url=b['link'])
                if data:
                    new_row = {col: '' for col in df.columns}
                    new_row['Author Name'] = author_name
                    new_row['Name of Series'] = data.get('Book_Title', '')
                    new_row['GoodReads series link'] = data.get('GoodReads_Series_URL') or data.get('GoodReads_Book_URL', '')
                    new_row['Number of PRIMARY books in the series'] = data.get('Num_Primary_Books', 1)
                    new_row['Rating (out of 5) of Primary Book 1'] = data.get('Book1_Rating', '')
                    new_row['Ratings (#) of Primary Book 1'] = data.get('Book1_Num_Ratings', '')
                    new_row['Synopsis (if available)'] = data.get('Description', '')
                    new_row['No. of pages in Book 1'] = data.get('Num_Pages', '')
                    
                    try:
                        num_pages = int(data.get('Num_Pages', 0))
                        num_books = int(data.get('Num_Primary_Books', 1))
                        new_row['Page Count (Sum of no. of pages in all primary books)'] = num_pages * num_books if num_pages else ''
                    except:
                        pass
                    
                    all_genres = data.get('All_Genres', [])
                    new_row['Genre tags- Up to 7 tags'] = ", ".join(all_genres)
                    
                    genre = map_genres(all_genres)
                    new_row['Genre'] = genre
                    new_row['Sub-Genre'] = 'Needs Mapping'
                    
                    results_list.append((idx, new_row))
                
            print(f"[{idx}] Successfully added new books for '{author_name}'!")
        except Exception as e:
            print(f"[{idx}] Error scraping '{author_name}': {e}")

async def run_scrape():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found!")
        return

    df = pd.read_excel(EXCEL_FILE)
    
    # Isolate original authors (first 297 rows in the original file)
    original_authors = df.iloc[:297]['Author Name'].dropna().unique().tolist()
    
    scraper = GoodreadsScraper(headless=False)
    tasks = []
    results_list = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        login_page = await context.new_page()
        await scraper.login_to_goodreads(login_page)
        
        semaphore = asyncio.Semaphore(MAX_CONCURRENT)
        
        for idx, author_name in enumerate(original_authors):
            author_name = str(author_name).strip()
            if not author_name or author_name.lower() == 'nan':
                continue
                
            # Find existing books for this author to avoid duplicates
            author_rows = df[df['Author Name'].astype(str).str.strip() == author_name]
            existing_titles = set(author_rows['Name of Series'].dropna().astype(str).str.lower().str.strip().tolist())
            existing_links = set(author_rows['GoodReads series link'].dropna().astype(str).str.lower().str.strip().tolist())
            
            tasks.append(process_author(context, scraper, idx, author_name, df, list(existing_titles), list(existing_links), semaphore, results_list))
                
        if tasks:
            await asyncio.gather(*tasks)
            
        await login_page.close()
        await browser.close()
        
    print("Processing results...")
    
    # Append new rows
    new_rows = [r[1] for r in sorted(results_list, key=lambda x: x[0])]
    if new_rows:
        new_df = pd.DataFrame(new_rows)
        df = pd.concat([df, new_df], ignore_index=True)
        
    print("Saving Excel...")
    df.to_excel(EXCEL_FILE, index=False)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        align_left_nowrap = Alignment(horizontal='left', vertical='top', wrap_text=False)
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 18
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = align_left_nowrap
        wb.save(EXCEL_FILE)
        print("Styling applied successfully.")
    except Exception as e:
        print(f"Styling error: {e}")
    
    print("Done!")

if __name__ == '__main__':
    asyncio.run(run_scrape())
