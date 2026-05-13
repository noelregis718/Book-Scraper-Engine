import asyncio
import json
import os
import pandas as pd
import re
import sys
from playwright.async_api import async_playwright
from goodreads_scraper import GoodreadsScraper
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# --- CONFIGURATION ---
BASE_URL = "https://sbrmedia.com/authors/"
OUTPUT_FILE = r"E:\Internship\PocketFM\SBR Media.xlsx"
STATE_FILE = r"E:\Internship\PocketFM\backend\sbr_state.json"
HEADLESS = False 

def format_excel(file_path):
    """Applies professional formatting to the SBR Media Excel file."""
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        return

    try:
        temp_path = file_path.replace(".xlsx", "_temp.xlsx")
        df = pd.read_excel(file_path, engine='openpyxl')
        writer = pd.ExcelWriter(temp_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='SBR Media Catalog')
        
        workbook = writer.book
        worksheet = writer.sheets['SBR Media Catalog']
        
        # SBR Media Blue Style (Using a professional blue/teal)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        column_widths = [35, 25, 15, 45, 15, 15, 15, 65, 15, 25, 15]
        for i, width in enumerate(column_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = left_align if cell.column in [1, 2, 4, 8, 10] else center_align
                cell.border = border

        worksheet.freeze_panes = "A2"
        writer.close()
        
        if os.path.exists(file_path): os.remove(file_path)
        os.rename(temp_path, file_path)
    except Exception as e:
        print(f"  [Error] Formatting failed: {e}", flush=True)

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f:
            try: return json.load(f)
            except: return {"last_author_index": 0}
    return {"last_author_index": 0}

def save_state(index):
    with open(STATE_FILE, 'w') as f:
        json.dump({"last_author_index": index}, f)

async def get_all_authors(page):
    """Scrapes all author links across all pages of SBR Media."""
    print(f"[SBR Media] Scanning authors starting from: {BASE_URL}", flush=True)
    authors = []
    current_page = 1
    
    while True:
        url = BASE_URL if current_page == 1 else f"{BASE_URL}page/{current_page}/"
        print(f"  Scanning page {current_page}...", flush=True)
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(3)
            
            page_authors = await page.evaluate("""() => {
                const results = [];
                // Target author titles/links in the archive
                const links = Array.from(document.querySelectorAll('a'));
                links.forEach(a => {
                    const href = a.href;
                    const name = a.innerText.trim();
                    if (href.includes('/authors/') && href !== 'https://sbrmedia.com/authors/' && name.length > 2) {
                         results.push({name: name, url: href});
                    }
                });
                return results;
            }""")
            
            if not page_authors:
                break
                
            # Deduplicate
            for auth in page_authors:
                if auth['url'] not in [a['url'] for a in authors]:
                    authors.append(auth)
            
            # Check for next page
            next_exists = await page.query_selector(f'a[href*="/authors/page/{current_page + 1}/"]')
            if not next_exists:
                break
            current_page += 1
        except:
            break
            
    print(f"[SBR Media] Found {len(authors)} total authors.", flush=True)
    return authors

async def scrape_author_carousel(page, author_name):
    """Scrapes book titles from the carousel on the author's page."""
    print(f"  [Scraping] {author_name} carousel...", flush=True)
    await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    await asyncio.sleep(4)
    
    titles = await page.evaluate("""() => {
        const results = [];
        const images = document.querySelectorAll('img');
        images.forEach(img => {
            let alt = img.alt || '';
            // Clean common SBR fluff
            if (alt.length > 3 && 
                !alt.includes('SBR') && 
                !alt.includes('Agent') && 
                !alt.includes('literary') &&
                !alt.includes('Designed') &&
                !alt.includes('Version')) {
                
                // Remove author name if appended with - 
                let clean = alt.split(' - ')[0].split(' – ')[0].trim();
                if (clean.length > 3 && !clean.toLowerCase().startsWith('image_') && !/^[\d_\-]+$/.test(clean)) {
                    results.push(clean);
                }
            }
        });
        return [...new Set(results)];
    }""")
    
    print(f"    -> Extracted {len(titles)} potential titles.", flush=True)
    return titles

async def process_book(title, author_name, context, gr_scraper, seen_titles, output_data_list, semaphore):
    """Processes a single book with a semaphore to control concurrency."""
    async with semaphore:
        if title in seen_titles: return
        print(f"    [Goodreads] Searching for: {title} ({author_name})", flush=True)
        try:
            gr_data = await gr_scraper.scrape_goodreads_data(context, title, author_name)
            if gr_data:
                final_author = gr_data.get('Author_Found', 'Unknown')
                if final_author == 'Unknown' or not final_author:
                    final_author = author_name

                new_row = {
                    'Name of Series': gr_data.get('Book_Title', title),
                    'Author Name': final_author,
                    'Publisher': 'SBR Media',
                    'GoodReads series link': gr_data.get('GoodReads_Series_URL', gr_data.get('GoodReads_Book_URL', 'N/A')),
                    'Number of PRIMARY books in the series': gr_data.get('Num_Primary_Books', '1'),
                    'Rating (out of 5) of Primary Book 1': gr_data.get('Book1_Rating', 'N/A'),
                    'Ratings (#) of Primary Book 1': gr_data.get('Book1_Num_Ratings', 'N/A'),
                    'Synopsis (if available)': gr_data.get('Description', 'N/A'),
                    'Is it Romantasy ?': gr_data.get('Romantasy_Subgenre', 'No'),
                    'Romantasy Sub-Genre of series': gr_data.get('Genre', 'N/A'),
                    'Name of agent': 'SBR Media'
                }
                output_data_list.append(new_row)
                seen_titles.add(title)
        except Exception as e:
            print(f"    [Error] Goodreads failed for {title}: {e}", flush=True)

async def run_mission():
    print("Starting SBR Media High-Speed Mission...", flush=True)
    
    if os.path.exists(OUTPUT_FILE):
        try: df = pd.read_excel(OUTPUT_FILE)
        except: df = pd.DataFrame()
    else:
        df = pd.DataFrame()

    seen_titles = set(df['Name of Series'].astype(str).tolist()) if not df.empty else set()
    gr_scraper = GoodreadsScraper(headless=HEADLESS)
    
    # SEMAPHORE: Controls how many books to "open at a go"
    semaphore = asyncio.Semaphore(10)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context()
        page = await context.new_page()

        authors = await get_all_authors(page)
        state = load_state()
        start_idx = state.get("last_author_index", 0)
        
        print(f"--- Mission Control: Starting from Author Index {start_idx} ---", flush=True)
        
        BATCH_SIZE = 50 
        for i in range(start_idx, len(authors), BATCH_SIZE):
            batch = authors[i:i+BATCH_SIZE]
            print(f"\n--- [Batch Start] Processing Authors {i+1} to {min(i+BATCH_SIZE, len(authors))} ---", flush=True)
            
            # Step 1: Collect all titles from the batch of authors
            all_book_tasks = []
            output_data_list = []
            
            for author in batch:
                author_name = author['name']
                author_url = author['url']
                try:
                    await page.goto(author_url, wait_until="networkidle", timeout=60000)
                    book_titles = await scrape_author_carousel(page, author_name)
                    for title in book_titles:
                        all_book_tasks.append(process_book(title, author_name, context, gr_scraper, seen_titles, output_data_list, semaphore))
                except Exception as e:
                    print(f"  [Error] Failed navigating to {author_name}: {e}", flush=True)
            
            # Step 2: Process all books in parallel (Limited to 10 at a time by semaphore)
            if all_book_tasks:
                print(f"  [Mission] Opening {len(all_book_tasks)} books on Goodreads (10 at a go)...", flush=True)
                await asyncio.gather(*all_book_tasks)
            
            # Update DataFrame and save
            if output_data_list:
                new_df = pd.DataFrame(output_data_list)
                if df.empty: df = new_df
                else: df = pd.concat([df, new_df], ignore_index=True)
                
                df.to_excel(OUTPUT_FILE, index=False)
                format_excel(OUTPUT_FILE)
            
            save_state(i + BATCH_SIZE)
            print(f"--- [Batch Complete] Saved. Total records: {len(df)} ---", flush=True)

        print(f"Mission Complete. Saved to {OUTPUT_FILE}", flush=True)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(run_mission())
