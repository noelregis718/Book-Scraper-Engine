import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def apply_premium_fixed_style(file_path):
    print(f"Applying premium fixed styling to {file_path}...")
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define Styles
    # Premium Header: Dark Slate Blue with White Text
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    
    # Alternating Row Colors
    row_fill_even = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Alignment and Borders
    align_center_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # For body, we use top alignment and wrap text, but rely on fixed row heights to prevent blowing up
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )

    # 1. Apply Header Styles & Fixed Header Height
    ws.row_dimensions[1].height = 35
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center_header
        cell.border = thin_border
        
    # 2. Apply Body Styles & Fixed Row Heights
    for row in range(2, ws.max_row + 1):
        # Set fixed row height so they don't become massive
        ws.row_dimensions[row].height = 60
        
        # Use white for all rows
        current_fill = row_fill_odd
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = current_fill
            cell.border = thin_border
            
            # Center align numbers/links, left align text like Title and Synopsis
            header_val = str(ws.cell(row=1, column=col).value).lower()
            if any(x in header_val for x in ["rating", "number", "#", "link", "url", "length", "rank", "language"]):
                cell.alignment = align_top_center
            else:
                cell.alignment = align_top_left

    # 3. Set Fixed Column Widths
    # Build a map from header name to width
    width_map = {
        "title": 30,
        "url": 15,
        "author": 20,
        "rating": 12,
        "customer reviews": 15,
        "goodreads rating": 15,
        "goodreads no. of ratings": 18,
        "publisher": 20,
        "language": 12,
        "print length": 12,
        "best sellers rank": 18,
        "synopsis/summary": 50, # Give synopsis more width but not too much
        "series link": 15,
        "# of primary books": 18,
        "part of series": 14,
        "book number": 12,
        "goodreads link": 15
    }
    
    # Apply widths based on header name
    for col in range(1, ws.max_column + 1):
        header_val = str(ws.cell(row=1, column=col).value).lower().strip()
        col_letter = ws.cell(row=1, column=col).column_letter
        
        # Find matching width
        assigned_width = 15 # default
        for key, w in width_map.items():
            if key in header_val:
                assigned_width = w
                break
                
        ws.column_dimensions[col_letter].width = assigned_width

    # Save
    wb.save(file_path)
    print("Premium styling with fixed sizes applied successfully!")

if __name__ == "__main__":
    target_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Crime_Thriller_Template.xlsx")
    apply_premium_fixed_style(target_file)
