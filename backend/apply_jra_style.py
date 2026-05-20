import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def apply_styling(file_path):
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply Header Styles
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # Apply Body Styles
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align_top
            cell.border = thin_border
            
    # Set Column Widths
    # Assumes the 11-column structure
    widths = {
        "A": 30,  # Name of Series
        "B": 20,  # Author Name
        "C": 25,  # Publisher
        "D": 35,  # GoodReads series link
        "E": 15,  # Number of PRIMARY books
        "F": 15,  # Rating
        "G": 15,  # Ratings (#)
        "H": 50,  # Synopsis
        "I": 15,  # Romantasy
        "J": 20,  # Sub-Genre
        "K": 25   # Name of agent
    }
    
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save the updated workbook
    wb.save(file_path)
    print("Styling applied successfully.")

if __name__ == "__main__":
    apply_styling('JRA_Bestsellers_Complete.xlsx')
