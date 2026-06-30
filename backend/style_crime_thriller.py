import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

input_csv = r"E:\Internship\PocketFM\Crime_Thriller_Amazon_Base_List - Series Base - 800 Titles (2).csv"
output_xlsx = r"E:\Internship\PocketFM\Crime_Thriller_Amazon_Base_List_Styled.xlsx"

print(f"Reading {input_csv}...")
df = pd.read_csv(input_csv)

print(f"Saving to {output_xlsx}...")
df.to_excel(output_xlsx, index=False)

print("Applying styles...")
wb = load_workbook(output_xlsx)
ws = wb.active
ws.title = "Crime Thriller"

# Define styles
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green

amazon_ratings_col_idx = None

# Style headers
for col_idx, cell in enumerate(ws[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    if cell.value == "Amazon Ratings":
        amazon_ratings_col_idx = col_idx

# Freeze the top row
ws.freeze_panes = "A2"

# Adjust column widths and highlight Amazon Ratings
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
            
        # Highlight Amazon Ratings column (skip header)
        if amazon_ratings_col_idx and cell.column == amazon_ratings_col_idx and cell.row > 1:
            cell.fill = highlight_fill
            
    adjusted_width = min(max_length + 2, 50) # Cap width at 50 to avoid massive columns
    ws.column_dimensions[column].width = adjusted_width

wb.save(output_xlsx)
print(f"Successfully created styled Excel file: {output_xlsx}")
