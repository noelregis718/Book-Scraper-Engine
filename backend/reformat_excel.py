import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CONFIGURATION ---
EXCEL_FILE = r"E:\Internship\PocketFM\Knight Agency.xlsx"

def apply_premium_styling(file_path):
    if not os.path.exists(file_path):
        return

    print(f"\n>>> [Styling Mission] Applying Premium Aesthetics to {os.path.basename(file_path)}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        if ws is None:
            raise ValueError("Active worksheet is None")
        
        # --- STYLE TOKENS ---
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border_style = Side(style='thin', color="D9D9D9")
        row_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # --- 1. HEADER STYLING ---
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = row_border
        
        # --- 2. DATA ROW STYLING ---
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_even = row_idx % 2 == 0
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = row_border
                if is_even: cell.fill = even_row_fill
        
        # --- 3. COLUMN WIDTHS (Dynamic Adjustment) ---
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(int(col[0].column))
            # Set defaults based on known content types
            if column == 'H': # Synopsis
                ws.column_dimensions[column].width = 70
            elif column in ['A', 'D']: # Title/Link
                ws.column_dimensions[column].width = 40
            else:
                ws.column_dimensions[column].width = 20

        # --- 4. UX ENHANCEMENTS ---
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(file_path)
        print(f"  [OK] Styling complete for {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"  [Warning] Styling failed for {os.path.basename(file_path)}: {e}")

if __name__ == "__main__":
    # Scan root folder and backend folder for Excel files
    search_paths = [r"E:\Internship\PocketFM", r"E:\Internship\PocketFM\backend"]
    for path in search_paths:
        if not os.path.exists(path): continue
        for file in os.listdir(path):
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(path, file)
                apply_premium_styling(full_path)
    
    # Final confirmation
    print(f"\n{'='*60}")
    print("GLOBAL STYLING COMPLETE: All project files are now premium.")
    print(f"{'='*60}\n")
